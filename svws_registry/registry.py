# -*- coding: utf-8 -*-
"""
SVWS Design Registry — module FastAPI
=====================================
Chuẩn thiết kế "3D Design" + Sổ đăng ký thiết kế, lưu PostgreSQL,
lịch sử chỉnh sửa, AI soạn đề bài (gọi Anthropic phía server).

Cách gắn vào app hiện có (main.py):

    from svws_registry import registry

    app.include_router(registry.router)

    @app.on_event("startup")
    def _init_registry():
        registry.init_db()

Biến môi trường:
    DATABASE_URL       — có sẵn trên Render (postgres://... tự đổi thành postgresql://)
    ANTHROPIC_API_KEY  — key API Claude (bắt buộc để dùng nút AI)
    ANTHROPIC_MODEL    — mặc định "claude-sonnet-5"

URL sau khi deploy:  https://<app>.onrender.com/registry
"""
import os
import re
import copy
import json
import datetime
from pathlib import Path
from typing import Optional

import httpx
from fastapi import APIRouter, HTTPException, Body, Response, UploadFile, File
from fastapi.responses import HTMLResponse, StreamingResponse
from sqlalchemy import Column, JSON, Text
from sqlalchemy.orm.attributes import flag_modified
from sqlmodel import SQLModel, Field, Session, create_engine, select

# ----------------------------------------------------------------------------
# DB
# ----------------------------------------------------------------------------
_DB_URL = os.getenv("DATABASE_URL", "sqlite:///./svws_registry.db")
if _DB_URL.startswith("postgres://"):  # Render cấp dạng cũ
    _DB_URL = _DB_URL.replace("postgres://", "postgresql://", 1)

engine = create_engine(_DB_URL, pool_pre_ping=True)

_HERE = Path(__file__).resolve().parent
_SEED = _HERE / "registry_seed.json"
_PAGE = _HERE / "static" / "registry.html"
_THREE = _HERE / "static" / "three.min.js"   # Three.js r128 (MIT) — chuẩn Rev.E
_SVWS3D = _HERE / "static" / "svws3d.js"     # bộ dựng hình 3D chuẩn của SVWS
_SVWSPID = _HERE / "static" / "svwspid.js"   # bộ dựng sơ đồ P&ID chuẩn
_SVWSGA = _HERE / "static" / "svwsga.js"     # bộ dựng mặt bằng GA chuẩn
_SVWSDIEN = _HERE / "static" / "svwsdien.js" # bộ dựng logic vận hành + bản vẽ điện
_SVWSCHE = _HERE / "static" / "svwsche.js"   # bộ dựng bản vẽ chế tạo (SD)
_SVWSVT = _HERE / "static" / "svwsvt.js"     # vật tư & phân tách thi công theo hình học

MAX_REVISIONS = 300  # giữ tối đa bao nhiêu bản lịch sử

# Địa chỉ công khai của app — tool tải về gọi ngược lại đây để dùng AI mà không
# phải ôm API key. Đổi được qua env nếu sau này dời tên miền.
_PUBLIC_URL = os.getenv("SVWS_PUBLIC_URL", "https://svws-app-iit7.onrender.com").rstrip("/")


def _now() -> str:
    return datetime.datetime.now(datetime.timezone.utc).isoformat(timespec="microseconds")


class RegistryDoc(SQLModel, table=True):
    """Tài liệu chính — luôn chỉ có 1 dòng id=1 (nguồn chuẩn duy nhất)."""
    __tablename__ = "svws_registry_doc"
    id: int = Field(default=1, primary_key=True)
    data: dict = Field(default_factory=dict, sa_column=Column(JSON))
    updated_at: str = Field(default_factory=_now)
    updated_by: str = ""


class RegistryRevision(SQLModel, table=True):
    """Mỗi lần lưu tạo 1 bản snapshot để tra lịch sử / khôi phục."""
    __tablename__ = "svws_registry_revision"
    id: Optional[int] = Field(default=None, primary_key=True)
    data: dict = Field(default_factory=dict, sa_column=Column(JSON))
    updated_at: str = Field(default_factory=_now)
    updated_by: str = ""
    note: str = Field(default="", sa_column=Column(Text))


_MUC_TAB7 = "Tab 7 — Logic vận hành & Sơ đồ mạch điện"
_MUC_TAB4 = "Bản vẽ chế tạo phải ĐỦ ĐỂ XƯỞNG LÀM ĐƯỢC"
_MUC_TAB5 = "Số lượng vật tư phải SUY TỪ HÌNH HỌC THẬT"
_MUC_TAB5B = "Bản vẽ từng công đoạn phải vẽ ĐÚNG TUYẾN THẬT"


def _seed_items(tien_to) -> list:
    """Lấy các mục trong seed bắt đầu bằng `tien_to` (chuỗi hoặc danh sách chuỗi).

    Giữ nguyên THỨ TỰ trong file seed để chèn vào tài liệu không bị đảo.
    """
    if not _SEED.exists():
        return []
    dau = (tien_to,) if isinstance(tien_to, str) else tuple(tien_to)
    ra = []
    for sec in json.loads(_SEED.read_text(encoding="utf-8")).get("sections") or []:
        for it in sec.get("items") or []:
            if (it.get("t") or "").startswith(dau):
                ra.append(it)
    return ra


def _bo_sung_muc(data: dict, dau_hieu: str, tien_to, nhom: str = "") -> bool:
    """Chèn các mục seed có tiền tố `tien_to` vào ngay sau cụm tab tương ứng.

    Idempotent: nhận diện bằng câu đặc trưng `dau_hieu` đã có trong tài liệu.
    Dùng cho mọi lần siết chuẩn về sau, không phải viết lại hàm mới mỗi lần.
    """
    if not nhom:                                       # "Tab 4 (bổ sung)" → "4"
        nhom = (tien_to if isinstance(tien_to, str) else tien_to[0]).split(" ")[1]
    for sec in data.get("sections") or []:
        if "tab bắt buộc" not in (sec.get("title") or "").lower():
            continue
        items = sec.get("items") or []
        if any(dau_hieu in (it.get("t") or "") for it in items):
            return False
        them = _seed_items(tien_to)
        if not them:
            return False
        vt = 0
        for i, it in enumerate(items):
            if (it.get("t") or "").startswith("Tab " + nhom):
                vt = i + 1
        sec["items"] = items[:vt] + them + items[vt:]
        return True
    return False


def _nang_cap_9_tab(data: dict) -> bool:
    """Bổ sung mục Tab 7 (Logic vận hành & mạch điện) vào tài liệu ĐANG CHẠY.

    Seed chỉ nạp khi DB trống, nên bản đã chạy trên Render sẽ không tự có mục
    mới. Hàm này chèn thêm, và đánh lại số hai tab cuối (O&M → 8, BOQ → 9).
    Chạy lại nhiều lần cũng chỉ thêm một lần; không đụng nội dung người dùng
    đã tự sửa. Trả về True nếu có thay đổi.
    """
    if not _SEED.exists():
        return False
    for sec in data.get("sections") or []:
        if "tab bắt buộc" not in (sec.get("title") or "").lower():
            continue
        items = sec.get("items") or []
        if any(_MUC_TAB7 in (it.get("t") or "") for it in items):
            return False                       # đã nâng cấp rồi

        seed_items = []
        for s_sec in json.loads(_SEED.read_text(encoding="utf-8")).get("sections") or []:
            for it in s_sec.get("items") or []:
                if (it.get("t") or "").startswith("Tab 7"):
                    seed_items.append(it)
        if not seed_items:
            return False

        # Đánh số lại TRƯỚC khi chèn, nếu không mục mới cũng bị đánh số theo.
        for it in items:
            t = it.get("t") or ""
            if t.startswith("Tab 8 —"):
                it["t"] = "Tab 9 —" + t[len("Tab 8 —"):]
            elif t.startswith("Tab 7 —"):
                it["t"] = "Tab 8 —" + t[len("Tab 7 —"):]
            elif t.startswith("Tab 7 ("):
                it["t"] = "Tab 8 (" + t[len("Tab 7 ("):]
            elif t.startswith("Tab 8 ("):
                it["t"] = "Tab 9 (" + t[len("Tab 8 ("):]

        vt = 0                                 # chèn ngay sau cụm Tab 6
        for i, it in enumerate(items):
            if (it.get("t") or "").startswith("Tab 6"):
                vt = i + 1
        sec["items"] = items[:vt] + seed_items + items[vt:]
        sec["title"] = "Cấu trúc 9 tab bắt buộc"
        return True
    return False


def init_db() -> None:
    """Tạo bảng nếu chưa có; nạp seed nếu DB trống; nâng cấp chuẩn nếu đã có."""
    SQLModel.metadata.create_all(engine)
    with Session(engine) as s:
        doc = s.get(RegistryDoc, 1)
        if doc is None:
            seed = json.loads(_SEED.read_text(encoding="utf-8")) if _SEED.exists() else {
                "meta": {"code": "SVWS-3D-REQ-001", "rev": "Rev.E", "date": "", "by": ""},
                "sections": [], "designs": [],
            }
            doc = RegistryDoc(id=1, data=seed, updated_at=_now(), updated_by="seed Rev.E")
            s.add(doc)
            s.add(RegistryRevision(data=seed, updated_at=doc.updated_at,
                                   updated_by="seed Rev.E", note="Khởi tạo từ file Rev.E"))
            s.commit()
            return

        # PHẢI sao chép SÂU: sao chép nông dùng chung các đối tượng con, sửa vào
        # đó là sửa luôn giá trị cũ → SQLAlchemy thấy "mới bằng cũ" và bỏ qua
        # cột JSON, kết quả là updated_by đổi mà nội dung thì không.
        data = copy.deepcopy(doc.data or {})
        ghi = []
        if _nang_cap_9_tab(data):
            ghi.append("Tab 7 — Logic vận hành & Sơ đồ mạch điện "
                       "(chuẩn chuyển từ 8 lên 9 tab)")
        if _bo_sung_muc(data, _MUC_TAB4, "Tab 4 (bổ sung)"):
            ghi.append("Siết chuẩn Tab 4 — bản vẽ chế tạo phải đủ để xưởng làm được")
        if _bo_sung_muc(data, _MUC_TAB5,
                        ["Tab 5 (bổ sung) — Số lượng vật tư",
                         "Tab 5 (bổ sung) — Danh sách công đoạn"], "5"):
            ghi.append("Siết chuẩn Tab 5 — vật tư suy từ hình học thật của bản vẽ 3D")
        if _bo_sung_muc(data, _MUC_TAB5B, "Tab 5 (bổ sung) — Bản vẽ từng công đoạn", "5"):
            ghi.append("Siết chuẩn Tab 5 — bản vẽ công đoạn vẽ đúng tuyến thật "
                       "kèm cao trình giá đỡ")
        if ghi:
            doc.data = data
            flag_modified(doc, "data")
            doc.updated_at = _now()
            doc.updated_by = "nâng cấp tự động"
            s.add(doc)
            s.add(RegistryRevision(data=data, updated_at=doc.updated_at,
                                   updated_by="nâng cấp tự động",
                                   note=" · ".join(ghi)))
            s.commit()
            for t in ghi:
                print("[REGISTRY] Đã bổ sung: " + t)


# ----------------------------------------------------------------------------
# Router
# ----------------------------------------------------------------------------
router = APIRouter(prefix="/registry", tags=["SVWS Design Registry"])


@router.get("", response_class=HTMLResponse, include_in_schema=False)
@router.get("/", response_class=HTMLResponse, include_in_schema=False)
def page() -> str:
    """Giao diện app (single-file HTML)."""
    if not _PAGE.exists():
        raise HTTPException(500, "Thiếu file static/registry.html trong module")
    return _PAGE.read_text(encoding="utf-8")


@router.get("/three.js", include_in_schema=False)
def three_js():
    """Three.js r128 để nhúng vào tool sinh ra.

    AI không thể gõ tay 600KB thư viện (vượt xa trần token) nên trước đây nó tự
    hạ cấp xuống canvas 2D. Nay AI chỉ viết phần cảnh, chỗ thư viện để trống
    bằng dấu hiệu, client nạp file này chèn vào → tool vẫn single-file offline.
    """
    if not _THREE.exists():
        raise HTTPException(500, "Thiếu static/three.min.js trong module")
    return Response(_THREE.read_text(encoding="utf-8"),
                    media_type="application/javascript; charset=utf-8",
                    headers={"Cache-Control": "public, max-age=31536000, immutable"})


@router.get("/svws3d.js", include_in_schema=False)
def svws3d_js():
    """Bộ dựng hình chuẩn: bố cục theo hàng, ống vuông góc, camera tự căn khung.

    AI gõ toạ độ mà không nhìn thấy kết quả nên bố cục hay hỏng. Thư viện này
    nhận phần đó; AI chỉ khai báo danh sách thiết bị và đường nối.
    """
    if not _SVWS3D.exists():
        raise HTTPException(500, "Thiếu static/svws3d.js trong module")
    return Response(_SVWS3D.read_text(encoding="utf-8"),
                    media_type="application/javascript; charset=utf-8",
                    headers={"Cache-Control": "no-cache"})


@router.get("/svwspid.js", include_in_schema=False)
def svwspid_js():
    """Bộ dựng P&ID: đo ký hiệu và đo chữ, tự né chồng lấn, nối ống vuông góc.

    AI gõ toạ độ SVG mà không nhìn thấy nên tự đặt hai con số mâu thuẫn (ký hiệu
    cao 38px mà khoảng cách xếp để 30px → chồng thành vệt đen). Thư viện nhận
    phần bố cục; AI chỉ khai báo dây chuyền.
    """
    if not _SVWSPID.exists():
        raise HTTPException(500, "Thiếu static/svwspid.js trong module")
    return Response(_SVWSPID.read_text(encoding="utf-8"),
                    media_type="application/javascript; charset=utf-8",
                    headers={"Cache-Control": "no-cache"})


@router.get("/svwsga.js", include_in_schema=False)
def svwsga_js():
    """Bộ dựng mặt bằng GA — dùng CHUNG bố cục với bản 3D nên hai bản vẽ khớp nhau.

    Ngoài việc vẽ đúng tỷ lệ và né chồng nhãn, nó còn kiểm KHE VẬN HÀNH giữa các
    thiết bị — đây là lỗi thiết kế thật, không phải lỗi vẽ.
    """
    if not _SVWSGA.exists():
        raise HTTPException(500, "Thiếu static/svwsga.js trong module")
    return Response(_SVWSGA.read_text(encoding="utf-8"),
                    media_type="application/javascript; charset=utf-8",
                    headers={"Cache-Control": "no-cache"})


@router.get("/svwsdien.js", include_in_schema=False)
def svwsdien_js():
    """Bộ dựng LOGIC VẬN HÀNH + BẢN VẼ ĐIỆN — đủ để thi công tủ và lập trình PLC.

    Ngoài phần vẽ (mạch động lực, thang điều khiển, đấu nối PLC, bố trí tủ), nó
    còn TỰ CHỌN THIẾT BỊ theo công suất: dòng định mức, MCCB, khởi động từ,
    rơ-le nhiệt, tiết diện cáp — và tự kiểm những lỗi chết người như thiếu nút
    dừng khẩn cấp, trùng địa chỉ PLC, thiết bị không đủ chỗ trong tủ.
    """
    if not _SVWSDIEN.exists():
        raise HTTPException(500, "Thiếu static/svwsdien.js trong module")
    return Response(_SVWSDIEN.read_text(encoding="utf-8"),
                    media_type="application/javascript; charset=utf-8",
                    headers={"Cache-Control": "no-cache"})


@router.get("/svwsche.js", include_in_schema=False)
def svwsche_js():
    """Bộ dựng BẢN VẼ CHẾ TẠO — đủ để xưởng cắt tôn, khoét lỗ và hàn.

    Đo được trong tool đã sinh: "3 hình chiếu" là ba ô chữ nhật rỗng ở toạ độ cố
    định, giống hệt nhau cho mọi thiết bị, không tỷ lệ và không nozzle. Thư viện
    này dựng hình theo kích thước thật, tự chọn tỷ lệ, tự ghi chuỗi kích thước,
    và tính những thứ xưởng cần: bề dày yêu cầu theo cột nước/áp suất, khai
    triển tôn, trọng lượng, bảng mối hàn.
    """
    if not _SVWSCHE.exists():
        raise HTTPException(500, "Thiếu static/svwsche.js trong module")
    return Response(_SVWSCHE.read_text(encoding="utf-8"),
                    media_type="application/javascript; charset=utf-8",
                    headers={"Cache-Control": "no-cache"})


@router.get("/svwsvt.js", include_in_schema=False)
def svwsvt_js():
    """Vật tư & phân tách thi công lấy số từ HÌNH HỌC THẬT của bản vẽ 3D.

    Trong tool đã sinh, mọi công đoạn đều ra 6 m ống và 2 co 90° vì đó là hằng
    số gõ cứng — đoạn dài 2 m và đoạn dài 30 m ra cùng một số, và BOQ lấy giá
    từ đó nên sai theo. Thư viện này đọc chiều dài thật, số khúc gãy thật và vị
    trí thật của thiết bị để ra số lượng, kèm cột "căn cứ tính" cho từng dòng.
    """
    if not _SVWSVT.exists():
        raise HTTPException(500, "Thiếu static/svwsvt.js trong module")
    return Response(_SVWSVT.read_text(encoding="utf-8"),
                    media_type="application/javascript; charset=utf-8",
                    headers={"Cache-Control": "no-cache"})


@router.get("/api")
def get_doc():
    """Lấy toàn bộ tài liệu (chuẩn + sổ thiết kế)."""
    with Session(engine) as s:
        doc = s.get(RegistryDoc, 1)
        if doc is None:
            raise HTTPException(404, "Chưa khởi tạo — gọi init_db() ở startup")
        return {"data": doc.data, "updated_at": doc.updated_at, "updated_by": doc.updated_by}


@router.put("/api")
def save_doc(payload: dict = Body(...)):
    """Lưu toàn bộ tài liệu.

    Body: {data: {...}, updated_by: "Tên", base_updated_at: "..."}
    base_updated_at dùng chống ghi đè: nếu người khác đã lưu sau thời điểm
    mình tải trang thì trả 409 kèm bản mới nhất.
    """
    data = payload.get("data")
    if not isinstance(data, dict) or "sections" not in data:
        raise HTTPException(422, 'Body thiếu "data.sections"')
    who = (payload.get("updated_by") or "").strip()[:120] or "không rõ"
    base = payload.get("base_updated_at") or ""

    with Session(engine) as s:
        doc = s.get(RegistryDoc, 1)
        if doc is None:
            raise HTTPException(404, "Chưa khởi tạo")
        if base and base != doc.updated_at:
            return {"conflict": True, "data": doc.data,
                    "updated_at": doc.updated_at, "updated_by": doc.updated_by}
        doc.data = data
        doc.updated_at = _now()
        doc.updated_by = who
        s.add(doc)
        s.add(RegistryRevision(data=data, updated_at=doc.updated_at,
                               updated_by=who, note=payload.get("note", "")[:300]))
        # cắt bớt lịch sử cũ
        ids = s.exec(select(RegistryRevision.id)
                     .order_by(RegistryRevision.id.desc())).all()
        for old_id in ids[MAX_REVISIONS:]:
            old = s.get(RegistryRevision, old_id)
            if old:
                s.delete(old)
        s.commit()
        return {"conflict": False, "updated_at": doc.updated_at, "updated_by": who}


@router.get("/api/history")
def history(limit: int = 40):
    """Danh sách các lần lưu gần nhất (không kèm nội dung)."""
    with Session(engine) as s:
        rows = s.exec(select(RegistryRevision)
                      .order_by(RegistryRevision.id.desc())
                      .limit(max(1, min(limit, 200)))).all()
        return [{"id": r.id, "updated_at": r.updated_at,
                 "updated_by": r.updated_by, "note": r.note,
                 "n_designs": len((r.data or {}).get("designs", [])),
                 "n_items": sum(len(sec.get("items", []))
                                for sec in (r.data or {}).get("sections", []))}
                for r in rows]


@router.get("/api/history/{rid}")
def history_item(rid: int):
    with Session(engine) as s:
        r = s.get(RegistryRevision, rid)
        if r is None:
            raise HTTPException(404, "Không có bản lịch sử này")
        return {"id": r.id, "data": r.data, "updated_at": r.updated_at,
                "updated_by": r.updated_by, "note": r.note}


@router.post("/api/restore/{rid}")
def restore(rid: int, payload: dict = Body(default={})):
    """Khôi phục về một bản lịch sử (tự tạo thêm 1 revision mới ghi nhận việc này)."""
    who = (payload.get("updated_by") or "").strip()[:120] or "không rõ"
    with Session(engine) as s:
        r = s.get(RegistryRevision, rid)
        if r is None:
            raise HTTPException(404, "Không có bản lịch sử này")
        doc = s.get(RegistryDoc, 1)
        doc.data = r.data
        doc.updated_at = _now()
        doc.updated_by = who
        s.add(doc)
        s.add(RegistryRevision(data=r.data, updated_at=doc.updated_at, updated_by=who,
                               note=f"Khôi phục từ bản #{rid}"))
        s.commit()
        return {"ok": True, "updated_at": doc.updated_at}


# ----------------------------------------------------------------------------
# AI — gọi Anthropic phía server (key không lộ ra trình duyệt)
# ----------------------------------------------------------------------------
_ANTHROPIC_URL = "https://api.anthropic.com/v1/messages"


def _loi_ai_vn(status: int, body: str) -> str:
    """Dịch lỗi của Anthropic sang tiếng Việt kèm CHỖ SỬA.

    Nguyên văn tiếng Anh dán lên màn hình thì người dùng không biết phải làm gì
    — mấy lỗi hay gặp đều là chuyện tài khoản, sửa ở Console chứ không phải lỗi app.
    """
    b = (body or "")
    low = b.lower()
    khi_nao = ""
    m = re.search(r"regain access on ([0-9]{4}-[0-9]{2}-[0-9]{2}) at ([0-9]{2}:[0-9]{2})", b)
    if m:
        try:  # giờ UTC → giờ Việt Nam cho dễ hình dung
            t = datetime.datetime.fromisoformat(m.group(1) + "T" + m.group(2) + ":00+00:00")
            vn = t.astimezone(datetime.timezone(datetime.timedelta(hours=7)))
            khi_nao = f" Hạn mức tự mở lại lúc {vn:%H:%M ngày %d/%m/%Y} (giờ Việt Nam)."
        except ValueError:
            khi_nao = f" Mở lại lúc {m.group(2)} UTC ngày {m.group(1)}."

    if "usage limit" in low or "spend limit" in low:
        return ("Tài khoản Anthropic đã CHẠM HẠN MỨC CHI TIÊU do chính mình đặt "
                "(không phải hết tiền, cũng không phải lỗi app)." + khi_nao +
                " Muốn dùng ngay thì vào console.anthropic.com → Settings → Limits "
                "để nâng hạn mức. Trong lúc chờ, nút '📄 Đề bài nhanh (không AI)' và "
                "'📋 Copy đề bài' vẫn dùng bình thường.")
    if "credit balance" in low:
        return ("Tài khoản Anthropic đã HẾT CREDIT. Vào console.anthropic.com → "
                "Plans & Billing để nạp thêm. Trong lúc chờ, nút '📄 Đề bài nhanh "
                "(không AI)' vẫn dùng được.")
    if status == 429 or "rate_limit" in low:
        return "Gọi AI quá nhanh (quá giới hạn số lượt/phút). Chờ khoảng một phút rồi thử lại."
    if status == 401 or "authentication" in low or "invalid x-api-key" in low:
        return ("Khóa ANTHROPIC_API_KEY sai hoặc đã bị thu hồi. Kiểm tra lại trong "
                "Render → Environment.")
    if "model" in low and ("not found" in low or "does not exist" in low):
        return ("Mã model trong ANTHROPIC_MODEL không còn phục vụ — đổi sang model "
                "hiện hành trong Render → Environment.")
    return f"Anthropic API lỗi {status}: {b[:300]}"


def _goi_ai_json(api_key: str, model: str, content, max_tokens: int,
                 system: str = "", timeout: float = 300.0, tra_tho: bool = False):
    """Gọi Claude và ép câu trả lời về JSON.

    BẪY: token SUY NGHĨ (adaptive thinking) TÍNH VÀO max_tokens. Để trần thấp thì
    model nghĩ hết hạn mức rồi JSON bị cắt giữa chừng, parse hỏng — và thông báo
    lỗi lại đổ oan cho "file khó đọc". Vì vậy: trần rộng, hãm mức suy nghĩ, và
    khi chạm trần thì nói đúng bệnh.

    tra_tho=True: parse hỏng thì trả {"__tho__": <văn bản>} thay vì ném lỗi.
    """
    body = {"model": model, "max_tokens": max_tokens,
            "messages": [{"role": "user", "content": content}]}
    if system:
        body["system"] = system
    effort = os.getenv("ANTHROPIC_EFFORT", "medium")
    if effort:
        body["output_config"] = {"effort": effort}
    try:
        resp = httpx.post(_ANTHROPIC_URL,
                          headers={"x-api-key": api_key,
                                   "anthropic-version": "2023-06-01",
                                   "content-type": "application/json"},
                          json=body, timeout=timeout)
    except httpx.HTTPError as e:
        raise HTTPException(502, f"Không gọi được Anthropic API: {e}")
    if resp.status_code != 200:
        raise HTTPException(502, _loi_ai_vn(resp.status_code, resp.text))

    out = resp.json()
    text = "\n".join(b.get("text", "") for b in out.get("content", [])
                     if b.get("type") == "text")
    stop = out.get("stop_reason")
    used = (out.get("usage") or {}).get("output_tokens", "?")
    clean = text.replace("```json", "").replace("```", "").strip()
    try:
        return json.loads(clean)
    except json.JSONDecodeError:
        pass
    m = re.search(r"[\{\[][\s\S]*[\}\]]", clean)   # vớt khối JSON lẫn trong lời dẫn
    if m:
        try:
            return json.loads(m.group(0))
        except json.JSONDecodeError:
            pass
    if tra_tho:
        return {"__tho__": clean}
    if stop == "max_tokens":
        raise HTTPException(502, f"AI bị cắt giữa chừng vì chạm trần {max_tokens:,} token "
                                 f"(đã dùng {used}) nên câu trả lời không đọc được. "
                                 f"Rút gọn file rồi thử lại, hoặc tăng trần token.")
    raise HTTPException(502, f"AI trả lời sai định dạng (stop={stop}, {used} token). "
                             f"Nội dung nhận được: {clean[:200] or '(rỗng)'}")


def _sections_summary(data: dict) -> str:
    lines = []
    for si, sec in enumerate(data.get("sections", []), 1):
        items = " | ".join(it.get("t", "") for it in sec.get("items", []))
        lines.append(f"{si}. {sec.get('title', '')}: {items}")
    return "\n".join(lines)


def _design_text(d: dict) -> str:
    def block(v):
        return "\n".join("    " + x for x in (v or "—").split("\n"))
    return "\n".join([
        f"• Mã thiết kế: {d.get('code') or '—'}",
        f"• Tên thiết kế: {d.get('name') or '—'}",
        f"• Công nghệ: {d.get('tech') or '—'}",
        f"• Công suất: {d.get('capacity') or '—'}",
        f"• Dự án / Khách hàng: {d.get('project') or '—'}",
        "• Thông số chính:", block(d.get("params")),
        "• Yêu cầu riêng:", block(d.get("notes")),
    ])


@router.post("/api/ai")
def ai_analyze(payload: dict = Body(...)):
    """AI đọc phiếu thiết kế, chỉ ra thông tin thiếu và chuẩn hóa mô tả.

    Body: {design: {...}, extra_answers: "..."}
    Trả về: {questions: [...], refined_A: "...", notes: "..."}
    """
    api_key = os.getenv("ANTHROPIC_API_KEY", "")
    if not api_key:
        raise HTTPException(503, "Server chưa cấu hình ANTHROPIC_API_KEY "
                                 "(Render → Environment → thêm biến này)")
    model = os.getenv("ANTHROPIC_MODEL", "claude-sonnet-5")

    design = payload.get("design") or {}
    extra = (payload.get("extra_answers") or "").strip()
    extra_block = ("\nTHÔNG TIN BỔ SUNG NGƯỜI DÙNG VỪA TRẢ LỜI:\n" + extra) if extra else ""

    with Session(engine) as s:
        doc = s.get(RegistryDoc, 1)
        summary = _sections_summary(doc.data if doc else {})

    user_msg = f"""Bạn là kỹ sư trưởng của SVWS (công ty EPC xử lý nước / nước thải / khí thải, TP.HCM). Dưới đây là PHIẾU MÔ TẢ một thiết kế và TÓM TẮT bộ chuẩn "3D Design" của công ty.

PHIẾU THIẾT KẾ:
{_design_text(design)}
{extra_block}

TÓM TẮT CHUẨN 3D DESIGN:
{summary}

NHIỆM VỤ: (1) liệt kê tối đa 6 câu hỏi về thông tin CÒN THIẾU quan trọng nhất để thiết kế được (nếu đủ thì để mảng rỗng); (2) viết "refined_A": mô tả thiết kế chuẩn hóa, gọn, đủ, kỹ thuật, tiếng Việt, dạng gạch đầu dòng, gộp cả thông tin bổ sung nếu có; (3) "notes": 1-2 lưu ý kỹ thuật đáng chú ý (nếu có).

CHỈ trả về JSON hợp lệ, không markdown, không giải thích:
{{"questions":["..."],"refined_A":"...","notes":"..."}}"""

    parsed = _goi_ai_json(api_key, model, user_msg, 8000, timeout=180.0, tra_tho=True)
    if "__tho__" in parsed:
        # AI trả không đúng JSON — vẫn đưa nội dung về cho người dùng đọc
        parsed = {"questions": [], "refined_A": parsed["__tho__"],
                  "notes": "AI trả lời không đúng định dạng JSON — nội dung thô ở trên."}
    return {
        "questions": parsed.get("questions") or [],
        "refined_A": parsed.get("refined_A") or "",
        "notes": parsed.get("notes") or "",
    }


@router.post("/api/execute")
def ai_execute(payload: dict = Body(...)):
    """AI THỰC HIỆN đề bài — tạo tool HTML, stream từng đoạn text về client.

    Body: {prompt: "đề bài hoàn chỉnh"}
    Trả về: text/plain stream (toàn bộ trả lời của Claude, client tự tách HTML).
    Stream giữ kết nối sống liên tục nên không vướng timeout proxy của Render.
    """
    api_key = os.getenv("ANTHROPIC_API_KEY", "")
    if not api_key:
        raise HTTPException(503, "Server chưa cấu hình ANTHROPIC_API_KEY "
                                 "(Render → Environment → thêm biến này)")
    model = os.getenv("ANTHROPIC_MODEL", "claude-sonnet-5")
    # Token suy nghĩ (adaptive thinking) ĐƯỢC TÍNH VÀO max_tokens — để 32000 thì
    # model nghĩ hết phần lớn hạn mức rồi bị cắt giữa chừng khi đang viết file.
    max_tokens = int(os.getenv("ANTHROPIC_MAX_TOKENS", "64000"))
    effort = os.getenv("ANTHROPIC_EFFORT", "medium")  # đặt "" để bỏ tham số
    prompt = (payload.get("prompt") or "").strip()
    if not prompt:
        raise HTTPException(422, 'Body thiếu "prompt"')

    user_msg = (prompt + "\n\nQUAN TRỌNG: Trả về TRỰC TIẾP file HTML hoàn chỉnh "
                "(bắt đầu bằng <!DOCTYPE html>, kết thúc bằng </html>), không lời dẫn, "
                "không markdown fence. Nếu đề bài thiếu thông tin thì tự chọn giá trị "
                "hợp lý và ghi chú ngay trong giao diện tool.\n"
                "JavaScript phải đúng cú pháp và chạy được ngay khi mở file bằng trình "
                "duyệt — rà lại trước khi kết thúc: khóa đối tượng động phải bọc ngoặc "
                "vuông ({['Màng '+t]: x} chứ không phải {'Màng '+t: x}), đóng đủ ngoặc "
                "và dấu nháy, không để hàm nào chưa định nghĩa.\n\n"
                "★★★ ĐIỀU QUAN TRỌNG NHẤT — ĐỌC TRƯỚC TIÊN ★★★\n"
                "Bảy thư viện sau ĐÃ CÓ SẴN dưới dạng biến toàn cục khi file chạy, do hệ "
                "thống tự chèn vào chỗ dấu <script>/*__SVWS_THREE__*/</script>:\n"
                "  THREE · SVWS3D · SVWSPID · SVWSGA · SVWSCHE · SVWSVT · SVWSDIEN\n"
                "TUYỆT ĐỐI KHÔNG được tự viết lại chúng. KHÔNG viết "
                "'const SVWSGA = (function(){…})();' hay bất kỳ khai báo nào đặt tên "
                "trùng bảy tên trên — khai báo trong file sẽ CHE MẤT bản thật, và bản "
                "bạn tự viết luôn kém hơn: đã đo trên một tool sinh ra trước đây, bản "
                "GA tự viết ghi 'Tỷ lệ 1:50' trong khi hệ số vẽ thực là 1:16,7 — nhãn "
                "nói dối; bản chọn thiết bị điện tự viết cho MỌI động cơ đều ra CB 10A "
                "và cáp 1,5 mm² bất kể công suất.\n"
                "Bạn CHỈ khai báo dữ liệu (EQUIP, PIPES, danh sách tải, nozzle…) rồi GỌI "
                "hàm của thư viện. Hệ thống sẽ tự động gỡ bỏ mọi định nghĩa trùng tên, "
                "nên viết lại chỉ tốn token vô ích và làm file dài gấp đôi.\n"
                "Cũng KHÔNG tự bịa hàm không có trong tài liệu dưới đây (ví dụ "
                "SVWS3D.widthOf không tồn tại — cần bề rộng thiết bị thì dùng "
                "SVWS3D.chanDe(e).w).\n\n"
                "THƯ VIỆN 3D — ĐỌC KỸ: Three.js r128 bản ĐẦY ĐỦ (biến toàn cục THREE) "
                "SẼ ĐƯỢC HỆ THỐNG TỰ ĐỘNG NHÚNG vào file sau khi bạn viết xong. Vì vậy:\n"
                "- TUYỆT ĐỐI KHÔNG tự viết mã thư viện Three.js, KHÔNG dùng CDN, và "
                "KHÔNG được hạ cấp xuống canvas 2D/isometric giả lập với lý do 'giới hạn "
                "offline' — WebGL 3D THẬT là bắt buộc theo chuẩn.\n"
                "- Đặt ĐÚNG một dòng sau vào trong <head>, hệ thống sẽ thay bằng thư viện thật:\n"
                "  <script>/*__SVWS_THREE__*/</script>\n"
                "- Rồi dùng THREE bình thường: THREE.Scene, THREE.WebGLRenderer, "
                "THREE.PerspectiveCamera, THREE.BoxGeometry, THREE.CylinderGeometry…\n"
                "- Kèm theo là thư viện SVWS3D (bộ dựng hình chuẩn của công ty). "
                "DÙNG NÓ, ĐỪNG tự gõ hình học — tự gõ toạ độ sẽ ra bố cục một đường "
                "thẳng, ống đi chéo xuyên thiết bị, tỷ lệ và màu lộn xộn.\n"
                "  Cách dùng (đơn vị mm, trục Y hướng lên):\n"
                "    const S = SVWS3D.scene(document.getElementById('view3d'));\n"
                "    const pos = SVWS3D.layout(EQUIP);   // tự xếp thành hàng có lối đi\n"
                "    EQUIP.forEach(e => S.addEquip(SVWS3D.build(e), pos[e.id], e));\n"
                "    S.addPipes(PIPES, pos);   // PHẢI gọi một lượt cả mảng, đừng gọi từng ống —\n"
                "                              // biết hết các tuyến thì thư viện mới chia được\n"
                "                              // tầng giá đỡ cho ống khỏi đè lên nhau\n"
                "    S.fit();                            // camera tự căn, không cắt mép\n"
                "    console.log(S.kiemTra());           // tự kiểm: ống có cắt qua thiết bị không\n"
                "  Bạn CHỈ khai báo dữ liệu, thư viện lo phần vẽ:\n"
                "    EQUIP = [{id,type,tag,name,...kích thước}] — type nhận: tank(d,h,"
                "service,level) · vessel(d,h,material,media,mediaColor) · cartridge(d,h) · "
                "pump(dup,soBom,dn) · roskid(vessels,memPerVessel,size,rows,bom,loc,tu) · "
                "edi(q,stacks,L,H,D) · panel(W,H,D) · dosing(d,h) · uv(d,L) · mixedbed\n"
                "    BƠM CẤP chạy 1 chạy 1 dự phòng: khai MỘT thiết bị pump với "
                "dup:true (tag ghi 'P-101A/B'), ĐỪNG khai thành hai thiết bị rời. Thư "
                "viện dựng đúng cách đấu thật: các bơm đứng SONG SONG giữa ống góp hút "
                "và ống góp đẩy chung, hai góp cùng dẫn về MỘT đầu để nối vào/ra nên "
                "mỗi nhánh thành hình chữ U; mỗi bơm có van chặn ở hút và ở đẩy để cô "
                "lập bơm mà không dừng hệ, cùng VAN MỘT CHIỀU ở đẩy. Thiếu van một "
                "chiều thì nước từ bơm đang chạy vòng ngược qua bơm dừng về góp hút, "
                "chạy lòng vòng trong cụm chứ không ra hệ thống — vẽ thiếu là thợ lắp "
                "thiếu. Bảng vật tư tự tính 2 van chặn + 1 van một chiều cho MỖI bơm, "
                "cộng hai tuyến ống góp và các tê rẽ nhánh.\n"
                "    BƠM CAO ÁP RO thì CHỈ MỘT, không chạy dự phòng — khai bom:1 trên "
                "roskid (hoặc pump không có dup).\n"
                "    P&ID: h1.bom({tag:'P-101A/B', dup:true}) vẽ đủ cụm song song kèm "
                "van chặn và van một chiều, không phải hai vòng tròn suông.\n"
                "    RO SKID thực tế XUẤT XƯỞNG ĐÃ GỒM bơm cao áp, vỏ lọc tinh và tủ "
                "điều khiển gắn ngay trên khung, đã đấu ống sẵn. Khai roskid với "
                "bom:2 (số bơm cao áp), loc:true (vỏ lọc tinh), tu:true (tủ điều khiển) "
                "để vẽ đúng như vậy — và KHI ĐÓ ĐỪNG khai chúng thành thiết bị riêng "
                "nữa. Khai riêng là vừa vẽ thừa thiết bị, vừa tính thừa hàng chục mét "
                "ống nối giữa chúng trong bảng vật tư, trong khi thực tế nhà chế tạo đã "
                "đấu sẵn trên skid. Đầu nối 'in' khi có loc:true tự chuyển sang miệng "
                "vỏ lọc cho đúng thứ tự công nghệ; thêm đầu nối 'dien' cho cáp vào tủ.\n"
                "    EDI là chồng TẤM–KHUNG (Veolia E-Cell MK-5 và tương đương), KHÔNG "
                "phải giàn vỏ màng nằm ngang như RO — đừng dùng roskid cho EDI. Khai q "
                "(m³/h) thì thư viện tự chia số stack theo định mức 5 m³/h MỖI STACK, "
                "hoặc khai thẳng stacks. Đầu nối riêng của EDI: in (nước cấp) · out "
                "(nước sản phẩm) · conc_in · conc (nước cô đặc) · dien (nước điện cực) · "
                "cip · dc (cáp nguồn một chiều).\n"
                "    PIPES = [{from,to,dn,service,fromPort,toPort}] — service nhận: raw · "
                "filtered · ro · di · chem · air · waste · steam · drain (màu tự theo lưu "
                "chất; ống tự đi vuông góc, tự chia tầng giá đỡ, tự chạy cao hơn thiết bị "
                "cao nhất nên không cắt qua bồn)\n"
                "    fromPort/toPort là TÊN ĐẦU NỐI, bỏ trống thì mặc định 'out' và 'in'. "
                "Mỗi thiết bị có sẵn các đầu nối đặt đúng vị trí kỹ thuật:\n"
                "      tank:      in (đỉnh) · out (thành, gần đáy) · tran (tràn) · xa (xả đáy)\n"
                "      vessel:    in (đỉnh) · out (đáy, qua đầu thu) · bw_in (rửa ngược vào, "
                "đáy) · bw_out (rửa ngược ra, đỉnh bên) · xa\n"
                "      cartridge: in (đáy bên) · out (đỉnh) · xa\n"
                "      pump:      in (hút, đầu trục) · out (đẩy, đỉnh buồng bơm)\n"
                "      roskid/edi: in (nước cấp) · out (nước thấm) · conc (nước cô đặc) · cip\n"
                "      dosing:    in (nạp) · out (đầu đẩy bơm định lượng)\n"
                "      uv:        in · out          |  panel: in · out (cáp)\n"
                "    Hãy vẽ ĐỦ cả nhánh phụ theo P&ID: rửa ngược, nước cô đặc, xả đáy, "
                "châm hoá chất — dùng đúng tên đầu nối ở trên.\n"
                "  Thư viện có sẵn: S.setView('iso'|'front'|'side'|'top'), "
                "S.layer('pipe'|'label'|'equip'|'ground', true/false), S.clear(), "
                "SVWS3D.PALETTE. Vòng lặp vẽ tự chạy — không cần tự viết render loop.\n"
                "  Khi người dùng đổi thông số: gọi S.clear() rồi dựng lại từ EQUIP/PIPES "
                "mới và S.fit() — đó là cách giữ tính parametric.\n"
                "  CHỈ tự viết Three.js thô cho chi tiết mà thư viện chưa có; phần bố cục "
                "và đường ống thì luôn dùng SVWS3D.\n"
                "- Tab 2 (P&ID): dùng thư viện SVWSPID, ĐỪNG tự gõ toạ độ SVG — tự gõ sẽ "
                "ra ký hiệu chồng lên nhau và nhãn đè lên thiết bị.\n"
                "    const D = SVWSPID.to({w:1560,h:980, ma:'<mã>', ten:'<tên>', dong2:'<cân bằng lưu lượng>'});\n"
                "    const h1 = D.hang(200);           // mỗi hàng là một mạch công nghệ\n"
                "    h1.nguon({nhan:'Nước cấp L-01 DN50'}); h1.van({});\n"
                "    h1.bon({tag:'TK-101', ghi:'9 m³ SS304'});\n"
                "    h1.bom({tag:'P-101A/B', ghi:'3 kW', dup:true});   // dup = 1 chạy 1 dự phòng\n"
                "    h1.cot({tag:'MMF-101', qty:2, ghi:'Ø1067'});      // qty = số cột song song\n"
                "    h1.loc({tag:'CF-101', ghi:'5 µm'});\n"
                "    h1.ro ({tag:'RO-101', vessels:5, ghi:'5×2 8040'});\n"
                "    h1.edi({tag:'EDI-101', stacks:4, ghi:'4 stack × 5 m³/h'});\n"
                "    h1.uv ({tag:'UV-101', ghi:'≥40 mJ/cm²'});\n"
                "    D.dungCu('TK-101','LIT-101');        // bầu đo, tự tìm chỗ trống\n"
                "    D.ghiChu('TK-101','LSL/LSH','duoi');\n"
                "    D.noi('RO-101','TK-101',{dong:'conc', nhan:'L-06 Conc RO1', phia:'duoi'});\n"
                "    el.innerHTML = D.ve();  console.log(D.kiemTra());\n"
                "  Thư viện tự đo ký hiệu và đo chữ, tự đẩy nhãn cho khỏi đè, tự nhảy con trỏ "
                "theo bề rộng thật, tự tìm làn trống khi nối. dong nhận: raw · di · conc · "
                "chem · cip. kiemTra() trả về danh sách chỗ còn đè nhau — phải rỗng.\n"
                "  Khổ giấy TỰ NỚI theo nội dung, nên cứ chia nhiều hàng cho dễ đọc "
                "(D.hang(180), D.hang(560), D.hang(940)…) thay vì nhồi cả dây chuyền "
                "vào một hàng — đừng lo tràn khổ.\n"
                'NÚT "ẨN / HIỆN WATERMARK" (mục 1.6 của chuẩn) — dùng bộ có sẵn, ĐỪNG '
                "tự viết: tự viết thì nút chỉ ẩn lớp phủ của riêng nó mà bỏ sót "
                "watermark vẽ chìm trong SVG các bản vẽ, bấm xong vẫn thấy chữ.\n"
                "    SVWSWM.gan();                       // gọi một lần khi mở tool\n"
                "    <button onclick=\"SVWSWM.dao()\">Ẩn / Hiện watermark</button>\n"
                "  SVWSWM.dat(true/false) đặt thẳng, SVWSWM.dangHien() hỏi trạng thái. "
                "Bộ này ẩn cả lớp phủ toàn trang lẫn chữ chìm trong mọi bản vẽ, và nhớ "
                "lựa chọn theo máy người dùng.\n"
                "- Tab 3 (GA — mặt bằng): dùng SVWSGA, và PHẢI dùng LẠI ĐÚNG bộ vị trí của "
                "tab 3D để hai bản vẽ khớp nhau:\n"
                "    const G = SVWSGA.to({ma:'<mã>', ten:'<tên>', hoVanHanh:800});\n"
                "    G.datCum(EQUIP, pos);   // pos CHÍNH LÀ SVWS3D.layout(EQUIP) dùng ở tab 1\n"
                "    G.rack('nước +300 / hoá chất +900');\n"
                "    G.kichThuoc();          // chuỗi kích thước tự sinh theo vị trí thật\n"
                "    el.innerHTML = G.ve();  console.log(G.kiemTra());\n"
                "  Thư viện tự chọn tỷ lệ cho vừa khổ A3, vẽ đúng tỷ lệ, tự sinh chuỗi kích "
                "thước và mũi tên hướng Bắc. kiemTra() báo cả nhãn đè nhau LẪN khe vận hành "
                "thiếu (dưới 800 mm thì không có chỗ đứng bảo trì) — phải rỗng.\n"
                "- Tab 4 (SD Chế tạo): dùng SVWSCHE, MỖI THIẾT BỊ MỘT TỜ. ĐỪNG tự vẽ "
                "ba ô chữ nhật giống nhau cho mọi thiết bị — thợ không chế tạo được.\n"
                "    const T = SVWSCHE.to({ma:'<mã>', duAn:'<dự án>', nguoiLap:'<tên>', "
                "ngay:'<dd/mm/yyyy>',\n"
                "      tag:'TK-101', ten:'Bể chứa nước thô', kieu:'bon',   // bon | cot\n"
                "      d:1800, h:2400, day:'phang',        // day: phang | chom | con\n"
                "      chan:'chande', caoChan:300,          // chan: chande | vay\n"
                "      vatLieu:'SS304', dayThan:5, dayDay:6, ap:0, mucNuoc:2200,\n"
                "      nozzle:[{ma:'N1', dv:'Nước vào', dn:65, cao:2250, goc:0, nho:150},\n"
                "              {ma:'N4', dv:'Xả đáy', dn:50, cao:120, goc:270, nho:130}]});\n"
                "    elVe.innerHTML = T.ve();  elBang.innerHTML = T.bang();\n"
                "    console.log(T.kiemTra());\n"
                "  cao = cao độ tâm nozzle đo TỪ MẶT TRONG ĐÁY (mm); goc = góc quay (độ, "
                "0° theo hình chiếu bằng); nho = độ nhô mặt bích khỏi thành. Bỏ trống "
                "mảng nozzle thì thư viện tự đặt bộ tiêu chuẩn theo loại thiết bị.\n"
                "  Thư viện tự chọn tỷ lệ, tự vẽ 3 hình chiếu đúng kích thước thật, tự "
                "ghi chuỗi kích thước và cao độ từng nozzle, vẽ hình chiếu bằng định vị "
                "góc quay, thêm khối ghi chú chế tạo + khung tên + watermark. T.bang() "
                "cho ra bảng nozzle, kiểm bề dày, bảng vật liệu, KHAI TRIỂN TÔN (kích "
                "thước phôi để cắt), bảng mối hàn, khối lượng, sơn phủ.\n"
                "  Kích thước d/h/vật liệu phải LẤY ĐÚNG từ EQUIP đã khai ở tab 3D — hai "
                "tab nói khác nhau là bản vẽ chế tạo sai.\n"
                "  kiemTra() bắt: bề dày không đủ chịu cột nước/áp suất, nozzle trùng mã, "
                "nozzle nằm ngoài thân, hai nozzle sát tới mức chồng vùng gia cường, "
                "thiếu xả đáy, thiếu cửa thăm. loi PHẢI rỗng.\n"
                "  Nút in tờ chế tạo PHẢI in cả HÌNH (T.ve()) chứ không chỉ bảng chữ.\n"
                "- Tab 5 (Phân tách thi công & vật tư): dùng SVWSVT. TUYỆT ĐỐI KHÔNG "
                "gõ hằng số kiểu qty:6 cho ống và qty:2 cho co 90° ở mọi công đoạn — "
                "đoạn dài 2 m và đoạn dài 30 m sẽ ra cùng một số, và tab BOQ lấy giá "
                "từ đây nên giá cũng sai.\n"
                "    // SAU khi đã dựng xong tab 3D (cần chính hình học đó):\n"
                "    const V = SVWSVT.to({ong:'UPVC', heSoHao:0.07, duTru:0.05});\n"
                "    V.nap(S.thongKeOng(), EQUIP, pos, PIPES);\n"
                "    elCongDoan.innerHTML = V.bangCongDoan();  // từng công đoạn + vật tư\n"
                "    elTongHop.innerHTML  = V.bangTongHop();   // gộp toàn hệ để đặt hàng\n"
                "    console.log(V.kiemTra());\n"
                "  S.thongKeOng() trả chiều dài THẬT (mm), số co 90° đếm theo khúc gãy "
                "thật và cao trình của từng tuyến đã vẽ, nên bảng vật tư không thể lệch "
                "với bản vẽ. Danh sách công đoạn tự sinh từ PIPES — đừng gõ tay danh "
                "sách công đoạn, gõ tay thì thêm thiết bị vào 3D mà quên thêm công đoạn "
                "là thiếu vật tư mà không ai phát hiện.\n"
                "  Thư viện tự tính: giá đỡ theo bước chuẩn của vật liệu, tê tại đầu nối "
                "nhiều nhánh, van một chiều ở đầu đẩy bơm, mặt bích theo số đầu nối và "
                "số van, móng theo chân đế thật, cáp điện theo khoảng cách Manhattan từ "
                "tủ tới từng động cơ. Mỗi dòng có cột CĂN CỨ TÍNH — giữ nguyên cột này.\n"
                "  kiemTra() bắt: công đoạn thiếu DN, tuyến ngắn bất thường (hai thiết bị "
                "trùng chỗ), THIẾT BỊ KHÔNG NỐI VÀO TUYẾN NÀO (thiếu vật tư), tuyến quá "
                "nhiều co gây tổn thất áp. loi PHẢI rỗng.\n"
                "  V.bangCongDoan() đã kèm sẵn BẢN VẼ từng công đoạn vẽ theo TUYẾN THẬT: "
                "mặt bằng tuyến đúng tỷ lệ + trắc dọc trải phẳng cho thấy ống lên giá cao "
                "bao nhiêu rồi hạ xuống đâu, có đánh dấu co 90°, vị trí giá đỡ, cao trình "
                "giá và trình tự lắp. Nút in công đoạn PHẢI in cả bản vẽ này (lấy riêng "
                "bằng V.veCongDoan(c) nếu cần), không chỉ in bảng vật tư.\n"
                "- Tab 9 (BOQ): LẤY DANH MỤC TỪ V.boq() hoặc V.bangBOQ(), KHÔNG tự gõ "
                "lại. V.boq() gộp đủ 4 nhóm: A thiết bị chính (bồn, bơm, cột lọc, cụm "
                "RO, EDI, UV, tủ điện — quy cách suy thẳng từ khai báo EQUIP nên không "
                "lệch bản vẽ), B đường ống & phụ kiện, C móng & kết cấu, D điện & điều "
                "khiển. Riêng V.bangThietBi() cho bảng thiết bị chính đứng riêng.\n"
                "  BOQ thiếu thiết bị chính là thiếu phần lớn tiền — kiemTra() báo lỗi "
                "nếu không có thiết bị nào hoặc có thiết bị chưa ghi quy cách.\n"
                "  Đơn giá do bạn ước tính điền vào (ghi rõ là số tham khảo, phải thay "
                "bằng báo giá nhà cung cấp trước khi phát hành); KHỐI LƯỢNG thì lấy "
                "nguyên từ V.boq().\n"
                "- Tab 7 (Logic vận hành & mạch điện): dùng SVWSDIEN. Đây là tab để "
                "THỢ ĐẤU TỦ và NGƯỜI LẬP TRÌNH PLC làm việc, nên phải đủ số liệu thật, "
                "ĐỪNG tự gõ toạ độ SVG cho sơ đồ điện.\n"
                "    // (1) mạch động lực — thư viện TỰ CHỌN CB, khởi động từ, rơ-le "
                "nhiệt, tiết diện cáp theo kW; chỉ khai báo tải:\n"
                "    const DL = SVWSDIEN.dongLuc({ma:'<mã>', ten:'<tên>', "
                "nguon:'3P+N 380/220 VAC 50 Hz'});\n"
                "    DL.tai({tag:'P-101A', ten:'Bơm cấp', kW:3, kieu:'DOL'});      "
                "// kieu: DOL | VFD ; pha:1 cho tải 1 pha\n"
                "    elDL.innerHTML = DL.ve();   elMotor.innerHTML = DL.bangMotor();\n"
                "    // (2) mạch điều khiển dạng thang — tự đánh số nấc, số dây, "
                "tham chiếu chéo:\n"
                "    const DK = SVWSDIEN.dieuKhien({ma:'<mã>'});\n"
                "    DK.mach({ten:'Cho phép chạy', pt:[{k:'estop',t:'ES-01'},"
                "{k:'nc',t:'F1'},{k:'cuon',t:'KA-01'}]});\n"
                "      k nhận: estop · no · nc · nut · chon · tg · di · do · cuon · "
                "den · coi. Cuộn/đèn/còi tự ép sát ray phải.\n"
                "    // (3) đấu nối PLC + bảng I/O:\n"
                "    const IO = SVWSDIEN.plc({ma:'<mã>', cpu:'S7-1200 CPU1214C'});\n"
                "    IO.module({kieu:'DI', ten:'DI on-board', kenh:[{dc:'I0.0',"
                "tag:'ES-01', mo:'Nút dừng khẩn cấp', tin:'NC'}]});   // kieu: DI/DO/AI/AO\n"
                "    elIO.innerHTML = IO.ve();   elBangIO.innerHTML = IO.bangIO();\n"
                "    // (4) bố trí tủ theo tỷ lệ thật (r = bề rộng mm, c = chiều cao mm, "
                "nhiet = W toả ra):\n"
                "    const TU = SVWSDIEN.tuDien({W:800, H:1200, D:300});\n"
                "    TU.thiet({ten:'MCCB 100A', r:105, c:165, nhiet:12});\n"
                "    // (5) bảng logic vận hành — phần người lập trình PLC đọc để viết "
                "chương trình:\n"
                "    const LG = SVWSDIEN.logic({});\n"
                "    LG.thietBi({tag:'P-101A', ten:'Bơm cấp', che:'AUTO/MAN', "
                "chay:'<điều kiện>', dung:'<điều kiện>', khoa:'<khoá liên động>', "
                "bao:'<báo động>'});\n"
                "    LG.trinhTu({b:1, viec:'...', dk:'...', t:'15 s', loi:'...'});\n"
                "    LG.baoDong({ma:'A-01', mo:'Dừng khẩn cấp', muc:'Sự cố', "
                "nguong:'ES-01=0', tacDong:'Cắt toàn bộ đầu ra', xuLy:'...'});\n"
                "    elLG.innerHTML = LG.tatCa();   // 4 bảng + ma trận khoá liên động\n"
                "  Nhúng SVWSDIEN.CSS một lần để các bảng có định dạng. Mỗi bộ có "
                "kiemTra() trả {loi, canhBao}: loi PHẢI RỖNG (thiếu nút dừng khẩn cấp, "
                "trùng địa chỉ PLC, thiết bị đè nhau, tủ không đủ chỗ); canhBao là "
                "khuyến nghị kỹ thuật, hiện lên cho người dùng đọc chứ không chặn.\n"
                "  Số liệu phải KHỚP giữa các tab: tag động cơ ở mạch động lực trùng "
                "tag bơm ở 3D/P&ID; mỗi động cơ phải có đủ DO lệnh chạy và DI phản hồi "
                "trong bảng I/O; mỗi dụng cụ đo trên P&ID phải có một kênh AI.\n"
                "- Dồn công sức vào MÔ TẢ ĐÚNG dây chuyền: đủ thiết bị theo Phần A, "
                "đúng thứ tự dòng công nghệ, tag thiết bị chuẩn (T-101, P-101, F-101…), "
                "kích thước suy từ thông số thiết kế.\n\n"
                'NÚT "CHỈNH BẢN VẼ BẰNG AI" (mục 3.3 của chuẩn) — làm ĐÚNG như sau:\n'
                "- TUYỆT ĐỐI KHÔNG hỏi người dùng nhập API key và KHÔNG gọi "
                "https://api.anthropic.com (trình duyệt chặn CORS, và để khóa trong file "
                "là lộ khóa cho bất kỳ ai mở file).\n"
                f"- Gọi CỬA NGÕ của công ty: POST {_PUBLIC_URL}/registry/api/claude\n"
                '  Body y hệt dạng Anthropic: {"max_tokens":2000,"messages":[{"role":"user",'
                '"content":"<yêu cầu người dùng + JSON thông số hiện tại + mô tả các trường>"}]}\n'
                "  KHÔNG gửi header x-api-key, KHÔNG gửi anthropic-version — server tự lo khóa.\n"
                "  Đáp án trả về đúng dạng Anthropic: đọc content[0].text (bảo AI trả JSON "
                "thông số đã cập nhật, rồi gán lại vào state và vẽ lại mọi tab).\n"
                "- Lỗi mạng thì báo: 'Cần kết nối mạng để chỉnh bằng AI — các chức năng "
                "còn lại vẫn chạy offline 100%.'")

    def gen():
        # Model suy nghĩ (adaptive thinking) có thể vài phút trước khi viết chữ
        # đầu tiên; Render cắt stream im lặng quá lâu → phải bơm "nhịp tim" (dấu
        # chấm) trong pha suy nghĩ. Client chỉ lấy từ <!DOCTYPE trở đi nên các
        # byte này tự bị loại khi ghép file.
        started_text = False
        last_beat = [datetime.datetime.now(datetime.timezone.utc)]

        def beat():
            """Trả về '.' tối đa 1 lần/2 giây khi chưa có nội dung HTML."""
            now = datetime.datetime.now(datetime.timezone.utc)
            if (now - last_beat[0]).total_seconds() >= 2.0:
                last_beat[0] = now
                return "."
            return ""

        body_json = {"model": model, "max_tokens": max_tokens, "stream": True,
                     "messages": [{"role": "user", "content": user_msg}]}
        if effort:
            body_json["output_config"] = {"effort": effort}

        try:
            with httpx.stream(
                "POST", _ANTHROPIC_URL,
                headers={"x-api-key": api_key,
                         "anthropic-version": "2023-06-01",
                         "content-type": "application/json"},
                json=body_json,
                timeout=httpx.Timeout(1800.0, connect=30.0),
            ) as resp:
                if resp.status_code != 200:
                    body = resp.read().decode("utf-8", "replace")
                    yield "\n[SVWS-LỖI] " + _loi_ai_vn(resp.status_code, body)
                    return
                yield "."  # mở stream ngay để proxy không buffer/timeout
                for line in resp.iter_lines():
                    if not line.startswith("data: "):
                        continue
                    try:
                        ev = json.loads(line[6:])
                    except json.JSONDecodeError:
                        continue
                    etype = ev.get("type")
                    if etype == "content_block_delta":
                        delta = ev.get("delta", {})
                        if delta.get("type") == "text_delta":
                            started_text = True
                            yield delta.get("text", "")
                        elif not started_text:  # thinking_delta… trong pha suy nghĩ
                            hb = beat()
                            if hb:
                                yield hb
                    elif etype == "message_delta":
                        # Luôn đính lý do dừng + số token vào cuối stream để chẩn
                        # đoán được khi file ra thiếu. Client cắt tại </html> nên
                        # dấu này tự biến mất khi file lành lặn.
                        stop = (ev.get("delta") or {}).get("stop_reason")
                        used = (ev.get("usage") or {}).get("output_tokens", "?")
                        if stop == "max_tokens":
                            yield (f"\n[SVWS-LỖI] File bị cắt vì chạm trần {max_tokens:,} token "
                                   f"(đã dùng {used}). Tăng ANTHROPIC_MAX_TOKENS trên Render, "
                                   f"hoặc chia đề bài thành phần nhỏ hơn.")
                            return
                        yield f"\n[SVWS-KET] stop={stop} out={used} tran={max_tokens}"
                    elif etype == "error":
                        yield "\n[SVWS-LỖI] " + _loi_ai_vn(
                            200, json.dumps(ev.get("error", {}), ensure_ascii=False))
                        return
                    elif not started_text:  # ping / content_block_start…
                        hb = beat()
                        if hb:
                            yield hb
        except httpx.HTTPError as e:
            yield f"\n[SVWS-LỖI] Không gọi được Anthropic API: {e}"

    # no-transform: chặn Cloudflare nén gzip (nén sẽ gom buffer làm trễ stream)
    return StreamingResponse(gen(), media_type="text/plain; charset=utf-8",
                             headers={"Cache-Control": "no-cache, no-transform",
                                      "X-Accel-Buffering": "no"})


MAX_SOW_MB = 20


def _docx_text(data: bytes) -> str:
    """Rút chữ từ file .docx (là file zip chứa word/document.xml) — không cần thư viện."""
    import io, zipfile
    from html import unescape
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as z:
            xml = z.read("word/document.xml").decode("utf-8", "replace")
    except Exception as e:
        raise ValueError(f"Không mở được file .docx: {e}")
    xml = re.sub(r"<w:tab[^>]*/>", "\t", xml)
    xml = re.sub(r"</w:p>", "\n", xml)
    xml = re.sub(r"<w:br[^>]*/>", "\n", xml)
    txt = unescape(re.sub(r"<[^>]+>", "", xml))
    return re.sub(r"\n{3,}", "\n\n", txt).strip()


def _xlsx_text(data: bytes) -> str:
    """Đổ bảng Excel thành text để AI đọc (SOW hay kèm bảng thiết bị/thông số)."""
    import io
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    out = []
    for ws in wb.worksheets:
        out.append(f"--- SHEET: {ws.title} ---")
        for row in ws.iter_rows(values_only=True):
            cells = ["" if c is None else str(c).strip() for c in row]
            if any(cells):
                out.append(" | ".join(cells))
    return "\n".join(out)[:120000]


def _khoi_file(data: bytes, content_type: str, filename: str) -> dict:
    """Dựng khối nội dung gửi Claude từ file người dùng tải lên."""
    import base64
    ct = (content_type or "").lower()
    fn = (filename or "").lower()
    if ct == "application/pdf" or fn.endswith(".pdf"):
        return {"type": "document",
                "source": {"type": "base64", "media_type": "application/pdf",
                           "data": base64.b64encode(data).decode()}}
    if ct.startswith("image/") or fn.endswith((".png", ".jpg", ".jpeg", ".webp", ".gif")):
        mt = ct if ct.startswith("image/") else ("image/png" if fn.endswith(".png") else "image/jpeg")
        return {"type": "image",
                "source": {"type": "base64", "media_type": mt,
                           "data": base64.b64encode(data).decode()}}
    if fn.endswith(".docx"):
        return {"type": "text", "text": _docx_text(data)[:120000]}
    if fn.endswith((".xlsx", ".xlsm")):
        return {"type": "text", "text": _xlsx_text(data)}
    if ct.startswith("text/") or fn.endswith((".txt", ".csv", ".md")):
        return {"type": "text", "text": data.decode("utf-8", errors="replace")[:120000]}
    if fn.endswith(".doc"):
        raise ValueError("File .doc là Word đời cũ — mở bằng Word rồi lưu lại thành .docx "
                         "(hoặc in ra PDF) là đọc được.")
    raise ValueError("Định dạng chưa hỗ trợ — dùng PDF, Word (.docx), Excel (.xlsx), "
                     "ảnh (PNG/JPG) hoặc TXT/CSV.")


@router.post("/api/sow")
async def ai_sow(file: UploadFile = File(...)):
    """Đọc file SOW (phạm vi công việc) → soạn sẵn PHIẾU THIẾT KẾ.

    Trả: {design: {...}, questions: [...], tom_tat: "..."}
    Người dùng xem lại rồi mới bấm thêm vào sổ.
    """
    api_key = os.getenv("ANTHROPIC_API_KEY", "")
    if not api_key:
        raise HTTPException(503, "Server chưa cấu hình ANTHROPIC_API_KEY "
                                 "(Render → Environment → thêm biến này)")
    model = os.getenv("ANTHROPIC_MODEL", "claude-sonnet-5")

    data = await file.read()
    if not data:
        raise HTTPException(422, "File rỗng")
    if len(data) > MAX_SOW_MB * 1024 * 1024:
        raise HTTPException(413, f"File nặng quá {MAX_SOW_MB}MB — nén lại hoặc tách bớt phần phụ lục.")
    try:
        khoi = _khoi_file(data, file.content_type or "", file.filename or "")
    except ValueError as e:
        raise HTTPException(415, str(e))

    with Session(engine) as s:
        doc = s.get(RegistryDoc, 1)
        d = doc.data if doc else {}
        summary = _sections_summary(d)
        ma_cu = ", ".join((x.get("code") or "") for x in (d.get("designs") or [])[:20])

    sys_msg = (
        "Bạn là kỹ sư trưởng của SVWS (công ty EPC xử lý nước / nước thải / khí thải, TP.HCM). "
        "Người dùng gửi một file SOW (Scope of Work — phạm vi công việc) của dự án. "
        "Hãy ĐỌC KỸ và soạn sẵn một PHIẾU THIẾT KẾ để đưa vào Sổ đăng ký thiết kế.\n\n"
        f"TÓM TẮT CHUẨN 3D DESIGN CỦA CÔNG TY:\n{summary}\n\n"
        f"CÁC MÃ THIẾT KẾ ĐÃ CÓ (đừng trùng): {ma_cu or '(chưa có)'}\n\n"
        "CHỈ trả về JSON hợp lệ, không markdown, không giải thích ngoài JSON:\n"
        '{"design":{"code":"<mã ngắn gọn theo kiểu SVWS-XXX, tự đặt từ nội dung SOW>",'
        '"name":"<tên thiết kế>","tech":"<công nghệ chính>","capacity":"<công suất kèm đơn vị>",'
        '"project":"<tên dự án / khách hàng>",'
        '"params":"<thông số chính, mỗi ý một dòng, gạch đầu dòng>",'
        '"notes":"<yêu cầu riêng / ràng buộc / tiêu chuẩn đầu ra, mỗi ý một dòng>"},'
        '"questions":["<tối đa 6 câu hỏi về thông tin SOW còn thiếu để thiết kế được>"],'
        '"tom_tat":"<2-3 câu tóm tắt phạm vi công việc>"}\n\n'
        "QUY TẮC: KHÔNG bịa số liệu không có trong SOW — thiếu thì để chuỗi rỗng và "
        "nêu thành câu hỏi trong \"questions\". Viết tiếng Việt, ngắn gọn, đúng kỹ thuật."
    )

    parsed = _goi_ai_json(
        api_key, model,
        [khoi, {"type": "text",
                "text": "Đọc SOW này và soạn phiếu thiết kế theo đúng định dạng JSON đã nêu."}],
        16000, system=sys_msg, timeout=420.0)
    dsg = parsed.get("design") or {}
    if not isinstance(dsg, dict) or not (dsg.get("name") or dsg.get("tech")):
        raise HTTPException(502, "AI không rút được nội dung thiết kế từ file này — "
                                 "kiểm tra đúng file SOW chưa.")
    design = {k: str(dsg.get(k) or "") for k in
              ("code", "name", "tech", "capacity", "project", "params", "notes")}
    design["status"] = "Ý tưởng"
    if not design["code"]:
        design["code"] = "SVWS-SOW"
    return {"design": design,
            "questions": [q for q in (parsed.get("questions") or []) if q][:6],
            "tom_tat": parsed.get("tom_tat") or "",
            "nguon": file.filename or ""}


@router.post("/api/claude")
def claude_proxy(payload: dict = Body(...)):
    """Cửa ngõ gọi Claude cho các TOOL sinh ra (nút "Chỉnh bản vẽ bằng AI").

    Tool tải về KHÔNG được ôm API key (lộ khóa cho bất kỳ ai mở file) và cũng
    không gọi thẳng api.anthropic.com được (trình duyệt chặn CORS). Endpoint này
    nhận ĐÚNG dạng body của Anthropic và trả NGUYÊN VĂN đáp án, nên tool chỉ cần
    đổi URL là chạy — không đổi gì khác.
    """
    api_key = os.getenv("ANTHROPIC_API_KEY", "")
    if not api_key:
        raise HTTPException(503, "Server chưa cấu hình ANTHROPIC_API_KEY")
    msgs = payload.get("messages")
    if not isinstance(msgs, list) or not msgs:
        raise HTTPException(422, 'Body thiếu "messages"')

    # CỐ TÌNH bỏ qua "model" do tool gửi lên: tool cũ hay ghim mã model đời cũ đã
    # ngừng phục vụ. Server luôn dùng model hiện hành → tool cũ tự sống lại.
    body = {"model": os.getenv("ANTHROPIC_MODEL", "claude-sonnet-5"),
            "max_tokens": max(256, min(int(payload.get("max_tokens") or 4000), 8000)),
            "messages": msgs}
    if payload.get("system"):
        body["system"] = payload["system"]
    effort = os.getenv("ANTHROPIC_EFFORT", "medium")
    if effort:
        body["output_config"] = {"effort": effort}
    if len(json.dumps(body)) > 400_000:      # chặn gửi cả file khổng lồ qua đây
        raise HTTPException(413, "Yêu cầu quá lớn — chỉ gửi thông số cần chỉnh, "
                                 "đừng gửi cả nội dung tool.")
    try:
        resp = httpx.post(_ANTHROPIC_URL,
                          headers={"x-api-key": api_key,
                                   "anthropic-version": "2023-06-01",
                                   "content-type": "application/json"},
                          json=body, timeout=300.0)
    except httpx.HTTPError as e:
        raise HTTPException(502, f"Không gọi được Anthropic API: {e}")
    if resp.status_code != 200:
        raise HTTPException(502, _loi_ai_vn(resp.status_code, resp.text))
    return resp.json()


@router.post("/api/fix")
def ai_fix(payload: dict = Body(...)):
    """Sửa lỗi cú pháp JS bằng PHÉP THAY THẾ CHUỖI, không chép lại cả file.

    Bắt AI chép lại file 50KB chỉ để sửa một dòng là tự chuốc rủi ro đứt giữa
    chừng; ở đây AI chỉ trả vài trăm ký tự nên nhanh, rẻ và gần như không hỏng.

    Body: {html: "...", error: "Unexpected token '+'"}
    Trả: {replacements: [{old, new}], note}
    """
    api_key = os.getenv("ANTHROPIC_API_KEY", "")
    if not api_key:
        raise HTTPException(503, "Server chưa cấu hình ANTHROPIC_API_KEY")
    model = os.getenv("ANTHROPIC_MODEL", "claude-sonnet-5")
    html = payload.get("html") or ""
    err = (payload.get("error") or "").strip()
    if len(html) < 200:
        raise HTTPException(422, 'Body thiếu "html"')

    user_msg = f"""File HTML dưới đây có LỖI CÚ PHÁP JavaScript nên mở ra trang trắng.
Trình duyệt báo: {err}

Tìm ĐÚNG chỗ sai và trả về các phép thay thế chuỗi TỐI THIỂU để sửa.
Quy tắc bắt buộc:
- "old" phải là đoạn văn bản NGUYÊN VĂN copy từ file, đủ dài để chỉ khớp DUY NHẤT một chỗ.
- "new" là đoạn thay thế đã sửa đúng cú pháp.
- Chỉ sửa lỗi cú pháp, KHÔNG đổi giao diện hay chức năng.
- Lỗi hay gặp: khóa đối tượng động thiếu ngoặc vuông — {{'Màng '+t: x}} phải là {{['Màng '+t]: x}}.

CHỈ trả JSON hợp lệ, không markdown, không giải thích ngoài JSON:
{{"replacements":[{{"old":"...","new":"..."}}],"note":"tóm tắt ngắn chỗ đã sửa"}}

===== FILE =====
{html}"""

    parsed = _goi_ai_json(api_key, model, user_msg, 12000, timeout=300.0)
    reps = [r for r in (parsed.get("replacements") or [])
            if isinstance(r, dict) and r.get("old") and r.get("new") is not None]
    if not reps:
        raise HTTPException(502, "AI không chỉ ra được chỗ cần sửa — làm lại tool từ đầu.")
    return {"replacements": reps, "note": parsed.get("note") or ""}
