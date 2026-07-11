"""
CỔNG AI — phân loại thư phản hồi + gợi ý trả lời.
Dev: FakeAIProvider (luật từ khóa tiếng Việt, chạy offline).
Prod: AnthropicAIProvider (gọi Claude API, trả JSON). Đổi AI_PROVIDER=ANTHROPIC + ANTHROPIC_API_KEY.

Kết quả: dict {y_dinh, khan, tom_tat, tra_loi}
  y_dinh ∈ QUAN_TAM/HOI_KY_THUAT/HEN_GAP/TU_CHOI/KHIEU_NAI/HUY_NHAN/VANG_MAT/SPAM/KHAC
  khan   ∈ CAO/TRUNG/THAP
"""
import json
from typing import Protocol
from .config import settings

Y_DINH = {"QUAN_TAM", "HOI_KY_THUAT", "HEN_GAP", "TU_CHOI",
          "KHIEU_NAI", "HUY_NHAN", "VANG_MAT", "SPAM", "KHAC"}

_TRA_LOI = {
    "QUAN_TAM": "Cảm ơn Quý công ty đã quan tâm. SVWS sẽ gửi báo giá chi tiết kèm phương án kỹ thuật phù hợp trong thời gian sớm nhất. Vui lòng cho biết công suất và yêu cầu cụ thể để chúng tôi tư vấn chính xác.",
    "HOI_KY_THUAT": "Cảm ơn câu hỏi của Quý công ty. Bộ phận kỹ thuật SVWS sẽ giải đáp chi tiết và có thể sắp lịch khảo sát nếu cần. Quý công ty vui lòng cung cấp thêm thông tin hiện trạng hệ thống.",
    "HEN_GAP": "SVWS rất sẵn lòng sắp xếp buổi khảo sát/gặp trao đổi. Vui lòng cho biết thời gian và địa điểm thuận tiện để chúng tôi chủ động bố trí nhân sự.",
    "TU_CHOI": "Cảm ơn Quý công ty đã phản hồi. SVWS rất mong có cơ hội hợp tác trong tương lai và luôn sẵn sàng hỗ trợ khi Quý công ty có nhu cầu.",
    "KHIEU_NAI": "SVWS rất tiếc về sự bất tiện này. Chúng tôi tiếp nhận phản ánh và sẽ cử bộ phận phụ trách kiểm tra, xử lý ưu tiên. Mong Quý công ty thông cảm và cho phép chúng tôi liên hệ ngay để khắc phục.",
    "KHAC": "Cảm ơn Quý công ty đã phản hồi. SVWS đã tiếp nhận và sẽ liên hệ lại trong thời gian sớm nhất.",
}

def _khan_cho(y_dinh: str, text: str) -> str:
    t = text.lower()
    if y_dinh == "KHIEU_NAI" or any(k in t for k in ["gấp", "khẩn", "ngay", "sớm nhất", "urgent"]):
        return "CAO"
    if y_dinh in ("QUAN_TAM", "HEN_GAP", "HOI_KY_THUAT"):
        return "TRUNG"
    return "THAP"


class AIProvider(Protocol):
    ten: str
    def phan_loai(self, tieu_de: str, noi_dung: str) -> dict: ...


class FakeAIProvider:
    ten = "DEMO"
    def phan_loai(self, tieu_de, noi_dung) -> dict:
        t = f"{tieu_de or ''} {noi_dung or ''}".lower()
        def has(*ks): return any(k in t for k in ks)
        if has("hủy đăng ký", "ngừng nhận", "không nhận email", "unsubscribe", "bỏ nhận tin", "ngưng gửi", "ngung gui"):
            y = "HUY_NHAN"
        elif has("vắng mặt", "out of office", "nghỉ phép", "auto-reply", "tự động trả lời", "đang đi công tác", "hồi âm tự động"):
            y = "VANG_MAT"
        elif has("khiếu nại", "không hài lòng", "sự cố", "hỏng", "hư hỏng", "lỗi", "kém", "thất vọng", "phàn nàn"):
            y = "KHIEU_NAI"
        elif has("khảo sát", "hẹn", "gặp", "lịch", "cuộc họp", "đến xem", "ghé"):
            y = "HEN_GAP"
        elif has("báo giá", "quan tâm", "tư vấn", "công suất", "m3", "mét khối", "cần mua", "muốn lắp", "đề nghị", "giá"):
            y = "QUAN_TAM"
        elif has("kỹ thuật", "thông số", "công nghệ", "màng", "ro", "mbr", "hỏi về", "tài liệu"):
            y = "HOI_KY_THUAT"
        elif has("không có nhu cầu", "từ chối", "không quan tâm", "đã có nhà cung cấp", "cảm ơn nhưng"):
            y = "TU_CHOI"
        elif has("trúng thưởng", "khuyến mãi", "vay vốn", "casino", "viagra"):
            y = "SPAM"
        else:
            y = "KHAC"
        tom_tat = (noi_dung or "").strip().replace("\n", " ")[:140]
        return {"y_dinh": y, "khan": _khan_cho(y, t),
                "tom_tat": tom_tat, "tra_loi": _TRA_LOI.get(y, _TRA_LOI["KHAC"])}


class AnthropicAIProvider:
    ten = "ANTHROPIC"
    def phan_loai(self, tieu_de, noi_dung) -> dict:
        import urllib.request
        sys = ("Bạn là trợ lý phân loại email phản hồi khách hàng cho công ty xử lý nước. "
               "CHỈ trả về JSON, không thêm chữ nào khác, dạng: "
               '{"y_dinh":"<một trong: QUAN_TAM,HOI_KY_THUAT,HEN_GAP,TU_CHOI,KHIEU_NAI,HUY_NHAN,VANG_MAT,SPAM,KHAC>",'
               '"khan":"<CAO|TRUNG|THAP>","tom_tat":"<tóm tắt 1 câu tiếng Việt>",'
               '"tra_loi":"<nội dung email trả lời lịch sự, chuyên nghiệp, tiếng Việt>"}')
        body = {"model": settings.anthropic_model, "max_tokens": 1024,
                "system": sys,
                "messages": [{"role": "user",
                              "content": f"Tiêu đề: {tieu_de}\nNội dung:\n{noi_dung}"}]}
        req = urllib.request.Request(
            "https://api.anthropic.com/v1/messages",
            data=json.dumps(body).encode("utf-8"),
            headers={"content-type": "application/json",
                     "x-api-key": settings.anthropic_api_key,
                     "anthropic-version": "2023-06-01"})
        try:
            with urllib.request.urlopen(req, timeout=30) as r:
                data = json.loads(r.read().decode("utf-8"))
            txt = "".join(b.get("text", "") for b in data.get("content", []) if b.get("type") == "text")
            txt = txt.strip().removeprefix("```json").removeprefix("```").removesuffix("```").strip()
            out = json.loads(txt)
            if out.get("y_dinh") not in Y_DINH:
                out["y_dinh"] = "KHAC"
            out.setdefault("khan", "TRUNG"); out.setdefault("tom_tat", ""); out.setdefault("tra_loi", "")
            return out
        except Exception:
            return FakeAIProvider().phan_loai(tieu_de, noi_dung)  # an toàn: rớt về luật


def lay_ai_provider() -> AIProvider:
    return AnthropicAIProvider() if settings.ai_provider.upper() == "ANTHROPIC" else FakeAIProvider()


def dich_vn_sang_en(texts: list[str]) -> list[str] | None:
    """Dịch danh sách đoạn văn hợp đồng VN -> EN bằng Claude.
    Trả None nếu chưa cấu hình ANTHROPIC_API_KEY hoặc gọi API thất bại."""
    if not settings.anthropic_api_key:
        return None
    import urllib.request
    sys_p = ("Bạn là biên dịch viên hợp đồng thương mại chuyên nghiệp. "
             "Người dùng gửi một MẢNG JSON các đoạn văn tiếng Việt của một hợp đồng dịch vụ. "
             "Dịch TỪNG phần tử sang tiếng Anh pháp lý trang trọng, giữ nguyên số liệu, tên riêng, "
             "ký hiệu kỹ thuật và các dấu xuống dòng (\\n). "
             "CHỈ trả về đúng một mảng JSON các chuỗi đã dịch, cùng số phần tử và cùng thứ tự, "
             "không thêm bất kỳ chữ nào khác.")
    body = {"model": settings.anthropic_model, "max_tokens": 8000, "system": sys_p,
            "messages": [{"role": "user", "content": json.dumps(texts, ensure_ascii=False)}]}
    req = urllib.request.Request(
        "https://api.anthropic.com/v1/messages",
        data=json.dumps(body).encode("utf-8"),
        headers={"content-type": "application/json",
                 "x-api-key": settings.anthropic_api_key,
                 "anthropic-version": "2023-06-01"})
    try:
        with urllib.request.urlopen(req, timeout=90) as r:
            data = json.loads(r.read().decode("utf-8"))
        txt = "".join(b.get("text", "") for b in data.get("content", []) if b.get("type") == "text")
        txt = txt.strip().removeprefix("```json").removeprefix("```").removesuffix("```").strip()
        out = json.loads(txt)
        if isinstance(out, list) and len(out) == len(texts):
            return [str(x) for x in out]
    except Exception:
        pass
    return None


# ============ AI ĐỌC FILE (báo giá / hồ sơ NCC) ============
def _excel_sang_text(data: bytes) -> str:
    """Giải mã file Excel (.xlsx/.xlsm) thành bảng văn bản cho AI đọc —
    mỗi sheet một khối, mỗi dòng là các ô cách nhau bằng ' | '."""
    import io
    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception:
        raise ValueError("Không mở được file Excel — kiểm tra file .xlsx không hỏng "
                         "và không đặt mật khẩu.")
    phan = []
    for ws in wb.worksheets:
        dong = []
        for row in ws.iter_rows(values_only=True):
            o = [("" if v is None else str(v)).strip() for v in row]
            if any(o):
                dong.append(" | ".join(o).rstrip(" |"))
            if len(dong) >= 1500:
                break
        if dong:
            phan.append(f"### Sheet: {ws.title}\n" + "\n".join(dong))
    wb.close()
    txt = "\n\n".join(phan)
    if not txt.strip():
        raise ValueError("File Excel không có dữ liệu để đọc.")
    return txt[:60000]


def _khoi_noi_dung_file(data: bytes, content_type: str, filename: str) -> dict:
    """Dựng khối nội dung gửi Claude từ file người dùng tải lên (PDF/ảnh/Excel/CSV-TXT)."""
    import base64
    ct = (content_type or "").lower()
    fn = (filename or "").lower()
    if ct == "application/pdf" or fn.endswith(".pdf"):
        return {"type": "document",
                "source": {"type": "base64", "media_type": "application/pdf",
                           "data": base64.b64encode(data).decode()}}
    if ct in ("image/png", "image/jpeg", "image/webp", "image/gif") or \
            fn.endswith((".png", ".jpg", ".jpeg", ".webp", ".gif")):
        mt = ct if ct.startswith("image/") else ("image/png" if fn.endswith(".png") else "image/jpeg")
        return {"type": "image",
                "source": {"type": "base64", "media_type": mt,
                           "data": base64.b64encode(data).decode()}}
    if ct == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" or \
            fn.endswith((".xlsx", ".xlsm")):
        return {"type": "text", "text": _excel_sang_text(data)}
    if ct.startswith("text/") or fn.endswith((".txt", ".csv")):
        return {"type": "text", "text": data.decode("utf-8", errors="replace")[:60000]}
    if fn.endswith(".xls") or ct == "application/vnd.ms-excel":
        raise ValueError("File .xls là Excel đời cũ — mở bằng Excel rồi lưu lại thành .xlsx là đọc được.")
    raise ValueError("Định dạng chưa hỗ trợ — dùng PDF, ảnh (PNG/JPG), Excel (.xlsx) hoặc CSV/TXT.")


def _loi_ai_ro_rang(e: Exception) -> str:
    """Đổi lỗi kỹ thuật khi gọi Claude thành thông báo tiếng Việt cụ thể."""
    import urllib.error
    if isinstance(e, urllib.error.HTTPError):
        chi_tiet = ""
        try:
            body = json.loads(e.read().decode("utf-8", errors="replace"))
            chi_tiet = str((body.get("error") or {}).get("message") or "")[:300]
        except Exception:
            pass
        if e.code == 401:
            return ("Khóa API không hợp lệ hoặc đã bị thu hồi (lỗi 401). "
                    "Vào Render → Environment → ANTHROPIC_API_KEY, dán khóa mới rồi Save.")
        if e.code == 400 and "credit" in chi_tiet.lower():
            return ("Tài khoản Anthropic hết credit (lỗi 400). "
                    "Vào console.anthropic.com → Billing để nạp thêm credit API.")
        if e.code == 404:
            return (f"Model AI không khả dụng (lỗi 404): {settings.anthropic_model}. "
                    "Kiểm tra biến ANTHROPIC_MODEL trên Render.")
        if e.code == 403:
            return "Khóa API không có quyền dùng model này (lỗi 403) — kiểm tra tài khoản Anthropic."
        if e.code == 429:
            return "Vượt giới hạn gọi AI (lỗi 429) — chờ khoảng 1 phút rồi thử lại."
        if e.code >= 500:
            return f"Máy chủ AI đang quá tải (lỗi {e.code}) — thử lại sau ít phút."
        return f"Gọi AI thất bại (lỗi {e.code}). {chi_tiet}".strip()
    if isinstance(e, urllib.error.URLError):
        return "Không kết nối được máy chủ AI (lỗi mạng) — thử lại sau."
    return f"Gọi AI thất bại ({type(e).__name__}) — thử lại sau."


def _goi_claude_json(khoi: dict, sys: str, cau_hoi: str, max_tokens: int = 4000, timeout: int = 120):
    """Gửi 1 khối nội dung + câu hỏi tới Claude, trả về text kết quả (đã bỏ ```json)."""
    import urllib.request
    body = {"model": settings.anthropic_model, "max_tokens": max_tokens, "system": sys,
            "messages": [{"role": "user", "content": [khoi, {"type": "text", "text": cau_hoi}]}]}
    req = urllib.request.Request(
        "https://api.anthropic.com/v1/messages",
        data=json.dumps(body).encode("utf-8"),
        headers={"content-type": "application/json",
                 "x-api-key": settings.anthropic_api_key,
                 "anthropic-version": "2023-06-01"})
    try:
        with urllib.request.urlopen(req, timeout=timeout) as r:
            resp = json.loads(r.read().decode("utf-8"))
    except Exception as e:
        raise ValueError(_loi_ai_ro_rang(e))
    txt = "".join(b.get("text", "") for b in resp.get("content", []) if b.get("type") == "text")
    return txt.strip().removeprefix("```json").removeprefix("```").removesuffix("```").strip()


def _vot_json_mang(txt: str) -> list:
    """Tách mảng JSON từ câu trả lời AI. Nếu mảng bị cắt cụt (trả lời dài quá
    giới hạn token), vẫn vớt các object phẳng hoàn chỉnh, bỏ phần tử dở dang."""
    import re as _re
    m = _re.search(r"\[[\s\S]*\]", txt)
    try:
        ds = json.loads(m.group(0) if m else txt)
        return ds if isinstance(ds, list) else [ds]
    except Exception:
        pass
    ds = []
    for mo in _re.finditer(r"\{[^{}]*\}", txt):
        try:
            ds.append(json.loads(mo.group(0)))
        except Exception:
            continue
    return ds


def doc_thong_tin_ncc(data: bytes, content_type: str, filename: str) -> list[dict]:
    """Đọc file báo giá / hồ sơ / danh sách NCC — file có thể chứa MỘT hoặc NHIỀU
    nhà cung cấp. Trả list dict {ten, ma_so_thue, dien_thoai, email, dia_chi, ghi_chu}."""
    if settings.ai_provider.upper() != "ANTHROPIC" or not settings.anthropic_api_key:
        raise ValueError("Chưa bật AI — cần AI_PROVIDER=ANTHROPIC và ANTHROPIC_API_KEY trên máy chủ.")
    khoi = _khoi_noi_dung_file(data, content_type, filename)
    sys = ("Bạn là trợ lý mua hàng của công ty xử lý nước SVWS. Người dùng gửi một file "
           "báo giá / hồ sơ / danh bạ nhà cung cấp — file có thể chứa MỘT hoặc NHIỀU nhà cung cấp. "
           "Hãy trích thông tin của TỪNG nhà cung cấp xuất hiện trong file "
           "(công ty phát hành báo giá hoặc được giới thiệu — KHÔNG phải khách nhận báo giá). "
           "CHỈ trả về đúng một MẢNG JSON, không thêm chữ nào khác, mỗi phần tử dạng: "
           '{"ten":"<tên công ty>","ma_so_thue":"<MST hoặc null>","dien_thoai":"<SĐT hoặc null>",'
           '"email":"<email hoặc null>","dia_chi":"<địa chỉ hoặc null>",'
           '"ghi_chu":"<người liên hệ, website, ngành hàng... nếu có, hoặc null>"} '
           "Mỗi công ty chỉ xuất hiện một lần. KHÔNG bịa thông tin không có trong file; "
           "thiếu trường nào để null trường đó.")
    txt = _goi_claude_json(khoi, sys,
                           "Trích danh sách nhà cung cấp trong file này thành mảng JSON.", 8000, 180)
    ds = _vot_json_mang(txt)
    if not ds:
        raise ValueError("AI không đọc được thông tin NCC trong file — kiểm tra lại nội dung. "
                         "Nếu file là ảnh scan mờ, thử file rõ nét hơn hoặc PDF gốc.")
    out = []
    for info in ds if isinstance(ds, list) else []:
        if not isinstance(info, dict) or not str(info.get("ten") or "").strip():
            continue
        item = {}
        for k, gh in (("ten", 200), ("ma_so_thue", 20), ("dien_thoai", 30),
                      ("email", 120), ("dia_chi", 300), ("ghi_chu", 1000)):
            v = info.get(k)
            item[k] = (str(v).strip()[:gh] if v else None)
        out.append(item)
    return out


def doc_bao_gia_file(data: bytes, content_type: str, filename: str) -> list[dict]:
    """Đọc file báo giá của NCC (PDF/ảnh/text-CSV) bằng Claude, trích danh sách sản phẩm.
    Trả list các dict {ten, ma_sp, mo_ta, nha_san_xuat, don_vi, don_gia, so_luong}.
    Ném ValueError nếu chưa bật AI / định dạng không hỗ trợ / đọc thất bại."""
    if settings.ai_provider.upper() != "ANTHROPIC" or not settings.anthropic_api_key:
        raise ValueError("Chưa bật AI — cần AI_PROVIDER=ANTHROPIC và ANTHROPIC_API_KEY trên máy chủ.")
    import urllib.request
    khoi = _khoi_noi_dung_file(data, content_type, filename)
    sys = ("Bạn là trợ lý mua hàng của công ty xử lý nước SVWS. Người dùng gửi một FILE BÁO GIÁ "
           "của nhà cung cấp. Hãy đọc và trích TOÀN BỘ các dòng sản phẩm/dịch vụ thành JSON. "
           "CHỈ trả về đúng một mảng JSON, không thêm chữ nào khác, mỗi phần tử dạng: "
           '{"ten":"<tên sản phẩm>","ma_sp":"<mã SP hoặc null>","mo_ta":"<mô tả/quy cách hoặc null>",'
           '"nha_san_xuat":"<hãng sản xuất/xuất xứ hoặc null>","don_vi":"<đơn vị tính hoặc null>",'
           '"don_gia":<đơn giá VND dạng số, hoặc null>,"so_luong":<số lượng trong báo giá dạng số, hoặc null>}. '
           "Đơn giá: bỏ dấu chấm/phẩy ngăn cách nghìn, quy về số VND; nếu giá bằng ngoại tệ thì để null. "
           "KHÔNG bịa thông tin không có trong file; thiếu trường nào để null trường đó.")
    body = {"model": settings.anthropic_model, "max_tokens": 8000, "system": sys,
            "messages": [{"role": "user", "content": [
                khoi,
                {"type": "text", "text": "Trích danh sách sản phẩm từ báo giá này thành mảng JSON."}]}]}
    req = urllib.request.Request(
        "https://api.anthropic.com/v1/messages",
        data=json.dumps(body).encode("utf-8"),
        headers={"content-type": "application/json",
                 "x-api-key": settings.anthropic_api_key,
                 "anthropic-version": "2023-06-01"})
    try:
        with urllib.request.urlopen(req, timeout=120) as r:
            resp = json.loads(r.read().decode("utf-8"))
    except Exception as e:
        raise ValueError(_loi_ai_ro_rang(e))
    txt = "".join(b.get("text", "") for b in resp.get("content", []) if b.get("type") == "text")
    txt = txt.strip().removeprefix("```json").removeprefix("```").removesuffix("```").strip()
    items = _vot_json_mang(txt)
    if not items:
        raise ValueError("AI không đọc được cấu trúc báo giá — kiểm tra file có bảng sản phẩm rõ ràng.")
    out = []
    for it in items if isinstance(items, list) else []:
        ten = str(it.get("ten") or "").strip()
        if not ten:
            continue
        gia = it.get("don_gia")
        try:
            gia = round(float(gia)) if gia is not None else None
        except (TypeError, ValueError):
            gia = None
        sl = it.get("so_luong")
        try:
            sl = float(sl) if sl is not None and float(sl) > 0 else None
        except (TypeError, ValueError):
            sl = None
        out.append({"ten": ten[:250], "so_luong": sl,
                    "ma_sp": (str(it.get("ma_sp")).strip()[:60] if it.get("ma_sp") else None),
                    "mo_ta": (str(it.get("mo_ta")).strip() if it.get("mo_ta") else None),
                    "nha_san_xuat": (str(it.get("nha_san_xuat")).strip()[:150] if it.get("nha_san_xuat") else None),
                    "don_vi": (str(it.get("don_vi")).strip()[:30] if it.get("don_vi") else None),
                    "don_gia": gia})
    return out


# ============ AI SOURCING AGENT — tự tìm/khuyến nghị NCC cho 1 đề xuất ============
def _ung_vien_nganh(ten: str):
    t = (ten or "").lower()
    g = []
    if any(k in t for k in ["hóa chất", "hoa chat", "pac", "pam", "clo", "polymer", "khử", "naoh", "phèn"]):
        g.append(("Nhà phân phối hóa chất xử lý nước", "hoa_chat",
                  "Kiểm chứng: GPKD hóa chất, MSDS/COA, nguồn gốc lô hàng."))
    if any(k in t for k in ["bơm", "bom", "pump"]):
        g.append(("Đại lý bơm công nghiệp (Ebara/Grundfos/Pentax/Wilo)", "thiet_bi",
                  "Kiểm chứng: chính hãng, bảo hành, sẵn hàng, đường cong bơm."))
    if any(k in t for k in ["màng", "mang", "mbr", "ro", "uf", "nf", "swro"]):
        g.append(("Nhà cung cấp màng lọc (Toray/DOW/Hydranautics/LG)", "mang",
                  "Kiểm chứng: chính hãng, model, lead time, điều kiện bảo quản."))
    if any(k in t for k in ["van", "valve", "ống", "ong", "phụ kiện", "phu kien"]):
        g.append(("Nhà cung cấp van & vật tư đường ống", "vat_tu",
                  "Kiểm chứng: tiêu chuẩn vật liệu, áp lực làm việc, chứng chỉ."))
    if not g:
        g.append(("Nhà cung cấp vật tư xử lý nước tổng hợp", "khac",
                  "Kiểm chứng năng lực, hồ sơ pháp lý và báo giá."))
    return [{"ten": x[0], "loai": x[1], "ghi_chu": x[2]} for x in g]


def _ai_source_fake(ten_hang, so_luong, ung_vien):
    trong = [u for u in ung_vien if u.get("trong_han_muc")]
    best = (trong or ung_vien)[0] if ung_vien else None
    rui_ro = []
    if best and not best.get("trong_han_muc"):
        rui_ro.append("NCC điểm cao nhất đang vượt/sát hạn mức công nợ — cân nhắc thanh toán bớt trước.")
    if len(ung_vien) <= 1:
        rui_ro.append("Chỉ có 1 nguồn cung nội bộ — rủi ro phụ thuộc; nên mời thêm báo giá.")
    over = [u["ten"] for u in ung_vien if not u.get("trong_han_muc")]
    if over:
        rui_ro.append("NCC rẻ nhưng vượt hạn mức công nợ: " + ", ".join(over) + ".")
    ly_do = ""
    if best:
        parts = []
        if best.get("gia_dung") or best.get("gia_gan_nhat"):
            _g = best.get("gia_dung") or best.get("gia_gan_nhat")
            _nhan = "báo giá" if best.get("nguon_gia") == "BAO_GIA" else "giá gần nhất"
            _hl = f", HL đến {best['hieu_luc_den']}" if best.get("hieu_luc_den") else ""
            parts.append(f"{_nhan} {int(_g):,}".replace(",", ".") + " đ" + _hl)
        if best.get("diem_danh_gia"):
            parts.append(f"đánh giá {best['diem_danh_gia']:.1f}/5")
        if best.get("ty_le_dung_han") is not None:
            parts.append(f"giao đúng hạn {round(best['ty_le_dung_han']*100)}%")
        ly_do = f"Đề xuất chọn {best['ten']} (điểm {best.get('diem_tong')}): " + ", ".join(parts) + "."
    top = [u["ten"] for u in ung_vien[:3]]
    hanh_dong = ("Gửi yêu cầu báo giá (RFQ) tới: " + ", ".join(top) + ".") if top \
        else "Bổ sung hồ sơ NCC cho mặt hàng này rồi xin báo giá."
    return {"khuyen_nghi_ncc_id": best["nha_cung_cap_id"] if best else None,
            "ten_khuyen_nghi": best["ten"] if best else None,
            "ly_do": ly_do, "rui_ro": rui_ro, "hanh_dong": hanh_dong,
            "ung_vien_moi": _ung_vien_nganh(ten_hang), "nguon": "DEMO"}


def _ai_source_anthropic(ten_hang, so_luong, ung_vien):
    import json as _json, urllib.request
    sys = ("Bạn là chuyên viên thu mua (sourcing) cho công ty xử lý nước SVWS. "
           "Dựa trên DANH SÁCH NCC nội bộ đã chấm điểm, hãy khuyến nghị lựa chọn và cảnh báo rủi ro. "
           "CHỈ chọn khuyen_nghi_ncc_id trong danh sách cung cấp (không bịa). "
           "ung_vien_moi chỉ là nhóm/thương hiệu gợi ý để tìm thêm báo giá, KHÔNG bịa tên công ty/SĐT cụ thể. "
           "CHỈ trả JSON: {\"khuyen_nghi_ncc_id\":<id|null>,\"ten_khuyen_nghi\":\"\",\"ly_do\":\"\","
           "\"rui_ro\":[\"\"],\"hanh_dong\":\"\",\"ung_vien_moi\":[{\"ten\":\"\",\"loai\":\"\",\"ghi_chu\":\"\"}]}")
    user = (f"Mặt hàng: {ten_hang}\nSố lượng: {so_luong}\n"
            f"NCC nội bộ (đã chấm điểm):\n{_json.dumps(ung_vien, ensure_ascii=False)}")
    body = {"model": settings.anthropic_model, "max_tokens": 1024, "system": sys,
            "messages": [{"role": "user", "content": user}]}
    req = urllib.request.Request("https://api.anthropic.com/v1/messages",
                                 data=_json.dumps(body).encode("utf-8"),
                                 headers={"content-type": "application/json",
                                          "x-api-key": settings.anthropic_api_key,
                                          "anthropic-version": "2023-06-01"})
    with urllib.request.urlopen(req, timeout=30) as r:
        data = _json.loads(r.read().decode("utf-8"))
    txt = "".join(b.get("text", "") for b in data.get("content", []) if b.get("type") == "text")
    txt = txt.strip().removeprefix("```json").removeprefix("```").removesuffix("```").strip()
    out = _json.loads(txt)
    ids = {u["nha_cung_cap_id"] for u in ung_vien}
    if out.get("khuyen_nghi_ncc_id") not in ids:
        out["khuyen_nghi_ncc_id"] = ung_vien[0]["nha_cung_cap_id"] if ung_vien else None
    out.setdefault("rui_ro", []); out.setdefault("ung_vien_moi", []); out["nguon"] = "ANTHROPIC"
    return out


def goi_y_ncc_ai(ten_hang, so_luong, ung_vien):
    if settings.ai_provider.upper() == "ANTHROPIC" and settings.anthropic_api_key:
        try:
            return _ai_source_anthropic(ten_hang, so_luong, ung_vien)
        except Exception:
            pass
    return _ai_source_fake(ten_hang, so_luong, ung_vien)


# ============ DÒ NCC MỚI TRÊN WEB (có kiểm chứng) ============
def _tim_ncc_web_anthropic(ten_hang, khu_vuc):
    import json as _json, re as _re, urllib.request
    sys = ("Bạn là chuyên viên thu mua. Dùng web search tìm nhà cung cấp khả dĩ cho mặt hàng "
           "tại khu vực yêu cầu. CHỈ liệt kê NCC CÓ THẬT tìm được trên web, kèm nguon_url; KHÔNG bịa. "
           "Trả về CUỐI CÙNG đúng một JSON: "
           "{\"ung_vien\":[{\"ten\":\"\",\"website\":\"\",\"khu_vuc\":\"\",\"ghi_chu\":\"\",\"nguon_url\":\"\"}]}. "
           "Tối đa 6 ứng viên.")
    body = {"model": settings.anthropic_model, "max_tokens": 1500, "system": sys,
            "tools": [{"type": settings.web_search_tool, "name": "web_search"}],
            "messages": [{"role": "user",
                          "content": f"Mặt hàng: {ten_hang}\nKhu vực: {khu_vuc}\nTìm nhà cung cấp."}]}
    req = urllib.request.Request("https://api.anthropic.com/v1/messages",
                                 data=_json.dumps(body).encode("utf-8"),
                                 headers={"content-type": "application/json",
                                          "x-api-key": settings.anthropic_api_key,
                                          "anthropic-version": "2023-06-01"})
    with urllib.request.urlopen(req, timeout=60) as r:
        data = _json.loads(r.read().decode("utf-8"))
    txt = "".join(b.get("text", "") for b in data.get("content", []) if b.get("type") == "text")
    m = _re.search(r"\{[\s\S]*\}", txt)
    obj = _json.loads(m.group(0) if m else txt)
    uv = obj.get("ung_vien", [])
    for u in uv:
        u["kiem_chung"] = False
    return {"nguon": "ANTHROPIC", "kha_dung": True, "ung_vien": uv}


def tim_ncc_web(ten_hang, khu_vuc="Việt Nam"):
    """Dò NCC mới trên web qua công cụ web_search của Claude. Kết quả CẦN KIỂM CHỨNG trước khi dùng."""
    if settings.ai_provider.upper() != "ANTHROPIC" or not settings.anthropic_api_key:
        return {"nguon": "DEMO", "kha_dung": False,
                "thong_bao": ("Đang ở chế độ DEMO. Bật AI_PROVIDER=ANTHROPIC + ANTHROPIC_API_KEY "
                              "(tài khoản có web search) để dò NCC mới trên web. Dưới đây là nhóm nguồn nên tìm:"),
                "ung_vien": [{"ten": u["ten"], "website": "", "khu_vuc": khu_vuc,
                              "ghi_chu": u["ghi_chu"], "nguon_url": "", "kiem_chung": False}
                             for u in _ung_vien_nganh(ten_hang)]}
    try:
        return _tim_ncc_web_anthropic(ten_hang, khu_vuc)
    except Exception as e:
        return {"nguon": "ANTHROPIC", "kha_dung": False,
                "thong_bao": f"Không gọi được web search ({type(e).__name__}). Tạm dùng nhóm nguồn gợi ý:",
                "ung_vien": [{"ten": u["ten"], "website": "", "khu_vuc": khu_vuc,
                              "ghi_chu": u["ghi_chu"], "nguon_url": "", "kiem_chung": False}
                             for u in _ung_vien_nganh(ten_hang)]}


# ============ AI CỐ VẤN TÀI CHÍNH (CFO ảo) — phân tích chỉ số + cảnh báo ============
def _fmt_vnd(x):
    try:
        return f"{float(x):,.0f}đ"
    except Exception:
        return str(x)


def _tu_van_tai_chinh_fake(payload: dict) -> dict:
    cs = payload.get("chi_so", {})
    cbs = payload.get("canh_bao", [])
    cao = [c for c in cbs if c.get("muc_do") == "CAO"]
    trung = [c for c in cbs if c.get("muc_do") == "TRUNG"]
    if cao:
        suc_khoe = "YEU"
    elif len(trung) >= 2:
        suc_khoe = "TRUNG_BINH"
    elif trung:
        suc_khoe = "KHA"
    else:
        suc_khoe = "TOT"
    nhan_dinh = []
    cr = cs.get("ty_so_thanh_toan_hien_hanh")
    if cr is not None:
        nhan_dinh.append(
            f"Hệ số thanh toán hiện hành {cr:.2f} — "
            + ("an toàn (>1,5)." if cr >= 1.5 else "ở mức trung bình." if cr >= 1 else "DƯỚI 1, rủi ro thanh khoản."))
    bg = cs.get("bien_loi_nhuan_gop")
    if bg is not None:
        nhan_dinh.append(f"Biên lợi nhuận gộp {bg*100:.1f}% — "
                         + ("tốt." if bg >= 0.25 else "cần cải thiện." if bg >= 0.1 else "quá mỏng."))
    dso = cs.get("ky_thu_tien_bq")
    if dso is not None:
        nhan_dinh.append(f"Kỳ thu tiền bình quân {dso:.0f} ngày — "
                         + ("tốt." if dso <= 30 else "hơi dài." if dso <= 60 else "quá dài, vốn bị chiếm dụng."))
    rw = cs.get("so_thang_tien_mat_con_lai")
    if rw is not None:
        nhan_dinh.append(f"Tiền mặt đủ duy trì ~{rw:.1f} tháng theo nhịp chi hiện tại.")
    uu_tien = []
    for c in cao + trung:
        uu_tien.append({"tieu_de": c.get("tieu_de"), "muc_do": c.get("muc_do"),
                        "hanh_dong": c.get("goi_y", "Rà soát và xử lý.")})
    if not uu_tien:
        uu_tien.append({"tieu_de": "Duy trì kỷ luật tài chính", "muc_do": "THAP",
                        "hanh_dong": "Tiếp tục theo dõi công nợ và dòng tiền hằng tuần."})
    dg = (f"Sức khỏe tài chính tổng thể: {suc_khoe.replace('_',' ').lower()}. "
          + ("Có rủi ro cần xử lý ngay. " if cao else "Chưa có rủi ro nghiêm trọng. " if not trung else "Một số điểm cần lưu ý. ")
          + " ".join(nhan_dinh[:3]))
    return {"nguon": "DEMO", "suc_khoe": suc_khoe, "danh_gia": dg,
            "nhan_dinh": nhan_dinh, "uu_tien": uu_tien}


def _tu_van_tai_chinh_anthropic(payload: dict) -> dict:
    import urllib.request
    sys = (
        "Bạn là Giám đốc Tài chính (CFO) giàu kinh nghiệm của một doanh nghiệp kỹ thuật "
        "xử lý nước (~20 nhân sự) tại Việt Nam. Phân tích các chỉ số tài chính được cung cấp, "
        "đưa nhận định sắc bén, thực tế, và khuyến nghị hành động ưu tiên. Dùng chuẩn mực VAS. "
        "CHỈ trả về JSON, không thêm chữ nào khác, dạng: "
        '{"suc_khoe":"<TOT|KHA|TRUNG_BINH|YEU>",'
        '"danh_gia":"<đánh giá tổng thể 3-5 câu tiếng Việt>",'
        '"nhan_dinh":["<gạch đầu dòng nhận định>", "..."],'
        '"uu_tien":[{"tieu_de":"<vấn đề>","muc_do":"<CAO|TRUNG|THAP>","hanh_dong":"<khuyến nghị cụ thể>"}]}')
    body = {"model": settings.anthropic_model, "max_tokens": 1500, "system": sys,
            "messages": [{"role": "user", "content": json.dumps(payload, ensure_ascii=False)}]}
    req = urllib.request.Request(
        "https://api.anthropic.com/v1/messages",
        data=json.dumps(body).encode("utf-8"),
        headers={"content-type": "application/json", "x-api-key": settings.anthropic_api_key,
                 "anthropic-version": "2023-06-01"})
    try:
        with urllib.request.urlopen(req, timeout=40) as r:
            data = json.loads(r.read().decode("utf-8"))
        txt = "".join(b.get("text", "") for b in data.get("content", []) if b.get("type") == "text")
        txt = txt.strip().removeprefix("```json").removeprefix("```").removesuffix("```").strip()
        out = json.loads(txt)
        out["nguon"] = "ANTHROPIC"
        out.setdefault("suc_khoe", "TRUNG_BINH"); out.setdefault("danh_gia", "")
        out.setdefault("nhan_dinh", []); out.setdefault("uu_tien", [])
        return out
    except Exception:
        return _tu_van_tai_chinh_fake(payload)


def tu_van_tai_chinh(payload: dict) -> dict:
    """Cố vấn tài chính AI: nhận {chi_so, canh_bao} → trả phân tích + khuyến nghị."""
    if settings.ai_provider.upper() == "ANTHROPIC":
        return _tu_van_tai_chinh_anthropic(payload)
    return _tu_van_tai_chinh_fake(payload)


# ============ AGENT PHÂN TÍCH NGUYÊN LÝ THIẾT KẾ ============
def _pt_thieu(info):
    thieu = []
    if not (info.get("mo_ta") or "").strip():
        thieu.append("Mô tả dự án (nguồn thải, ngành, mục tiêu xả/tái sử dụng)")
    cts = info.get("chi_tieu") or []
    if not cts:
        thieu.append("Chỉ tiêu chất lượng đầu vào/đầu ra")
    else:
        if not any(c.get("gia_tri_vao") is not None for c in cts):
            thieu.append("Giá trị chất lượng ĐẦU VÀO (chưa nhập số nào)")
        if not any(c.get("gioi_han_ra") is not None for c in cts):
            thieu.append("Giới hạn ĐẦU RA theo tiêu chuẩn")
    if not (info.get("cong_suat") or "").strip():
        thieu.append("Công suất/lưu lượng thiết kế")
    return thieu


def _phan_tich_heuristic(info):
    """Phân tích sơ bộ DỰA TRÊN dữ liệu đã nhập (không bịa số). Quy tắc kỹ thuật chung, cần kỹ sư thẩm định."""
    cts = info.get("chi_tieu") or []
    loai = (info.get("loai_du_an") or "").upper()
    ten = lambda c: (c.get("ten") or "").lower()
    def co(*keys):  # chỉ tiêu có giá trị đầu vào và cần xử lý cao
        for c in cts:
            if any(k in ten(c) for k in keys) and (c.get("can_xu_ly") or 0) >= 40:
                return True
        return False
    ng, sodo = [], []
    if loai == "KHI_THAI":
        if co("bụi", "pm"): ng.append("Tách bụi: cyclone → lọc túi vải/ESP tùy nồng độ và nhiệt độ khí."); sodo.append("Chụp hút → tách bụi")
        if co("so2", "lưu huỳnh"): ng.append("Khử SO₂ bằng hấp thụ kiềm (vôi/NaOH) – tháp hấp thụ ướt.")
        if co("nox", "nitơ oxit"): ng.append("Khử NOx bằng SNCR/SCR (xúc tác) tùy nhiệt độ và yêu cầu.")
        if co("voc", "hữu cơ", "toluene", "xylene"): ng.append("Xử lý VOC: hấp phụ than hoạt tính hoặc thiêu đốt/RTO.")
        if not ng: ng.append("Chưa đủ chỉ tiêu vượt ngưỡng để khuyến nghị; cần bổ sung dữ liệu đo khí thực tế.")
        sodo += ["Tháp hấp thụ/khử", "Quạt hút – ống khói (kèm quan trắc)"]
    else:  # nước
        cao_mau = co("màu"); cao_codbod = co("cod", "bod"); cao_tss = co("tss", "rắn lơ lửng")
        cao_np = co("nitơ", "photpho", "amoni", "phosphat"); cao_kl = co("crom", "chì", "đồng", "kẽm", "niken", "asen", "thủy ngân")
        if cao_tss: ng.append("Tách rắn đầu vào: song chắn → lắng cát → keo tụ/tạo bông → lắng hoặc tuyển nổi (DAF).")
        if cao_kl: ng.append("Kim loại nặng: kết tủa hóa học theo pH (hydroxit/sunfua) trước xử lý sinh học.")
        if cao_codbod: ng.append("Hữu cơ (COD/BOD): xử lý sinh học – kỵ khí (UASB) nếu tải cao, sau đó hiếu khí (Aerotank/MBR).")
        if cao_mau: ng.append("Độ màu (đặc trưng dệt nhuộm): keo tụ + oxy hóa nâng cao (Fenton/O₃) và/hoặc hấp phụ than hoạt tính.")
        if cao_np: ng.append("Dinh dưỡng N/P: bố trí thiếu khí–hiếu khí (nitrat hóa/khử nitrat) và khử P (sinh học hoặc hóa học).")
        # tái sử dụng / giới hạn rất chặt
        if "tái sử dụng" in (info.get("mo_ta") or "").lower() or "ro" in (info.get("mo_ta") or "").lower():
            ng.append("Mục tiêu tái sử dụng: bổ sung lọc tinh + màng (UF → RO/NF); xử lý dòng thải đậm (reject) bằng AOP.")
        if not ng: ng.append("Chưa đủ chỉ tiêu vượt ngưỡng để khuyến nghị; cần nhập giá trị đầu vào/đầu ra.")
        sodo = ["Tiền xử lý (tách rắn/điều hòa)", "Xử lý sinh học (kỵ khí/hiếu khí)"]
        if cao_mau: sodo.append("Oxy hóa nâng cao/khử màu")
        sodo += ["Lắng/lọc hoàn thiện", "Khử trùng / xả hoặc tái sử dụng"]
    return {
        "tom_tat": f"Phân tích sơ bộ cho dự án {info.get('loai_du_an') or ''} (công suất: {info.get('cong_suat') or 'chưa rõ'}). "
                   f"Dựa trên {sum(1 for c in cts if c.get('gia_tri_vao') is not None)} chỉ tiêu có số liệu đầu vào.",
        "nguyen_ly": ng,
        "so_do_cong_nghe": sodo,
        "can_kiem_chung": ["Cân bằng vật chất theo lưu lượng & tải lượng thực tế",
                            "Jar test/bench test xác định hóa chất & liều lượng",
                            "Kiểm tra tính khả thi mặt bằng và chi phí vận hành"],
        "du_lieu_thieu": _pt_thieu(info),
        "nguon": "HEURISTIC",
        "luu_y": "Đây là gợi ý sơ bộ theo quy tắc kỹ thuật chung, KHÔNG thay thế thiết kế. Cần kỹ sư thẩm định và thí nghiệm thực tế.",
    }


def _phan_tich_anthropic(info):
    import json as _json, urllib.request
    sys = ("Bạn là kỹ sư thiết kế hệ thống xử lý nước/khí thải của công ty SVWS. "
           "Nhiệm vụ: phân tích MÔ TẢ và DỮ LIỆU CHẤT LƯỢNG do người dùng cung cấp để đề xuất NGUYÊN LÝ THIẾT KẾ và SƠ ĐỒ CÔNG NGHỆ. "
           "RÀNG BUỘC: chỉ dựa trên dữ liệu được cung cấp; KHÔNG bịa số liệu, KHÔNG tự đưa ra giá trị giới hạn quy chuẩn nếu không có trong dữ liệu; "
           "nếu thiếu dữ liệu thì nêu rõ ở 'du_lieu_thieu'. Đề xuất mang tính định hướng, cần kỹ sư thẩm định. Trả lời bằng tiếng Việt. "
           "CHỈ trả JSON đúng dạng: {\"tom_tat\":\"\",\"nguyen_ly\":[\"\"],\"so_do_cong_nghe\":[\"\"],\"can_kiem_chung\":[\"\"],\"du_lieu_thieu\":[\"\"]}")
    user = ("Loại dự án: " + str(info.get("loai_du_an")) + "\n"
            "Công suất: " + str(info.get("cong_suat")) + "\n"
            "Tiêu chuẩn đầu ra: " + str(info.get("tieu_chuan_dau_ra")) + "\n"
            "Mô tả: " + str(info.get("mo_ta") or "(trống)") + "\n"
            "Chỉ tiêu chất lượng (đầu vào / giới hạn ra / % cần xử lý):\n"
            + _json.dumps(info.get("chi_tieu") or [], ensure_ascii=False))
    body = {"model": settings.anthropic_model, "max_tokens": 1500, "system": sys,
            "messages": [{"role": "user", "content": user}]}
    req = urllib.request.Request("https://api.anthropic.com/v1/messages",
                                 data=_json.dumps(body).encode("utf-8"),
                                 headers={"content-type": "application/json",
                                          "x-api-key": settings.anthropic_api_key,
                                          "anthropic-version": "2023-06-01"})
    with urllib.request.urlopen(req, timeout=45) as r:
        data = _json.loads(r.read().decode("utf-8"))
    txt = "".join(b.get("text", "") for b in data.get("content", []) if b.get("type") == "text")
    txt = txt.strip().removeprefix("```json").removeprefix("```").removesuffix("```").strip()
    out = _json.loads(txt)
    for k in ("nguyen_ly", "so_do_cong_nghe", "can_kiem_chung", "du_lieu_thieu"):
        out.setdefault(k, [])
    out.setdefault("tom_tat", "")
    out["nguon"] = "ANTHROPIC"
    out["luu_y"] = "Đề xuất do AI phân tích, mang tính định hướng — cần kỹ sư thẩm định và thí nghiệm thực tế."
    return out


def phan_tich_thiet_ke(info):
    if settings.ai_provider.upper() == "ANTHROPIC" and settings.anthropic_api_key:
        try:
            return _phan_tich_anthropic(info)
        except Exception:
            pass
    return _phan_tich_heuristic(info)
