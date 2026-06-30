"""
Mô tả công việc (JD) · KPI theo vị trí · Đánh giá kết quả theo kỳ (tuần/tháng/quý/năm).
Thuộc module RBAC 'nhan_su'. Xem: XEM · Đánh giá/nhập KPI: THAO_TAC · Sửa JD/KPI: DUYET.
"""
import json
from decimal import Decimal
from datetime import datetime
from fastapi import APIRouter, Depends, HTTPException, status
from pydantic import BaseModel, Field
from sqlalchemy.orm import Session
from sqlalchemy import func
from ..database import get_db
from ..rbac import yeu_cau
from ..deps import nhan_vien_id_cua
from ..audit import ghi_audit
from ..models import (NguoiDung, NhanVien, VaiTro, MoTaCongViec, KpiViTri,
                      DanhGiaKy, DanhGiaCt, BangLuong, KyLuong, CfgThuongKpi)
from fastapi.responses import StreamingResponse
import io

router = APIRouter(prefix="/nhan-su", tags=["nhan_su_kpi"])
MODULE = "nhan_su"

LOAI_KY = {"TUAN", "THANG", "QUY", "NAM"}


def _loads(s, mac_dinh):
    try:
        return json.loads(s) if s else mac_dinh
    except Exception:
        return mac_dinh


def _vai_tro_nv(db: Session, nv: NhanVien) -> str | None:
    """Suy ra mã vị trí của nhân viên qua tài khoản đăng nhập."""
    if nv.nguoi_dung_id:
        u = db.get(NguoiDung, nv.nguoi_dung_id)
        if u:
            vt = db.get(VaiTro, u.vai_tro_id)
            return vt.ma if vt else None
    return None


def _xep_loai(diem: float) -> str:
    if diem >= 85:
        return "A"
    if diem >= 70:
        return "B"
    if diem >= 50:
        return "C"
    return "D"


def _tinh(chieu: str, muc_tieu, thuc, trong_so) -> tuple[float | None, float]:
    """Trả (phần_trăm_đạt, điểm). Điểm = trọng_số × min(%đạt,100)/100."""
    ts = float(trong_so or 0)
    if thuc is None:
        return None, 0.0
    thuc = float(thuc)
    if muc_tieu is None:
        pct = 100.0  # không có mục tiêu số → coi như đạt khi có nhập
    else:
        mt = float(muc_tieu)
        if chieu == "THAP":          # càng thấp càng tốt
            if mt == 0:
                pct = 100.0 if thuc <= 0 else max(0.0, 100.0 - thuc * 25.0)
            else:
                pct = 100.0 if thuc <= mt else (mt / thuc * 100.0)
        else:                         # CAO: càng cao càng tốt
            pct = 100.0 if mt == 0 else (thuc / mt * 100.0)
    pct_hien = round(min(pct, 200.0), 2)
    diem = round(ts * min(pct, 100.0) / 100.0, 2)
    return pct_hien, diem


# ===================== MÔ TẢ CÔNG VIỆC (JD) =====================
@router.get("/mo-ta-cv")
def ds_mo_ta(db: Session = Depends(get_db), _=Depends(yeu_cau(MODULE, "XEM"))):
    ten_vt = {v.ma: v.ten for v in db.query(VaiTro).all()}
    out = []
    for j in db.query(MoTaCongViec).all():
        sl_kpi = db.query(func.count(KpiViTri.id)).filter_by(vai_tro=j.vai_tro).scalar()
        out.append({"vai_tro": j.vai_tro, "chuc_danh": j.chuc_danh, "cap_bac": j.cap_bac,
                    "bao_cao_cho": j.bao_cao_cho, "ten_vai_tro": ten_vt.get(j.vai_tro, j.vai_tro),
                    "so_kpi": sl_kpi or 0})
    # giữ thứ tự theo id vai_tro
    thu_tu = {v.ma: v.id for v in db.query(VaiTro).all()}
    out.sort(key=lambda x: thu_tu.get(x["vai_tro"], 999))
    return out


@router.get("/mo-ta-cv/{vai_tro}")
def chi_tiet_mo_ta(vai_tro: str, db: Session = Depends(get_db), _=Depends(yeu_cau(MODULE, "XEM"))):
    j = db.query(MoTaCongViec).filter_by(vai_tro=vai_tro).first()
    if j is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Chưa có mô tả công việc cho vị trí này")
    vt = db.query(VaiTro).filter_by(ma=vai_tro).first()
    kpis = db.query(KpiViTri).filter_by(vai_tro=vai_tro).order_by(KpiViTri.thu_tu, KpiViTri.id).all()
    return {
        "vai_tro": j.vai_tro, "ten_vai_tro": vt.ten if vt else vai_tro,
        "chuc_danh": j.chuc_danh, "cap_bac": j.cap_bac, "bao_cao_cho": j.bao_cao_cho,
        "muc_dich": j.muc_dich,
        "trach_nhiem": _loads(j.trach_nhiem, []),
        "quyen_han": _loads(j.quyen_han, []),
        "yeu_cau": _loads(j.yeu_cau, {}),
        "cap_nhat_luc": str(j.cap_nhat_luc) if j.cap_nhat_luc else None,
        "kpis": [{"id": k.id, "ten": k.ten, "don_vi": k.don_vi, "trong_so": float(k.trong_so or 0),
                  "muc_tieu": float(k.muc_tieu) if k.muc_tieu is not None else None,
                  "chieu": k.chieu, "chu_ky": k.chu_ky, "mo_ta": k.mo_ta} for k in kpis],
    }


class MoTaSua(BaseModel):
    chuc_danh: str | None = None
    cap_bac: str | None = None
    bao_cao_cho: str | None = None
    muc_dich: str | None = None
    trach_nhiem: list[str] | None = None
    quyen_han: list[str] | None = None
    yeu_cau: dict | None = None


@router.put("/mo-ta-cv/{vai_tro}")
def sua_mo_ta(vai_tro: str, data: MoTaSua, db: Session = Depends(get_db),
              nd: NguoiDung = Depends(yeu_cau(MODULE, "DUYET"))):
    j = db.query(MoTaCongViec).filter_by(vai_tro=vai_tro).first()
    if j is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Không tìm thấy mô tả công việc")
    if data.chuc_danh is not None:
        j.chuc_danh = data.chuc_danh
    if data.cap_bac is not None:
        j.cap_bac = data.cap_bac
    if data.bao_cao_cho is not None:
        j.bao_cao_cho = data.bao_cao_cho
    if data.muc_dich is not None:
        j.muc_dich = data.muc_dich
    if data.trach_nhiem is not None:
        j.trach_nhiem = json.dumps(data.trach_nhiem, ensure_ascii=False)
    if data.quyen_han is not None:
        j.quyen_han = json.dumps(data.quyen_han, ensure_ascii=False)
    if data.yeu_cau is not None:
        j.yeu_cau = json.dumps(data.yeu_cau, ensure_ascii=False)
    j.cap_nhat_luc = datetime.now()
    ghi_audit(db, nd.id, "SUA", "mo_ta_cong_viec", j.id, moi={"vai_tro": vai_tro})
    db.commit()
    return {"vai_tro": vai_tro, "trang_thai": "DA_LUU"}


# ===================== KPI THEO VỊ TRÍ =====================
class KpiVao(BaseModel):
    vai_tro: str
    ten: str
    don_vi: str | None = None
    trong_so: Decimal = 0
    muc_tieu: Decimal | None = None
    chieu: str = Field(default="CAO", pattern="^(CAO|THAP)$")
    chu_ky: str = Field(default="THANG", pattern="^(TUAN|THANG|QUY|NAM)$")
    mo_ta: str | None = None


@router.post("/kpi", status_code=201)
def them_kpi(data: KpiVao, db: Session = Depends(get_db),
             nd: NguoiDung = Depends(yeu_cau(MODULE, "DUYET"))):
    n = db.query(func.count(KpiViTri.id)).filter_by(vai_tro=data.vai_tro).scalar()
    k = KpiViTri(vai_tro=data.vai_tro, ten=data.ten, don_vi=data.don_vi,
                 trong_so=data.trong_so, muc_tieu=data.muc_tieu, chieu=data.chieu,
                 chu_ky=data.chu_ky, mo_ta=data.mo_ta, thu_tu=n or 0)
    db.add(k); db.flush()
    ghi_audit(db, nd.id, "TAO", "kpi_vi_tri", k.id, moi={"vai_tro": data.vai_tro, "ten": data.ten})
    db.commit()
    return {"id": k.id}


class KpiSua(BaseModel):
    ten: str | None = None
    don_vi: str | None = None
    trong_so: Decimal | None = None
    muc_tieu: Decimal | None = None
    chieu: str | None = Field(default=None, pattern="^(CAO|THAP)$")
    chu_ky: str | None = Field(default=None, pattern="^(TUAN|THANG|QUY|NAM)$")
    mo_ta: str | None = None


@router.put("/kpi/{kpi_id}")
def sua_kpi(kpi_id: int, data: KpiSua, db: Session = Depends(get_db),
            nd: NguoiDung = Depends(yeu_cau(MODULE, "DUYET"))):
    k = db.get(KpiViTri, kpi_id)
    if k is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Không tìm thấy KPI")
    for f in ("ten", "don_vi", "trong_so", "muc_tieu", "chieu", "chu_ky", "mo_ta"):
        v = getattr(data, f)
        if v is not None:
            setattr(k, f, v)
    ghi_audit(db, nd.id, "SUA", "kpi_vi_tri", k.id)
    db.commit()
    return {"id": k.id, "trang_thai": "DA_LUU"}


@router.delete("/kpi/{kpi_id}")
def xoa_kpi(kpi_id: int, db: Session = Depends(get_db),
            nd: NguoiDung = Depends(yeu_cau(MODULE, "DUYET"))):
    k = db.get(KpiViTri, kpi_id)
    if k is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Không tìm thấy KPI")
    db.delete(k)
    ghi_audit(db, nd.id, "XOA", "kpi_vi_tri", kpi_id)
    db.commit()
    return {"id": kpi_id, "trang_thai": "DA_XOA"}


# ===================== ĐÁNH GIÁ THEO KỲ =====================
class DanhGiaCtVao(BaseModel):
    kpi_id: int
    gia_tri_thuc: Decimal | None = None


class DanhGiaVao(BaseModel):
    nhan_vien_id: int
    loai_ky: str = Field(pattern="^(TUAN|THANG|QUY|NAM)$")
    ky: str
    nhan_xet: str | None = None
    chi_tiet: list[DanhGiaCtVao] = Field(min_length=1)


@router.post("/danh-gia", status_code=201)
def tao_danh_gia(data: DanhGiaVao, db: Session = Depends(get_db),
                 nd: NguoiDung = Depends(yeu_cau(MODULE, "THAO_TAC"))):
    nv = db.get(NhanVien, data.nhan_vien_id)
    if nv is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Không tìm thấy nhân viên")
    # upsert: xóa bản đánh giá cũ cùng (nhân viên, loại kỳ, kỳ)
    cu = db.query(DanhGiaKy).filter_by(nhan_vien_id=data.nhan_vien_id,
                                       loai_ky=data.loai_ky, ky=data.ky).all()
    for c in cu:
        db.delete(c)
    db.flush()
    dg = DanhGiaKy(nhan_vien_id=data.nhan_vien_id, vai_tro=_vai_tro_nv(db, nv),
                   loai_ky=data.loai_ky, ky=data.ky, nhan_xet=data.nhan_xet,
                   nguoi_danh_gia=nhan_vien_id_cua(db, nd.id))
    db.add(dg); db.flush()
    tong = 0.0
    for ct in data.chi_tiet:
        k = db.get(KpiViTri, ct.kpi_id)
        if k is None:
            continue
        pct, diem = _tinh(k.chieu, k.muc_tieu, ct.gia_tri_thuc, k.trong_so)
        tong += diem
        db.add(DanhGiaCt(danh_gia_id=dg.id, kpi_id=k.id, ten=k.ten, trong_so=k.trong_so,
                         muc_tieu=k.muc_tieu, gia_tri_thuc=ct.gia_tri_thuc,
                         phan_tram_dat=pct, diem=diem))
    dg.tong_diem = round(tong, 2)
    dg.xep_loai = _xep_loai(tong)
    ghi_audit(db, nd.id, "TAO", "danh_gia_ky", dg.id,
              moi={"nhan_vien_id": data.nhan_vien_id, "ky": data.ky, "diem": dg.tong_diem})
    db.commit()
    return {"id": dg.id, "tong_diem": float(dg.tong_diem), "xep_loai": dg.xep_loai}


@router.get("/danh-gia")
def ds_danh_gia(loai_ky: str | None = None, ky: str | None = None,
                db: Session = Depends(get_db), _=Depends(yeu_cau(MODULE, "XEM"))):
    q = db.query(DanhGiaKy)
    if loai_ky:
        q = q.filter_by(loai_ky=loai_ky)
    if ky:
        q = q.filter_by(ky=ky)
    out = []
    ten_vt = {v.ma: v.ten for v in db.query(VaiTro).all()}
    for d in q.order_by(DanhGiaKy.ky.desc(), DanhGiaKy.tong_diem.desc()).all():
        nv = db.get(NhanVien, d.nhan_vien_id)
        out.append({"id": d.id, "nhan_vien_id": d.nhan_vien_id,
                    "ho_ten": nv.ho_ten if nv else "", "ma": nv.ma if nv else None,
                    "vai_tro": d.vai_tro, "ten_vai_tro": ten_vt.get(d.vai_tro, d.vai_tro),
                    "loai_ky": d.loai_ky, "ky": d.ky,
                    "tong_diem": float(d.tong_diem or 0), "xep_loai": d.xep_loai})
    return out


@router.get("/danh-gia/{dg_id}")
def chi_tiet_danh_gia(dg_id: int, db: Session = Depends(get_db), _=Depends(yeu_cau(MODULE, "XEM"))):
    d = db.get(DanhGiaKy, dg_id)
    if d is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Không tìm thấy bản đánh giá")
    nv = db.get(NhanVien, d.nhan_vien_id)
    return {
        "id": d.id, "nhan_vien_id": d.nhan_vien_id, "ho_ten": nv.ho_ten if nv else "",
        "vai_tro": d.vai_tro, "loai_ky": d.loai_ky, "ky": d.ky,
        "tong_diem": float(d.tong_diem or 0), "xep_loai": d.xep_loai, "nhan_xet": d.nhan_xet,
        "chi_tiet": [{"ten": c.ten, "trong_so": float(c.trong_so or 0),
                      "muc_tieu": float(c.muc_tieu) if c.muc_tieu is not None else None,
                      "gia_tri_thuc": float(c.gia_tri_thuc) if c.gia_tri_thuc is not None else None,
                      "phan_tram_dat": float(c.phan_tram_dat) if c.phan_tram_dat is not None else None,
                      "diem": float(c.diem or 0)} for c in d.chi_tiet],
    }


@router.get("/danh-gia-tong-ket")
def tong_ket(loai_ky: str | None = None, ky: str | None = None,
             db: Session = Depends(get_db), _=Depends(yeu_cau(MODULE, "XEM"))):
    q = db.query(DanhGiaKy)
    if loai_ky:
        q = q.filter_by(loai_ky=loai_ky)
    if ky:
        q = q.filter_by(ky=ky)
    rows = q.all()
    phan_bo = {"A": 0, "B": 0, "C": 0, "D": 0}
    tong = 0.0
    for r in rows:
        phan_bo[r.xep_loai or "D"] = phan_bo.get(r.xep_loai or "D", 0) + 1
        tong += float(r.tong_diem or 0)
    n = len(rows)
    return {"so_danh_gia": n, "diem_trung_binh": round(tong / n, 2) if n else None,
            "phan_bo": phan_bo,
            "cac_ky": sorted({r.ky for r in db.query(DanhGiaKy).all()}, reverse=True)}


# ===================== (MR1) XU HƯỚNG ĐIỂM THEO KỲ =====================
@router.get("/danh-gia-xu-huong")
def xu_huong(loai_ky: str = "THANG", nhan_vien_id: int | None = None,
             db: Session = Depends(get_db), _=Depends(yeu_cau(MODULE, "XEM"))):
    rows = db.query(DanhGiaKy).filter_by(loai_ky=loai_ky).all()
    by_ky: dict[str, list[float]] = {}
    for r in rows:
        by_ky.setdefault(r.ky, []).append(float(r.tong_diem or 0))
    tong_quan = [{"ky": k, "diem_tb": round(sum(v) / len(v), 2), "so": len(v)}
                 for k, v in sorted(by_ky.items())]
    theo_nv = None
    if nhan_vien_id:
        nr = sorted([r for r in rows if r.nhan_vien_id == nhan_vien_id], key=lambda x: x.ky)
        theo_nv = [{"ky": r.ky, "diem": float(r.tong_diem or 0), "xep_loai": r.xep_loai} for r in nr]
    return {"loai_ky": loai_ky, "tong_quan": tong_quan, "theo_nhan_vien": theo_nv}


# ===================== (MR2) XUẤT EXCEL =====================
_LK_VN = {"TUAN": "Tuần", "THANG": "Tháng", "QUY": "Quý", "NAM": "Năm"}


@router.get("/danh-gia-xuat")
def xuat_excel(loai_ky: str | None = None, ky: str | None = None,
               db: Session = Depends(get_db), _=Depends(yeu_cau(MODULE, "XEM"))):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    q = db.query(DanhGiaKy)
    if loai_ky:
        q = q.filter_by(loai_ky=loai_ky)
    if ky:
        q = q.filter_by(ky=ky)
    rows = q.order_by(DanhGiaKy.ky.desc(), DanhGiaKy.tong_diem.desc()).all()
    ten_vt = {v.ma: v.ten for v in db.query(VaiTro).all()}

    wb = Workbook()
    hdr_fill = PatternFill("solid", fgColor="0E7490")
    hdr_font = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center")
    thin = Border(*[Side(style="thin", color="D7DEE3")] * 4)

    ws = wb.active
    ws.title = "Tổng hợp"
    head = ["Mã NV", "Họ tên", "Vị trí", "Loại kỳ", "Kỳ", "Tổng điểm", "Xếp loại"]
    ws.append(head)
    for c in ws[1]:
        c.fill = hdr_fill; c.font = hdr_font; c.alignment = center; c.border = thin
    for r in rows:
        nv = db.get(NhanVien, r.nhan_vien_id)
        ws.append([nv.ma if nv else "", nv.ho_ten if nv else "",
                   ten_vt.get(r.vai_tro, r.vai_tro or ""), _LK_VN.get(r.loai_ky, r.loai_ky),
                   r.ky, float(r.tong_diem or 0), r.xep_loai or ""])
    for w, i in zip([12, 26, 30, 10, 12, 12, 10], range(1, 8)):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    # Sheet chi tiết KPI
    ws2 = wb.create_sheet("Chi tiết KPI")
    head2 = ["Họ tên", "Kỳ", "Chỉ tiêu", "Trọng số (%)", "Mục tiêu", "Thực tế", "% đạt", "Điểm"]
    ws2.append(head2)
    for c in ws2[1]:
        c.fill = hdr_fill; c.font = hdr_font; c.alignment = center; c.border = thin
    for r in rows:
        nv = db.get(NhanVien, r.nhan_vien_id)
        for ct in r.chi_tiet:
            ws2.append([nv.ho_ten if nv else "", r.ky, ct.ten,
                        float(ct.trong_so or 0),
                        float(ct.muc_tieu) if ct.muc_tieu is not None else "",
                        float(ct.gia_tri_thuc) if ct.gia_tri_thuc is not None else "",
                        float(ct.phan_tram_dat) if ct.phan_tram_dat is not None else "",
                        float(ct.diem or 0)])
    for w, i in zip([24, 12, 36, 12, 12, 12, 10, 10], range(1, 9)):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A2"

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    fn = f"danh-gia-kpi_{loai_ky or 'tatca'}_{ky or 'tatca'}.xlsx"
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fn}"'})


# ===================== (MR3) THƯỞNG KPI → LƯƠNG =====================
def _cfg_thuong(db: Session) -> CfgThuongKpi:
    c = db.get(CfgThuongKpi, 1)
    if c is None:
        c = CfgThuongKpi(id=1); db.add(c); db.commit(); db.refresh(c)
    return c


class CfgThuongVao(BaseModel):
    muc_co_so: Decimal | None = None
    hs_a: Decimal | None = None
    hs_b: Decimal | None = None
    hs_c: Decimal | None = None
    hs_d: Decimal | None = None


@router.get("/cau-hinh-thuong-kpi")
def lay_cfg_thuong(db: Session = Depends(get_db), _=Depends(yeu_cau(MODULE, "XEM"))):
    c = _cfg_thuong(db)
    return {"muc_co_so": float(c.muc_co_so or 0),
            "hs": {"A": float(c.hs_a or 0), "B": float(c.hs_b or 0),
                   "C": float(c.hs_c or 0), "D": float(c.hs_d or 0)}}


@router.put("/cau-hinh-thuong-kpi")
def cap_nhat_cfg_thuong(data: CfgThuongVao, db: Session = Depends(get_db),
                        nd: NguoiDung = Depends(yeu_cau(MODULE, "DUYET"))):
    c = _cfg_thuong(db)
    if data.muc_co_so is not None:
        c.muc_co_so = data.muc_co_so
    for k in ("hs_a", "hs_b", "hs_c", "hs_d"):
        v = getattr(data, k)
        if v is not None:
            setattr(c, k, v)
    ghi_audit(db, nd.id, "SUA", "cfg_thuong_kpi", 1)
    db.commit()
    return lay_cfg_thuong(db)


@router.post("/ky-luong/{thang}/ap-thuong-kpi")
def ap_thuong_kpi(thang: str, db: Session = Depends(get_db),
                  nd: NguoiDung = Depends(yeu_cau(MODULE, "DUYET"))):
    """Áp thưởng KPI vào bảng lương tháng theo đánh giá THÁNG cùng kỳ (xếp loại → hệ số)."""
    from .nhan_su import _ap_dung_tinh, _cap_nhat_tong_ky
    rows = db.query(BangLuong).filter_by(thang=thang).all()
    if not rows:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Kỳ chưa có bảng lương — hãy sinh bảng lương trước.")
    c = _cfg_thuong(db)
    hs = {"A": float(c.hs_a or 0), "B": float(c.hs_b or 0), "C": float(c.hs_c or 0), "D": float(c.hs_d or 0)}
    co_so = float(c.muc_co_so or 0)
    ap, bo_qua, chi_tiet = 0, [], []
    for bl in rows:
        if bl.trang_thai == "DA_DUYET":
            bo_qua.append(bl.nhan_vien_id); continue
        dg = db.query(DanhGiaKy).filter_by(nhan_vien_id=bl.nhan_vien_id,
                                           loai_ky="THANG", ky=thang).first()
        if dg is None:
            bl.thuong_kpi = 0
            nv = db.get(NhanVien, bl.nhan_vien_id)
            _ap_dung_tinh(db, nv, bl)
            continue
        he_so = hs.get(dg.xep_loai or "D", 0)
        bl.thuong_kpi = round(co_so * he_so)
        nv = db.get(NhanVien, bl.nhan_vien_id)
        _ap_dung_tinh(db, nv, bl)
        ap += 1
        chi_tiet.append({"nhan_vien_id": bl.nhan_vien_id, "ho_ten": nv.ho_ten if nv else "",
                         "xep_loai": dg.xep_loai, "he_so": he_so, "thuong_kpi": float(bl.thuong_kpi)})
    ky = db.get(KyLuong, thang)
    if ky:
        _cap_nhat_tong_ky(db, ky)
    ghi_audit(db, nd.id, "SUA", "bang_luong", None,
              moi={"thang": thang, "ap_thuong_kpi": ap})
    db.commit()
    return {"thang": thang, "so_ap": ap, "bo_qua_da_chot": len(bo_qua),
            "muc_co_so": co_so, "he_so": hs, "chi_tiet": chi_tiet}
