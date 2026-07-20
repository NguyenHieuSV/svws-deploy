"""
Module NHÂN SỰ / LƯƠNG — lát cắt dọc thứ sáu.
Điểm mới: QUY TRÌNH DUYỆT TUẦN TỰ (không theo hạn mức tiền):
  NV_HCNS lập lương -> KTT duyệt + hạch toán -> CEO ký cuối.
Dùng primitive chi_vai_tro(...) cho các bước gắn đúng vai trò.
Tính lương (BHXH/TNCN) tách trong luong_service; bút toán trong hach_toan.
"""
from decimal import Decimal
from datetime import date
from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File
from sqlalchemy.orm import Session
from sqlalchemy import func
from ..database import get_db
from ..rbac import yeu_cau, chi_vai_tro
from ..deps import nhan_vien_id_cua, lay_nguoi_dung_hien_tai
from ..audit import ghi_audit
from ..luong_service import tinh_luong
from ..hach_toan import hach_toan_luong
from ..models import (NguoiDung, NhanVien, ChamCong, NghiPhep, BangLuong, KyLuong, ThamSoLuong,
                      NgayNghiOt, BaoCaoNgay, ButToan, NhacViec)
from ..schemas import (NhanVienRa, ChamCongVao, NghiPhepVao, TinhLuongVao, BangLuongRa,
                        HoSoLuongVao, KyLuongVao, ChamCongLuongVao, ThamSoLuongVao, ChamCongImportVao,
                        NgayNghiOtVao, DangKyOtVao, TuChoiOtVao)
import json as _json
from datetime import date, datetime
from ..email_gateway import lay_email_provider
from ..config import settings
from decimal import Decimal as _Dec

router = APIRouter(prefix="/nhan-su", tags=["nhan_su"])
MODULE = "nhan_su"


@router.get("/nhan-vien", response_model=list[NhanVienRa])
def ds_nhan_vien(db: Session = Depends(get_db), _=Depends(yeu_cau(MODULE, "XEM"))):
    return db.query(NhanVien).order_by(NhanVien.id).all()


from pydantic import BaseModel as _NVBase


class NhanVienMoiVao(_NVBase):
    ho_ten: str
    ma: str | None = None
    chuc_danh: str | None = None
    luong_co_ban: float = 0
    luong_dong_bh: float = 0
    so_phu_thuoc: int = 0
    email: str | None = None
    so_tai_khoan: str | None = None
    ngan_hang: str | None = None
    tk_chi_phi: str = "642"


@router.post("/nhan-vien", status_code=201)
def tao_nhan_vien(data: NhanVienMoiVao, db: Session = Depends(get_db),
                  nd: NguoiDung = Depends(yeu_cau(MODULE, "THAO_TAC"))):
    """Thêm nhân viên mới vào hồ sơ lương (kỳ lương tạo sau sẽ tự gồm người này)."""
    if not data.ho_ten.strip():
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Họ tên là bắt buộc")
    if data.ma:
        trung = db.query(NhanVien).filter_by(ma=data.ma).first()
        if trung:
            raise HTTPException(status.HTTP_400_BAD_REQUEST, f"Mã NV '{data.ma}' đã tồn tại")
    nv = NhanVien(ma=data.ma, ho_ten=data.ho_ten.strip(), chuc_danh=data.chuc_danh,
                  luong_co_ban=data.luong_co_ban, luong_dong_bh=data.luong_dong_bh,
                  so_phu_thuoc=data.so_phu_thuoc, email=data.email,
                  so_tai_khoan=data.so_tai_khoan, ngan_hang=data.ngan_hang,
                  tk_chi_phi=data.tk_chi_phi or "642", trang_thai="DANG_LAM")
    db.add(nv); db.flush()
    ghi_audit(db, nd.id, "TAO", "nhan_vien", nv.id, moi={"ho_ten": nv.ho_ten, "chuc_danh": nv.chuc_danh})
    db.commit(); db.refresh(nv)
    return {"id": nv.id, "ma": nv.ma, "ho_ten": nv.ho_ten, "chuc_danh": nv.chuc_danh}


# ----- DUYET: xóa nhân viên (chỉ khi chưa có bảng lương/chấm công) -----
@router.delete("/nhan-vien/{nv_id}")
def xoa_nhan_vien(nv_id: int, db: Session = Depends(get_db),
                  nd: NguoiDung = Depends(chi_vai_tro("CEO", "ADMIN"))):
    """Xóa nhân viên thêm nhầm. Chặn khi đã có bảng lương, chấm công, nghỉ phép
    hoặc được tham chiếu trong chứng từ khác (người tạo/duyệt) — khi đó nên
    chuyển trạng thái NGHỈ VIỆC thay vì xóa."""
    from sqlalchemy import text as _sql
    from sqlalchemy.exc import IntegrityError
    nv = db.get(NhanVien, nv_id)
    if nv is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Không tìm thấy nhân viên")
    refs = [("bảng lương", "bang_luong"), ("chấm công", "cham_cong"),
            ("nghỉ phép", "nghi_phep")]
    ban = []
    for ten_ref, bang in refs:
        try:
            n = db.execute(_sql(f"SELECT COUNT(*) FROM {bang} WHERE nhan_vien_id = :i"),
                           {"i": nv_id}).scalar() or 0
        except Exception:
            n = 0
        if n:
            ban.append(f"{n} {ten_ref}")
    if ban:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            f"Nhân viên '{nv.ho_ten}' đang gắn với {', '.join(ban)} — không thể xóa. "
            "Hãy dùng trạng thái NGHỈ VIỆC để giữ hồ sơ.")
    ten_cu, ma_cu = nv.ho_ten, nv.ma
    try:
        db.execute(_sql("UPDATE nguoi_dung SET nhan_vien_id = NULL WHERE nhan_vien_id = :i"),
                   {"i": nv_id})
        db.execute(_sql("DELETE FROM kpi_danh_gia WHERE nhan_vien_id = :i"), {"i": nv_id})
    except Exception:
        pass
    try:
        db.delete(nv)
        db.flush()
    except IntegrityError:
        db.rollback()
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            f"Nhân viên '{ten_cu}' được tham chiếu trong chứng từ khác (người tạo/duyệt/phụ trách) "
            "— không thể xóa. Hãy dùng trạng thái NGHỈ VIỆC để giữ hồ sơ.")
    ghi_audit(db, nd.id, "XOA", "nhan_vien", nv_id, cu={"ma": ma_cu, "ho_ten": ten_cu})
    db.commit()
    return {"ok": True, "ho_ten": ten_cu}


# ----- Chấm công -----
@router.post("/cham-cong", status_code=201)
def cham_cong(data: ChamCongVao, db: Session = Depends(get_db),
              nd: NguoiDung = Depends(yeu_cau(MODULE, "THAO_TAC"))):
    cc = db.query(ChamCong).filter_by(nhan_vien_id=data.nhan_vien_id, ngay=data.ngay).first()
    if cc is None:
        cc = ChamCong(nhan_vien_id=data.nhan_vien_id, ngay=data.ngay)
        db.add(cc)
    cc.gio_vao = data.gio_vao or cc.gio_vao
    cc.gio_ra = data.gio_ra or cc.gio_ra
    db.commit()
    return {"nhan_vien_id": data.nhan_vien_id, "ngay": str(data.ngay), "trang_thai": "DA_GHI"}


# ----- Nghỉ phép -----
@router.post("/nghi-phep", status_code=201)
def xin_nghi(data: NghiPhepVao, db: Session = Depends(get_db),
             nd: NguoiDung = Depends(yeu_cau(MODULE, "THAO_TAC"))):
    np = NghiPhep(nhan_vien_id=data.nhan_vien_id, tu_ngay=data.tu_ngay,
                  den_ngay=data.den_ngay, loai=data.loai, trang_thai="CHO_DUYET")
    db.add(np); db.flush()
    ghi_audit(db, nd.id, "TAO", "nghi_phep", np.id, moi={"loai": data.loai})
    db.commit()
    return {"id": np.id, "trang_thai": "CHO_DUYET"}


@router.post("/nghi-phep/{np_id}/duyet")
def duyet_nghi(np_id: int, db: Session = Depends(get_db),
               nd: NguoiDung = Depends(yeu_cau(MODULE, "THAO_TAC"))):
    np = db.get(NghiPhep, np_id)
    if np is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Không tìm thấy đơn")
    np.trang_thai = "DA_DUYET"
    np.nguoi_duyet = nhan_vien_id_cua(db, nd.id)
    db.commit()
    return {"id": np.id, "trang_thai": "DA_DUYET"}


# ----- Tính lương cả tháng (NV_HCNS) -> trạng thái CHO_DUYET -----
@router.post("/tinh-luong")
def tinh_luong_thang(data: TinhLuongVao, db: Session = Depends(get_db),
                     nd: NguoiDung = Depends(yeu_cau(MODULE, "THAO_TAC"))):
    nvs = db.query(NhanVien).filter_by(trang_thai="DANG_LAM").all()
    tong = Decimal(0)
    n = 0
    for nv in nvs:
        bl = db.query(BangLuong).filter_by(nhan_vien_id=nv.id, thang=data.thang).first()
        if bl and bl.trang_thai == "DA_DUYET":
            continue  # đã chốt, không tính lại
        kq = tinh_luong(nv.luong_co_ban)
        if bl is None:
            bl = BangLuong(nhan_vien_id=nv.id, thang=data.thang)
            db.add(bl)
        for k, v in kq.items():
            setattr(bl, k, v)
        bl.trang_thai = "CHO_DUYET"
        bl.nguoi_duyet_ktt = None
        bl.nguoi_ky_ceo = None
        tong += kq["thuc_linh"]; n += 1
    ghi_audit(db, nd.id, "TAO", "bang_luong", None, moi={"thang": data.thang, "so_nv": n})
    db.commit()
    return {"thang": data.thang, "so_bang_luong": n, "tong_thuc_linh": float(tong)}


@router.get("/bang-luong", response_model=list[BangLuongRa])
def ds_bang_luong(thang: str | None = None, db: Session = Depends(get_db),
                  _=Depends(yeu_cau(MODULE, "XEM"))):
    q = db.query(BangLuong)
    if thang:
        q = q.filter_by(thang=thang)
    return q.order_by(BangLuong.id).all()


# ----- KTT duyệt + hạch toán (bước 1, chỉ KTT/CEO) -----
@router.post("/bang-luong/{thang}/duyet-ktt")
def ktt_duyet(thang: str, db: Session = Depends(get_db),
              nd: NguoiDung = Depends(yeu_cau(MODULE, "XEM")),
              __: NguoiDung = Depends(chi_vai_tro("KTT", "CEO"))):
    rows = db.query(BangLuong).filter_by(thang=thang, trang_thai="CHO_DUYET").all()
    chua = [r for r in rows if r.nguoi_duyet_ktt is None]
    if not chua:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Không có bảng lương chờ KTT duyệt")
    me = nhan_vien_id_cua(db, nd.id)
    so_bt = 0
    for bl in chua:
        bl.nguoi_duyet_ktt = me
        so_bt += len(hach_toan_luong(db, bl))
    ghi_audit(db, nd.id, "DUYET", "bang_luong", None,
              moi={"thang": thang, "so_bang": len(chua), "so_but_toan": so_bt})
    db.commit()
    return {"thang": thang, "so_bang_ktt_duyet": len(chua), "so_but_toan_luong": so_bt,
            "buoc_tiep": "CEO ký"}


# ----- CEO ký cuối (bước 2, chỉ CEO; yêu cầu KTT đã duyệt) -----
@router.post("/bang-luong/{thang}/ky-ceo")
def ceo_ky(thang: str, db: Session = Depends(get_db),
           nd: NguoiDung = Depends(yeu_cau(MODULE, "XEM")),
           __: NguoiDung = Depends(chi_vai_tro("CEO"))):
    rows = db.query(BangLuong).filter_by(thang=thang, trang_thai="CHO_DUYET").all()
    if not rows:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Không có bảng lương chờ ký")
    chua_ktt = [r for r in rows if r.nguoi_duyet_ktt is None]
    if chua_ktt:
        raise HTTPException(status.HTTP_400_BAD_REQUEST,
                            f"Còn {len(chua_ktt)} bảng chưa được KTT duyệt — chưa thể ký.")
    me = nhan_vien_id_cua(db, nd.id)
    for bl in rows:
        bl.trang_thai = "DA_DUYET"
        bl.nguoi_ky_ceo = me
    ghi_audit(db, nd.id, "DUYET", "bang_luong", None, moi={"thang": thang, "ky": len(rows)})
    db.commit()
    return {"thang": thang, "so_bang_da_ky": len(rows), "trang_thai": "DA_DUYET"}


# ====================== QUY TRÌNH LƯƠNG ĐẦY ĐỦ ======================
def _f(x):
    return float(x or 0)


def _lay_tham_so_luong(db):
    ts = db.get(ThamSoLuong, 1)
    if ts is None:
        ts = ThamSoLuong(id=1); db.add(ts); db.commit(); db.refresh(ts)
    return ts


def _cfg_luong(db):
    ts = _lay_tham_so_luong(db)
    try:
        bac = _json.loads(ts.bac_thue) if ts.bac_thue else None
    except Exception:
        bac = None
    return {
        "tl_bhxh_nv": ts.tl_bhxh_nv, "tl_bhyt_nv": ts.tl_bhyt_nv, "tl_bhtn_nv": ts.tl_bhtn_nv,
        "tl_bhxh_dn": ts.tl_bhxh_dn, "tl_bhyt_dn": ts.tl_bhyt_dn, "tl_bhtn_dn": ts.tl_bhtn_dn,
        "tran_bhxh_bhyt": ts.tran_bhxh_bhyt, "tran_bhtn": ts.tran_bhtn,
        "giam_tru_ban_than": ts.giam_tru_ban_than, "giam_tru_phu_thuoc": ts.giam_tru_phu_thuoc,
        "mien_thue_an": ts.mien_thue_an, "hs_ot_thuong": ts.hs_ot_thuong,
        "hs_ot_cuoi_tuan": ts.hs_ot_cuoi_tuan, "hs_ot_le": ts.hs_ot_le,
        "bac_thue": bac,
    }


def _ap_dung_tinh(db, nv: NhanVien, bl: BangLuong, cfg=None):
    cfg = cfg if cfg is not None else _cfg_luong(db)
    kq = tinh_luong(nv.luong_co_ban, nv.luong_dong_bh, bl.cong_chuan, bl.cong_thuc_te,
                    bl.gio_ot_thuong, bl.gio_ot_cuoi_tuan, bl.gio_ot_le,
                    nv.phu_cap_an, nv.phu_cap_di_lai, nv.phu_cap_dien_thoai, nv.phu_cap_trach_nhiem,
                    nv.so_phu_thuoc, bl.tam_ung,
                    (bl.phu_cap_khac or 0) + (bl.thuong_kpi or 0), bl.ngay_nghi_kpep, bl.so_phut_di_tre, bl.khau_tru_khac, cfg)
    bl.luong_co_ban = nv.luong_co_ban
    for k in ("luong_thuc_te", "ot", "phu_cap", "bhxh", "bhyt", "bhtn",
              "bhxh_dn", "bhyt_dn", "bhtn_dn", "thu_nhap_chiu_thue",
              "thue_tncn", "khau_tru", "thuc_linh", "chi_phi_dn",
              "khau_tru_nghi", "khau_tru_tre"):
        setattr(bl, k, kq[k])
    return kq


def _pl_dict(db, bl: BangLuong, nv: NhanVien = None):
    nv = nv or db.get(NhanVien, bl.nhan_vien_id)
    tong_thu_nhap = _f(bl.luong_thuc_te) + _f(bl.phu_cap) + _f(bl.ot)
    return {
        "id": bl.id, "nhan_vien_id": bl.nhan_vien_id, "ho_ten": nv.ho_ten if nv else "",
        "chuc_danh": nv.chuc_danh if nv else "", "thang": bl.thang,
        "ma_so_thue": nv.ma_so_thue if nv else None,
        "so_tai_khoan": nv.so_tai_khoan if nv else None, "ngan_hang": nv.ngan_hang if nv else None,
        "email": (nv.email if nv and nv.email else None),
        "cong_chuan": _f(bl.cong_chuan), "cong_thuc_te": _f(bl.cong_thuc_te),
        "gio_ot_thuong": _f(bl.gio_ot_thuong), "gio_ot_cuoi_tuan": _f(bl.gio_ot_cuoi_tuan),
        "gio_ot_le": _f(bl.gio_ot_le),
        "luong_co_ban": _f(bl.luong_co_ban), "luong_thuc_te": _f(bl.luong_thuc_te),
        "phu_cap": _f(bl.phu_cap), "phu_cap_khac": _f(bl.phu_cap_khac), "thuong_kpi": _f(bl.thuong_kpi), "ot": _f(bl.ot),
        "ngay_nghi_kpep": _f(bl.ngay_nghi_kpep), "so_phut_di_tre": int(bl.so_phut_di_tre or 0),
        "khau_tru_nghi": _f(bl.khau_tru_nghi), "khau_tru_tre": _f(bl.khau_tru_tre),
        "khau_tru_khac": _f(bl.khau_tru_khac), "tong_thu_nhap": tong_thu_nhap,
        "bhxh": _f(bl.bhxh), "bhyt": _f(bl.bhyt), "bhtn": _f(bl.bhtn),
        "bhxh_dn": _f(bl.bhxh_dn), "bhyt_dn": _f(bl.bhyt_dn), "bhtn_dn": _f(bl.bhtn_dn),
        "thu_nhap_chiu_thue": _f(bl.thu_nhap_chiu_thue), "thue_tncn": _f(bl.thue_tncn),
        "tam_ung": _f(bl.tam_ung), "khau_tru": _f(bl.khau_tru), "thuc_linh": _f(bl.thuc_linh),
        "chi_phi_dn": _f(bl.chi_phi_dn), "trang_thai": bl.trang_thai,
        "email_sent": bool(bl.email_sent),
    }


def _cap_nhat_tong_ky(db, ky: KyLuong):
    rows = db.query(BangLuong).filter_by(thang=ky.thang).all()
    ky.tong_thu_nhap = sum((_Dec(str(_f(r.luong_thuc_te) + _f(r.phu_cap) + _f(r.ot))) for r in rows), _Dec(0))
    ky.tong_thuc_linh = sum((_Dec(str(_f(r.thuc_linh))) for r in rows), _Dec(0))
    ky.tong_chi_phi_dn = sum((_Dec(str(_f(r.chi_phi_dn))) for r in rows), _Dec(0))


# ----- Tham số lương theo luật -----
@router.get("/tham-so-luong")
def lay_tham_so_luong_ep(db: Session = Depends(get_db), _=Depends(yeu_cau(MODULE, "XEM"))):
    ts = _lay_tham_so_luong(db)
    try:
        bac = _json.loads(ts.bac_thue) if ts.bac_thue else []
    except Exception:
        bac = []
    return {k: _f(getattr(ts, k)) for k in (
        "tl_bhxh_nv", "tl_bhyt_nv", "tl_bhtn_nv", "tl_bhxh_dn", "tl_bhyt_dn", "tl_bhtn_dn",
        "tran_bhxh_bhyt", "tran_bhtn", "giam_tru_ban_than", "giam_tru_phu_thuoc",
        "mien_thue_an", "hs_ot_thuong", "hs_ot_cuoi_tuan", "hs_ot_le",
        "luong_co_so", "luong_toi_thieu_vung")} | {"bac_thue": bac}


@router.put("/tham-so-luong")
def cap_nhat_tham_so_luong(data: ThamSoLuongVao, db: Session = Depends(get_db),
                           nd: NguoiDung = Depends(yeu_cau(MODULE, "DUYET"))):
    ts = _lay_tham_so_luong(db)
    for k in ("tl_bhxh_nv", "tl_bhyt_nv", "tl_bhtn_nv", "tl_bhxh_dn", "tl_bhyt_dn", "tl_bhtn_dn",
              "tran_bhxh_bhyt", "tran_bhtn", "giam_tru_ban_than", "giam_tru_phu_thuoc",
              "mien_thue_an", "hs_ot_thuong", "hs_ot_cuoi_tuan", "hs_ot_le",
              "luong_co_so", "luong_toi_thieu_vung"):
        v = getattr(data, k)
        if v is not None:
            setattr(ts, k, v)
    if data.bac_thue is not None:
        ts.bac_thue = _json.dumps(data.bac_thue)
    ghi_audit(db, nd.id, "CAP_NHAT", "tham_so_luong", 1, moi={"cap_nhat": True})
    db.commit()
    return lay_tham_so_luong_ep(db)


# ----- Nhận dữ liệu chấm công từ app chấm công (đẩy theo kỳ) -----
@router.post("/ky-luong/{thang}/cham-cong-import")
def import_cham_cong(thang: str, data: ChamCongImportVao, db: Session = Depends(get_db),
                     nd: NguoiDung = Depends(yeu_cau(MODULE, "THAO_TAC"))):
    ky = db.get(KyLuong, thang)
    if ky is None or ky.trang_thai == "DA_CHOT":
        raise HTTPException(status.HTTP_400_BAD_REQUEST,
                            "Kỳ chưa tạo hoặc đã chốt — không thể nhập chấm công.")
    cfg = _cfg_luong(db)
    cap_nhat, khong_thay = 0, []
    for rec in data.ban_ghi:
        nv = None
        if rec.nhan_vien_id:
            nv = db.get(NhanVien, rec.nhan_vien_id)
        elif rec.ma:
            nv = db.query(NhanVien).filter_by(ma=rec.ma).first()
        if nv is None:
            khong_thay.append(rec.ma or rec.nhan_vien_id); continue
        bl = db.query(BangLuong).filter_by(nhan_vien_id=nv.id, thang=thang).first()
        if bl is None:
            khong_thay.append(nv.ma or nv.id); continue
        for k in ("cong_thuc_te", "gio_ot_thuong", "gio_ot_cuoi_tuan", "gio_ot_le",
                  "ngay_nghi_kpep", "so_phut_di_tre"):
            v = getattr(rec, k)
            if v is not None:
                setattr(bl, k, v)
        _ap_dung_tinh(db, nv, bl, cfg)
        cap_nhat += 1
    _cap_nhat_tong_ky(db, ky)
    ghi_audit(db, nd.id, "IMPORT", "cham_cong", None, moi={"thang": thang, "so_ban_ghi": cap_nhat})
    db.commit()
    return {"thang": thang, "cap_nhat": cap_nhat, "khong_tim_thay": khong_thay}


# ===================== WORKING TIME & OVERTIME (BLLĐ 2019) =====================
_WT_OT = ("OT_THUONG", "OT_CUOI_TUAN", "OT_LE")
_WT_NGHI = ("NGHI_PHEP", "KHONG_PHEP", "NGHI_LE",
            "VIEC_RIENG_CO_LUONG", "VIEC_RIENG_KHONG_LUONG", "NGHI_BU")
# Trần luật: Đ105 (8h/ngày), Đ107 (OT ≤50%/ngày = 4h, ≤40h/tháng, ≤200h/năm — đặc thù 300h),
# Đ113 (phép năm 12 ngày), Đ125 (không phép ≥5 ngày/tháng hoặc ≥20 ngày/năm → căn cứ sa thải)
_WT_OT_NGAY, _WT_OT_THANG, _WT_OT_NAM, _WT_PHEP_NAM = 4, 40, 200, 12


def _wt_canh_bao(db, nv_id: int, ngay):
    """Cảnh báo vượt trần luật cho 1 NV tính đến tháng của `ngay`."""
    from datetime import date as _d
    y, m = ngay.year, ngay.month
    dau_thang = _d(y, m, 1)
    cuoi_thang = _d(y + 1, 1, 1) if m == 12 else _d(y, m + 1, 1)
    rows = db.query(NgayNghiOt).filter(NgayNghiOt.nhan_vien_id == nv_id,
                                       NgayNghiOt.trang_thai == "DA_DUYET",
                                       NgayNghiOt.ngay >= _d(y, 1, 1),
                                       NgayNghiOt.ngay <= _d(y, 12, 31)).all()
    cb = []
    ot_ngay = sum(float(r.so_gio or 0) for r in rows if r.loai in _WT_OT and r.ngay == ngay)
    ot_thang = sum(float(r.so_gio or 0) for r in rows
                   if r.loai in _WT_OT and dau_thang <= r.ngay < cuoi_thang)
    ot_nam = sum(float(r.so_gio or 0) for r in rows if r.loai in _WT_OT)
    phep_nam = sum(float(r.so_ngay or 0) for r in rows if r.loai == "NGHI_PHEP")
    kp_thang = sum(float(r.so_ngay or 0) for r in rows
                   if r.loai == "KHONG_PHEP" and dau_thang <= r.ngay < cuoi_thang)
    kp_nam = sum(float(r.so_ngay or 0) for r in rows if r.loai == "KHONG_PHEP")
    if ot_ngay > _WT_OT_NGAY:
        cb.append(f"OT {ot_ngay:g}h trong ngày {ngay:%d/%m} — vượt 50% giờ làm bình thường (tối đa 4h/ngày, Đ107)")
    if ot_thang > _WT_OT_THANG:
        cb.append(f"OT tháng {m:02d} đã {ot_thang:g}h — vượt trần 40h/tháng (Đ107)")
    if ot_nam > 300:
        cb.append(f"OT năm {y} đã {ot_nam:g}h — vượt cả trần 300h/năm ngành đặc thù (Đ107)")
    elif ot_nam > _WT_OT_NAM:
        cb.append(f"OT năm {y} đã {ot_nam:g}h — vượt trần 200h/năm (Đ107; tối đa 300h ngành đặc thù có thông báo)")
    if phep_nam > _WT_PHEP_NAM:
        cb.append(f"Đã nghỉ phép {phep_nam:g} ngày trong năm — vượt 12 ngày chuẩn (Đ113; kiểm tra thâm niên +1 ngày/5 năm, Đ114)")
    if kp_thang >= 5:
        cb.append(f"Nghỉ không phép {kp_thang:g} ngày trong tháng — đủ căn cứ xử lý kỷ luật sa thải (≥5 ngày cộng dồn/30 ngày, Đ125)")
    elif kp_nam >= 20:
        cb.append(f"Nghỉ không phép {kp_nam:g} ngày trong năm — đủ căn cứ xử lý kỷ luật sa thải (≥20 ngày cộng dồn/365 ngày, Đ125)")
    return cb


def _wt_so_lieu_luong(db, nv_id: int, thang: str):
    """Gom số liệu WT ĐÃ DUYỆT của 1 NV cho 1 tháng lương: giờ OT theo loại +
    ngày nghỉ bị khấu trừ (không phép, việc riêng không lương, phép vượt 12 ngày/năm).
    Trả None nếu tháng đó NV không có bản ghi WT nào (giữ nguyên chấm công tay)."""
    from datetime import date as _d
    y, m = int(thang[:4]), int(thang[5:7])
    dau_thang = _d(y, m, 1)
    cuoi_thang = _d(y + 1, 1, 1) if m == 12 else _d(y, m + 1, 1)
    rows = db.query(NgayNghiOt).filter(NgayNghiOt.nhan_vien_id == nv_id,
                                       NgayNghiOt.trang_thai == "DA_DUYET",
                                       NgayNghiOt.ngay >= _d(y, 1, 1),
                                       NgayNghiOt.ngay < cuoi_thang).all()
    thang_rs = [r for r in rows if r.ngay >= dau_thang]
    if not thang_rs:
        return None
    gio = lambda lo: sum(float(r.so_gio or 0) for r in thang_rs if r.loai == lo)
    ngay_ = lambda lo: sum(float(r.so_ngay or 0) for r in thang_rs if r.loai == lo)
    # phép vượt quota 12 ngày/năm PHÁT SINH trong tháng này → ngày không lương (Đ113)
    phep_truoc = sum(float(r.so_ngay or 0) for r in rows
                     if r.loai == "NGHI_PHEP" and r.ngay < dau_thang)
    phep_vuot = (max(0.0, phep_truoc + ngay_("NGHI_PHEP") - _WT_PHEP_NAM)
                 - max(0.0, phep_truoc - _WT_PHEP_NAM))
    khong_luong = ngay_("KHONG_PHEP") + ngay_("VIEC_RIENG_KHONG_LUONG") + phep_vuot
    return {"gio_ot_thuong": gio("OT_THUONG"), "gio_ot_cuoi_tuan": gio("OT_CUOI_TUAN"),
            "gio_ot_le": gio("OT_LE"), "ngay_nghi_kpep": round(khong_luong, 1)}


def _wt_dong_bo_luong(db, nv_id: int, ngay) -> bool:
    """Đồng bộ WT tháng chứa `ngay` vào bảng lương của NV (nếu kỳ đã sinh & chưa chốt).
    Gọi sau khi duyệt / HR ghi / xóa bản ghi WT — cột tăng ca & khấu trừ tự cập nhật."""
    thang = f"{ngay.year:04d}-{ngay.month:02d}"
    ky = db.get(KyLuong, thang)
    if ky is None or ky.trang_thai == "DA_CHOT":
        return False
    bl = db.query(BangLuong).filter_by(nhan_vien_id=nv_id, thang=thang).first()
    nv = db.get(NhanVien, nv_id)
    if bl is None or nv is None:
        return False
    wt = _wt_so_lieu_luong(db, nv_id, thang) or \
        {"gio_ot_thuong": 0, "gio_ot_cuoi_tuan": 0, "gio_ot_le": 0, "ngay_nghi_kpep": 0}
    bl.gio_ot_thuong = wt["gio_ot_thuong"]
    bl.gio_ot_cuoi_tuan = wt["gio_ot_cuoi_tuan"]
    bl.gio_ot_le = wt["gio_ot_le"]
    bl.ngay_nghi_kpep = wt["ngay_nghi_kpep"]
    _ap_dung_tinh(db, nv, bl)
    _cap_nhat_tong_ky(db, ky)
    return True


@router.get("/working-time")
def ds_working_time(thang: str | None = None, db: Session = Depends(get_db),
                    _=Depends(yeu_cau(MODULE, "XEM"))):
    """Bảng kiểm soát Working time & Overtime: chi tiết tháng + tổng hợp năm/tháng
    từng NV kèm cảnh báo vượt trần Bộ luật Lao động 2019."""
    from datetime import date as _d
    thang = thang or _d.today().strftime("%Y-%m")
    try:
        y, m = int(thang[:4]), int(thang[5:7])
        dau_thang = _d(y, m, 1)
    except (ValueError, IndexError):
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Tháng không hợp lệ (YYYY-MM)")
    cuoi_thang = _d(y + 1, 1, 1) if m == 12 else _d(y, m + 1, 1)
    rows_ca = db.query(NgayNghiOt).filter(NgayNghiOt.ngay >= _d(y, 1, 1),
                                          NgayNghiOt.ngay <= _d(y, 12, 31)).all()
    rows_nam = [r for r in rows_ca if r.trang_thai == "DA_DUYET"]
    nvs = [v for v in db.query(NhanVien).order_by(NhanVien.id).all()
           if (v.trang_thai or "DANG_LAM") == "DANG_LAM"]
    ten = {v.id: v for v in nvs}
    danh_sach = [{"id": r.id, "nhan_vien_id": r.nhan_vien_id,
                  "ma": ten[r.nhan_vien_id].ma if r.nhan_vien_id in ten else None,
                  "ho_ten": ten[r.nhan_vien_id].ho_ten if r.nhan_vien_id in ten else f"NV #{r.nhan_vien_id}",
                  "ngay": str(r.ngay), "loai": r.loai, "so_gio": _f(r.so_gio),
                  "so_ngay": _f(r.so_ngay), "ghi_chu": r.ghi_chu,
                  "trang_thai": r.trang_thai, "ly_do_tu_choi": r.ly_do_tu_choi}
                 for r in sorted(rows_ca, key=lambda x: (x.ngay, x.nhan_vien_id))
                 if dau_thang <= r.ngay < cuoi_thang]
    # danh sách chờ duyệt (mọi tháng) cho người có quyền duyệt
    cho_duyet = [{"id": r.id, "nhan_vien_id": r.nhan_vien_id,
                  "ho_ten": ten[r.nhan_vien_id].ho_ten if r.nhan_vien_id in ten else f"NV #{r.nhan_vien_id}",
                  "ngay": str(r.ngay), "loai": r.loai, "so_gio": _f(r.so_gio),
                  "ghi_chu": r.ghi_chu}
                 for r in sorted((x for x in rows_ca if x.trang_thai == "CHO_DUYET"),
                                 key=lambda x: x.ngay, reverse=True)][:100]
    tong_hop = []
    for v in nvs:
        rs = [r for r in rows_nam if r.nhan_vien_id == v.id]
        thang_rs = [r for r in rs if dau_thang <= r.ngay < cuoi_thang]
        gio = lambda lst, lo: sum(float(r.so_gio or 0) for r in lst if r.loai == lo)
        ngay_ = lambda lst, lo: sum(float(r.so_ngay or 0) for r in lst if r.loai == lo)
        ot_thang = sum(gio(thang_rs, lo) for lo in _WT_OT)
        ot_nam = sum(gio(rs, lo) for lo in _WT_OT)
        cb = []
        if ot_thang > _WT_OT_THANG:
            cb.append(f"OT tháng {ot_thang:g}h > 40h (Đ107)")
        if ot_nam > 300:
            cb.append(f"OT năm {ot_nam:g}h > 300h (Đ107)")
        elif ot_nam > _WT_OT_NAM:
            cb.append(f"OT năm {ot_nam:g}h > 200h (Đ107)")
        if ngay_(rs, "NGHI_PHEP") > _WT_PHEP_NAM:
            cb.append(f"Phép {ngay_(rs, 'NGHI_PHEP'):g} ngày > 12 (Đ113)")
        if ngay_(thang_rs, "KHONG_PHEP") >= 5:
            cb.append("Không phép ≥5 ngày/tháng — căn cứ sa thải (Đ125)")
        elif ngay_(rs, "KHONG_PHEP") >= 20:
            cb.append("Không phép ≥20 ngày/năm — căn cứ sa thải (Đ125)")
        tong_hop.append({
            "nhan_vien_id": v.id, "ma": v.ma, "ho_ten": v.ho_ten, "chuc_danh": v.chuc_danh,
            "phep_nam": ngay_(rs, "NGHI_PHEP"), "quota_phep": _WT_PHEP_NAM,
            "khong_phep_thang": ngay_(thang_rs, "KHONG_PHEP"),
            "khong_phep_nam": ngay_(rs, "KHONG_PHEP"),
            "viec_rieng_nam": ngay_(rs, "VIEC_RIENG_CO_LUONG") + ngay_(rs, "VIEC_RIENG_KHONG_LUONG"),
            "nghi_le_nam": ngay_(rs, "NGHI_LE"), "nghi_bu_nam": ngay_(rs, "NGHI_BU"),
            "ot_thuong": gio(thang_rs, "OT_THUONG"), "ot_cuoi_tuan": gio(thang_rs, "OT_CUOI_TUAN"),
            "ot_le": gio(thang_rs, "OT_LE"), "ot_thang": ot_thang, "ot_nam": ot_nam,
            "canh_bao": cb})
    return {"thang": thang, "danh_sach": danh_sach, "tong_hop": tong_hop, "cho_duyet": cho_duyet}


@router.post("/working-time", status_code=201)
def ghi_working_time(data: NgayNghiOtVao, db: Session = Depends(get_db),
                     nd: NguoiDung = Depends(yeu_cau(MODULE, "THAO_TAC"))):
    """Ghi nhận nghỉ / tăng ca. Nghỉ nhiều ngày (den_ngay) → tạo từng dòng/ngày, bỏ Chủ nhật.
    Trùng (NV + ngày + loại) → cập nhật dòng cũ. Trả kèm cảnh báo vượt trần luật."""
    from datetime import timedelta
    nv = db.get(NhanVien, data.nhan_vien_id)
    if nv is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Không tìm thấy nhân viên")
    loai = (data.loai or "").upper()
    if loai not in _WT_OT + _WT_NGHI:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Loại không hợp lệ")
    la_ot = loai in _WT_OT
    if la_ot:
        gio = float(data.so_gio or 0)
        if not (0 < gio <= 16):
            raise HTTPException(status.HTTP_400_BAD_REQUEST, "Nhập số giờ tăng ca (0–16h)")
        cac_ngay = [data.ngay]
    else:
        den = data.den_ngay or data.ngay
        if den < data.ngay:
            raise HTTPException(status.HTTP_400_BAD_REQUEST, "Đến ngày phải sau Từ ngày")
        if (den - data.ngay).days > 60:
            raise HTTPException(status.HTTP_400_BAD_REQUEST, "Khoảng nghỉ tối đa 60 ngày/lần ghi")
        cac_ngay = [data.ngay + timedelta(days=i) for i in range((den - data.ngay).days + 1)]
        if len(cac_ngay) > 1:            # nghỉ dài ngày: bỏ Chủ nhật (ngày nghỉ hằng tuần)
            cac_ngay = [d for d in cac_ngay if d.weekday() != 6]
    so_ngay = float(data.so_ngay or 1)
    if not la_ot and so_ngay not in (0.5, 1):
        so_ngay = 1
    nguoi_ghi = nhan_vien_id_cua(db, nd.id)
    them, cap_nhat = 0, 0
    for d in cac_ngay:
        cu = db.query(NgayNghiOt).filter_by(nhan_vien_id=nv.id, ngay=d, loai=loai).first()
        if cu is not None:
            cu.so_gio = data.so_gio or 0
            cu.so_ngay = so_ngay if not la_ot else 0
            cu.ghi_chu = data.ghi_chu
            cu.trang_thai = "DA_DUYET"      # HR ghi trực tiếp = đã duyệt
            cu.nguoi_duyet = nguoi_ghi
            cu.ly_do_tu_choi = None
            cap_nhat += 1
        else:
            db.add(NgayNghiOt(nhan_vien_id=nv.id, ngay=d, loai=loai,
                              so_gio=(data.so_gio or 0) if la_ot else 0,
                              so_ngay=0 if la_ot else so_ngay, ghi_chu=data.ghi_chu,
                              nguoi_tao=nguoi_ghi, trang_thai="DA_DUYET",
                              nguoi_duyet=nguoi_ghi))
            them += 1
    db.flush()
    canh_bao = _wt_canh_bao(db, nv.id, data.ngay)
    # gợi ý chọn đúng loại theo thứ trong tuần (Chủ nhật = ngày nghỉ hằng tuần)
    if la_ot and loai == "OT_THUONG" and data.ngay.weekday() == 6:
        canh_bao.append("Ngày này là Chủ nhật — tăng ca ngày nghỉ hằng tuần phải trả ≥200% (Đ98): chọn loại Tăng ca cuối tuần")
    if la_ot and loai == "OT_CUOI_TUAN" and data.ngay.weekday() < 5:
        canh_bao.append("Ngày này là ngày làm việc trong tuần — kiểm tra lại loại tăng ca (ngày thường ≥150%)")
    # tự cập nhật bảng lương các tháng bị ảnh hưởng (nếu kỳ đã sinh & chưa chốt)
    dong_bo = False
    for thang_d in {(d.year, d.month) for d in cac_ngay}:
        from datetime import date as _d
        if _wt_dong_bo_luong(db, nv.id, _d(thang_d[0], thang_d[1], 1)):
            dong_bo = True
    ghi_audit(db, nd.id, "TAO", "ngay_nghi_ot", nv.id,
              moi={"loai": loai, "ngay": str(data.ngay), "so_dong": them + cap_nhat,
                   "dong_bo_luong": dong_bo})
    db.commit()
    return {"them": them, "cap_nhat": cap_nhat, "canh_bao": canh_bao, "dong_bo_luong": dong_bo}


@router.delete("/working-time/{wt_id}")
def xoa_working_time(wt_id: int, db: Session = Depends(get_db),
                     nd: NguoiDung = Depends(chi_vai_tro("CEO", "ADMIN"))):
    r = db.get(NgayNghiOt, wt_id)
    if r is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Không tìm thấy bản ghi")
    nv_id, ngay_xoa = r.nhan_vien_id, r.ngay
    ghi_audit(db, nd.id, "XOA", "ngay_nghi_ot", wt_id, cu={"loai": r.loai, "ngay": str(r.ngay)})
    db.delete(r); db.flush()
    _wt_dong_bo_luong(db, nv_id, ngay_xoa)   # rút số liệu khỏi bảng lương nếu kỳ chưa chốt
    db.commit()
    return {"da_xoa": True}


# ----- Nhân viên tự đăng ký tăng ca → chờ duyệt (chỉ cần đăng nhập) -----
def _wt_nv_cua_toi(db, nd) -> int:
    nv_id = nhan_vien_id_cua(db, nd.id)
    if not nv_id:
        raise HTTPException(status.HTTP_400_BAD_REQUEST,
                            "Tài khoản của bạn chưa gắn với hồ sơ nhân viên — nhờ HR gắn ở Hồ sơ lương (nút 🔑 Tạo tài khoản).")
    return nv_id


# Lịch nghỉ lễ nhà nước (Đ112): lễ dương lịch cố định hằng năm + lễ âm lịch/liền kề chốt theo năm
_WT_LE_CO_DINH = {"01-01", "04-30", "05-01", "09-02"}
_WT_LE_THEO_NAM = {
    2026: {"2026-02-16", "2026-02-17", "2026-02-18", "2026-02-19", "2026-02-20",  # Tết Bính Ngọ (5 ngày)
           "2026-04-26",                                                          # Giỗ Tổ 10/3 ÂL
           "2026-09-01"},                                                         # liền kề Quốc khánh
}


def _wt_loai_ngay(d) -> str:
    """Phân loại OT theo lịch: ngày lễ → OT_LE; Chủ nhật → OT_CUOI_TUAN; còn lại OT_THUONG.
    (Công chuẩn 26 — thứ Bảy tính ngày làm việc bình thường.)"""
    s = str(d)
    if s[5:] in _WT_LE_CO_DINH or s in _WT_LE_THEO_NAM.get(d.year, set()):
        return "OT_LE"
    return "OT_CUOI_TUAN" if d.weekday() == 6 else "OT_THUONG"


@router.post("/working-time/dang-ky", status_code=201)
def dang_ky_ot(data: DangKyOtVao, db: Session = Depends(get_db),
               nd: NguoiDung = Depends(lay_nguoi_dung_hien_tai)):
    """NV tự đăng ký tăng ca cho CHÍNH MÌNH — vào hàng chờ duyệt (Đ107: OT phải có sự đồng ý).
    Khai Từ giờ/ngày → Đến giờ/ngày: loại + số giờ TỰ TÍNH theo lịch; ca qua đêm tách 2 phần theo ngày."""
    from datetime import timedelta
    nv_id = _wt_nv_cua_toi(db, nd)
    phan = []                                    # [(ngay, so_gio, khung_gio)]
    if data.tu_gio is not None and data.den_gio is not None:
        t1, t2 = int(data.tu_gio), int(data.den_gio)
        n1, n2 = data.ngay, data.den_ngay or data.ngay
        if not (0 <= t1 <= 23) or not (1 <= t2 <= 24):
            raise HTTPException(status.HTTP_400_BAD_REQUEST, "Giờ không hợp lệ (Từ: 0–23, Đến: 1–24)")
        if n2 == n1:
            if t2 <= t1:
                raise HTTPException(status.HTTP_400_BAD_REQUEST, "Giờ Đến phải sau giờ Từ")
            phan = [(n1, t2 - t1, f"{t1}h–{t2}h")]
        elif n2 == n1 + timedelta(days=1):
            phan = [(n1, 24 - t1, f"{t1}h–24h")]
            if t2 > 0 and t2 < 24:
                phan.append((n2, t2, f"0h–{t2}h"))
        else:
            raise HTTPException(status.HTTP_400_BAD_REQUEST,
                                "Chỉ đăng ký trong ngày hoặc ca qua đêm (Đến ngày = ngày kế tiếp)")
        if not (0 < sum(p[1] for p in phan) <= 16):
            raise HTTPException(status.HTTP_400_BAD_REQUEST, "Tổng giờ tăng ca phải trong khoảng 0–16h")
    else:
        # tương thích cách cũ: gửi thẳng loai + so_gio
        loai_cu = (data.loai or "").upper()
        if loai_cu not in _WT_OT:
            raise HTTPException(status.HTTP_400_BAD_REQUEST, "Thiếu khung giờ đăng ký (Từ / Đến)")
        gio = float(data.so_gio or 0)
        if not (0 < gio <= 16):
            raise HTTPException(status.HTTP_400_BAD_REQUEST, "Số giờ tăng ca phải trong khoảng 0–16h")
        phan = [(data.ngay, gio, None)]
    ket_qua = []
    for ng, gio, khung in phan:
        loai = _wt_loai_ngay(ng) if (data.tu_gio is not None) else (data.loai or "").upper()
        cu = db.query(NgayNghiOt).filter_by(nhan_vien_id=nv_id, ngay=ng, loai=loai).first()
        if cu is not None and cu.trang_thai == "DA_DUYET":
            raise HTTPException(status.HTTP_400_BAD_REQUEST,
                                f"Ngày {ng:%d/%m} đã có tăng ca loại này ĐÃ DUYỆT — liên hệ quản lý nếu cần sửa.")
        gc = " · ".join(x for x in ((data.ghi_chu or "").strip(), khung) if x)[:300] or None
        if cu is not None:
            cu.so_gio = gio; cu.ghi_chu = gc
            cu.trang_thai = "CHO_DUYET"; cu.ly_do_tu_choi = None; cu.nguoi_duyet = None
        else:
            db.add(NgayNghiOt(nhan_vien_id=nv_id, ngay=ng, loai=loai, so_gio=gio,
                              so_ngay=0, ghi_chu=gc, nguoi_tao=nv_id, trang_thai="CHO_DUYET"))
        ket_qua.append({"ngay": str(ng), "loai": loai, "so_gio": gio})
    ghi_audit(db, nd.id, "DANG_KY_OT", "ngay_nghi_ot", nv_id,
              moi={"phan": ket_qua})
    db.commit()
    return {"ok": True, "trang_thai": "CHO_DUYET", "phan": ket_qua}


@router.get("/working-time/cua-toi")
def wt_cua_toi(db: Session = Depends(get_db),
               nd: NguoiDung = Depends(lay_nguoi_dung_hien_tai)):
    """Đăng ký tăng ca của chính mình trong năm + tổng OT đã duyệt tháng này / năm nay."""
    from datetime import date as _d
    nv_id = _wt_nv_cua_toi(db, nd)
    hom_nay = _d.today()
    rows = db.query(NgayNghiOt).filter(
        NgayNghiOt.nhan_vien_id == nv_id, NgayNghiOt.loai.in_(_WT_OT),
        NgayNghiOt.ngay >= _d(hom_nay.year, 1, 1)).order_by(NgayNghiOt.ngay.desc()).all()
    thang_nay = hom_nay.strftime("%Y-%m")
    ot_thang = sum(float(r.so_gio or 0) for r in rows
                   if r.trang_thai == "DA_DUYET" and str(r.ngay)[:7] == thang_nay)
    ot_nam = sum(float(r.so_gio or 0) for r in rows if r.trang_thai == "DA_DUYET")
    return {"danh_sach": [{"id": r.id, "ngay": str(r.ngay), "loai": r.loai,
                           "so_gio": _f(r.so_gio), "ghi_chu": r.ghi_chu,
                           "trang_thai": r.trang_thai, "ly_do_tu_choi": r.ly_do_tu_choi}
                          for r in rows],
            "ot_thang": ot_thang, "ot_nam": ot_nam}


@router.delete("/working-time/cua-toi/{wt_id}")
def huy_dang_ky_ot(wt_id: int, db: Session = Depends(get_db),
                   nd: NguoiDung = Depends(lay_nguoi_dung_hien_tai)):
    """NV hủy đăng ký tăng ca của mình khi còn CHỜ DUYỆT."""
    nv_id = _wt_nv_cua_toi(db, nd)
    r = db.get(NgayNghiOt, wt_id)
    if r is None or r.nhan_vien_id != nv_id:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Không tìm thấy đăng ký của bạn")
    if r.trang_thai != "CHO_DUYET":
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Chỉ hủy được đăng ký còn chờ duyệt")
    db.delete(r); db.commit()
    return {"da_huy": True}


@router.post("/working-time/{wt_id}/duyet")
def duyet_ot(wt_id: int, db: Session = Depends(get_db),
             nd: NguoiDung = Depends(yeu_cau(MODULE, "DUYET"))):
    r = db.get(NgayNghiOt, wt_id)
    if r is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Không tìm thấy đăng ký")
    if r.trang_thai != "CHO_DUYET":
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Đăng ký này không ở trạng thái chờ duyệt")
    r.trang_thai = "DA_DUYET"
    r.nguoi_duyet = nhan_vien_id_cua(db, nd.id)
    db.flush()
    canh_bao = _wt_canh_bao(db, r.nhan_vien_id, r.ngay)
    dong_bo = _wt_dong_bo_luong(db, r.nhan_vien_id, r.ngay)   # tự cập nhật bảng lương nếu kỳ đã sinh
    ghi_audit(db, nd.id, "DUYET", "ngay_nghi_ot", wt_id,
              moi={"ngay": str(r.ngay), "so_gio": _f(r.so_gio), "dong_bo_luong": dong_bo})
    db.commit()
    return {"ok": True, "canh_bao": canh_bao, "dong_bo_luong": dong_bo}


@router.post("/working-time/{wt_id}/tu-choi")
def tu_choi_ot(wt_id: int, data: TuChoiOtVao = TuChoiOtVao(), db: Session = Depends(get_db),
               nd: NguoiDung = Depends(yeu_cau(MODULE, "DUYET"))):
    r = db.get(NgayNghiOt, wt_id)
    if r is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Không tìm thấy đăng ký")
    if r.trang_thai != "CHO_DUYET":
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Đăng ký này không ở trạng thái chờ duyệt")
    r.trang_thai = "TU_CHOI"
    r.nguoi_duyet = nhan_vien_id_cua(db, nd.id)
    r.ly_do_tu_choi = (data.ly_do or "").strip()[:300] or None
    ghi_audit(db, nd.id, "TU_CHOI", "ngay_nghi_ot", wt_id, moi={"ly_do": r.ly_do_tu_choi})
    db.commit()
    return {"ok": True}


# ============ Daily Report — báo cáo công việc hằng ngày ============
class BaoCaoNgayVao(_NVBase):
    ngay: str | None = None
    da_lam: str = ""
    ket_qua: str = ""
    kho_khan: str | None = None
    ke_hoach: str | None = None
    so_gio: float | None = 0
    ma_lien_quan: str | None = None


def _bc_ra(r: BaoCaoNgay, nv: NhanVien | None = None) -> dict:
    return {"id": r.id, "nhan_vien_id": r.nhan_vien_id, "ngay": str(r.ngay),
            "da_lam": r.da_lam or "", "ket_qua": r.ket_qua or "",
            "kho_khan": r.kho_khan, "ke_hoach": r.ke_hoach,
            "so_gio": _f(r.so_gio), "ma_lien_quan": r.ma_lien_quan,
            "ho_ten": (nv.ho_ten if nv else None), "ma_nv": (nv.ma if nv else None)}


@router.post("/bao-cao-ngay", status_code=201)
def gui_bao_cao_ngay(data: BaoCaoNgayVao, db: Session = Depends(get_db),
                     nd: NguoiDung = Depends(lay_nguoi_dung_hien_tai)):
    """Nhân viên gửi/cập nhật báo cáo công việc của CHÍNH MÌNH cho một ngày (upsert theo ngày)."""
    nv_id = _wt_nv_cua_toi(db, nd)
    ngay = (data.ngay or "").strip() or str(date.today())
    try:
        ng = date.fromisoformat(ngay)
    except ValueError:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Ngày không hợp lệ (YYYY-MM-DD)")
    if not (data.da_lam or "").strip():
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Vui lòng nhập nội dung công việc đã làm")
    r = db.query(BaoCaoNgay).filter(BaoCaoNgay.nhan_vien_id == nv_id,
                                    BaoCaoNgay.ngay == ng).first()
    moi = r is None
    if moi:
        r = BaoCaoNgay(nhan_vien_id=nv_id, ngay=ng)
        db.add(r)
    r.da_lam = (data.da_lam or "").strip()
    r.ket_qua = (data.ket_qua or "").strip()
    r.kho_khan = (data.kho_khan or "").strip() or None
    r.ke_hoach = (data.ke_hoach or "").strip() or None
    r.so_gio = _Dec(str(data.so_gio or 0))
    r.ma_lien_quan = (data.ma_lien_quan or "").strip()[:120] or None
    db.flush()
    ghi_audit(db, nd.id, "TAO" if moi else "SUA", "bao_cao_ngay", r.id,
              moi={"ngay": str(ng), "so_gio": _f(r.so_gio)})
    db.commit()
    return {"ok": True, "id": r.id, "moi": moi}


@router.get("/bao-cao-ngay/cua-toi")
def bao_cao_cua_toi(thang: str | None = None, db: Session = Depends(get_db),
                    nd: NguoiDung = Depends(lay_nguoi_dung_hien_tai)):
    """Danh sách báo cáo của chính mình trong tháng (mặc định tháng hiện tại)."""
    nv_id = _wt_nv_cua_toi(db, nd)
    thang = (thang or str(date.today())[:7]).strip()
    q = db.query(BaoCaoNgay).filter(BaoCaoNgay.nhan_vien_id == nv_id)
    if len(thang) == 7:
        q = q.filter(func.to_char(BaoCaoNgay.ngay, "YYYY-MM") == thang)
    rows = q.order_by(BaoCaoNgay.ngay.desc()).all()
    return {"danh_sach": [_bc_ra(r) for r in rows], "nhan_vien_id": nv_id, "thang": thang}


@router.get("/bao-cao-ngay")
def ds_bao_cao_ngay(ngay: str | None = None, thang: str | None = None,
                    nhan_vien_id: int | None = None, db: Session = Depends(get_db),
                    _=Depends(yeu_cau(MODULE, "XEM"))):
    """Quản lý xem báo cáo của toàn công ty — lọc theo ngày/tháng/nhân viên."""
    q = db.query(BaoCaoNgay, NhanVien).join(NhanVien, BaoCaoNgay.nhan_vien_id == NhanVien.id)
    if nhan_vien_id:
        q = q.filter(BaoCaoNgay.nhan_vien_id == nhan_vien_id)
    if ngay:
        q = q.filter(BaoCaoNgay.ngay == ngay)
    elif thang and len(thang) == 7:
        q = q.filter(func.to_char(BaoCaoNgay.ngay, "YYYY-MM") == thang)
    else:
        q = q.filter(BaoCaoNgay.ngay == str(date.today()))
    rows = q.order_by(BaoCaoNgay.ngay.desc(), NhanVien.ma).all()
    return {"danh_sach": [_bc_ra(r, nv) for r, nv in rows]}


@router.delete("/bao-cao-ngay/{bc_id}")
def xoa_bao_cao_ngay(bc_id: int, db: Session = Depends(get_db),
                     nd: NguoiDung = Depends(lay_nguoi_dung_hien_tai)):
    """Xóa báo cáo — chủ nhân tự xóa được, hoặc CEO/ADMIN xóa bất kỳ."""
    r = db.get(BaoCaoNgay, bc_id)
    if r is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Không tìm thấy báo cáo")
    nv_id = nhan_vien_id_cua(db, nd.id)
    la_qtv = (nd.vai_tro.ma if nd.vai_tro else "").upper() in ("CEO", "ADMIN")
    if r.nhan_vien_id != nv_id and not la_qtv:
        raise HTTPException(status.HTTP_403_FORBIDDEN, "Chỉ được xóa báo cáo của chính mình")
    ghi_audit(db, nd.id, "XOA", "bao_cao_ngay", bc_id, cu={"ngay": str(r.ngay)})
    db.delete(r); db.commit()
    return {"da_xoa": True}


@router.get("/danh-ba")
def danh_ba_nhan_vien(db: Session = Depends(get_db),
                      nd: NguoiDung = Depends(lay_nguoi_dung_hien_tai)):
    """Danh bạ tối giản (id · mã · họ tên · chức danh) cho MỌI nhân viên đã đăng nhập —
    phục vụ droplist “nhắc ai” / “người hỗ trợ”. KHÔNG chứa lương hay dữ liệu nhạy cảm."""
    rows = (db.query(NhanVien).filter(NhanVien.trang_thai == "DANG_LAM")
            .order_by(NhanVien.ho_ten).all())
    return [{"id": nv.id, "ma": nv.ma, "ho_ten": nv.ho_ten, "chuc_danh": nv.chuc_danh}
            for nv in rows]


# ============ Work Reminder — nhắc việc cho mình / cho người khác ============
_NV_MUC_DO = ("THAP", "BINH_THUONG", "CAO", "KHAN")
_NV_TRANG_THAI = ("CHO_LAM", "DANG_LAM", "XONG", "HUY")


def _gio_vn():
    """Giờ Việt Nam — máy chủ chạy UTC, còn thoi_diem/xong_luc lưu giờ địa phương."""
    from ..nhac_viec_service import gio_hien_tai
    return gio_hien_tai()


class NhacViecVao(_NVBase):
    nhan_vien_id: int | None = None          # nhắc ai (trống = nhắc chính mình)
    nhac_tat_ca: bool = False                # True = nhắc TOÀN BỘ nhân viên đang làm
    tieu_de: str = ""
    thoi_diem: str = ""                      # 'YYYY-MM-DDTHH:MM' giờ địa phương — lúc NHẮC
    han_hoan_thanh: str | None = None        # hạn PHẢI HOÀN THÀNH (tùy chọn)
    ma_lien_quan: str | None = None
    chuan_bi: str | None = None
    nguoi_ho_tro_id: int | None = None
    ho_tro_gi: str | None = None
    muc_do: str = "BINH_THUONG"
    ghi_chu: str | None = None
    trang_thai: str | None = None            # chỉ dùng khi SỬA


def _nv_doc_thoi_diem(s: str) -> datetime:
    s = (s or "").strip().replace(" ", "T")
    if not s:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Vui lòng chọn thời điểm nhắc")
    try:
        return datetime.fromisoformat(s[:16])
    except ValueError:
        raise HTTPException(status.HTTP_400_BAD_REQUEST,
                            "Thời điểm không hợp lệ (định dạng YYYY-MM-DD HH:MM)")


def _nv_ra(r: NhacViec, ten: dict) -> dict:
    return {"id": r.id, "nguoi_tao": r.nguoi_tao, "nhan_vien_id": r.nhan_vien_id,
            "tieu_de": r.tieu_de, "thoi_diem": r.thoi_diem.isoformat(timespec="minutes"),
            "han_hoan_thanh": (r.han_hoan_thanh.isoformat(timespec="minutes")
                               if r.han_hoan_thanh else None),
            "nhom": r.nhom,
            "ma_lien_quan": r.ma_lien_quan, "chuan_bi": r.chuan_bi,
            "nguoi_ho_tro_id": r.nguoi_ho_tro_id, "ho_tro_gi": r.ho_tro_gi,
            "muc_do": r.muc_do, "trang_thai": r.trang_thai, "ghi_chu": r.ghi_chu,
            "xong_luc": r.xong_luc.isoformat(timespec="minutes") if r.xong_luc else None,
            "nguoi_tao_ten": ten.get(r.nguoi_tao), "nhan_vien_ten": ten.get(r.nhan_vien_id),
            "ho_tro_ten": ten.get(r.nguoi_ho_tro_id)}


def _nv_ten_map(db, rows) -> dict:
    ids = set()
    for r in rows:
        ids.update([r.nguoi_tao, r.nhan_vien_id, r.nguoi_ho_tro_id])
    ids.discard(None)
    if not ids:
        return {}
    return {nv.id: nv.ho_ten for nv in db.query(NhanVien).filter(NhanVien.id.in_(ids)).all()}


def _nv_sua_duoc(db, r: NhacViec, nd: NguoiDung) -> bool:
    """Người tạo, người được nhắc, hoặc CEO/ADMIN."""
    me = nhan_vien_id_cua(db, nd.id)
    la_qtv = (nd.vai_tro.ma if nd.vai_tro else "").upper() in ("CEO", "ADMIN")
    return la_qtv or (me is not None and me in (r.nguoi_tao, r.nhan_vien_id))


@router.post("/nhac-viec", status_code=201)
def tao_nhac_viec(data: NhacViecVao, db: Session = Depends(get_db),
                  nd: NguoiDung = Depends(lay_nguoi_dung_hien_tai)):
    """Mọi nhân viên đặt lời nhắc cho chính mình hoặc cho người khác."""
    me = _wt_nv_cua_toi(db, nd)
    tieu_de = (data.tieu_de or "").strip()
    if not tieu_de:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Vui lòng nhập việc cần làm")
    if data.nguoi_ho_tro_id and db.get(NhanVien, data.nguoi_ho_tro_id) is None:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Không tìm thấy người hỗ trợ")
    thoi_diem = _nv_doc_thoi_diem(data.thoi_diem)
    han = _nv_doc_thoi_diem(data.han_hoan_thanh) if (data.han_hoan_thanh or "").strip() else None
    if han is not None and han < thoi_diem:
        raise HTTPException(status.HTTP_400_BAD_REQUEST,
                            "Hạn hoàn thành phải sau thời điểm nhắc")
    # Nhắc tất cả: mỗi nhân viên MỘT dòng riêng (để từng người tự đánh dấu xong),
    # các dòng cùng lô chia sẻ cùng mã `nhom` để gom nhóm khi hiển thị.
    if data.nhac_tat_ca:
        nguoi_nhan_ids = [nv.id for nv in db.query(NhanVien)
                          .filter(NhanVien.trang_thai == "DANG_LAM").all()]
        if not nguoi_nhan_ids:
            raise HTTPException(status.HTTP_400_BAD_REQUEST, "Không có nhân viên đang làm việc")
        nhom = f"ALL-{me}-{datetime.now().strftime('%Y%m%d%H%M%S')}"
    else:
        nguoi_nhan = data.nhan_vien_id or me
        if db.get(NhanVien, nguoi_nhan) is None:
            raise HTTPException(status.HTTP_400_BAD_REQUEST, "Không tìm thấy người được nhắc")
        nguoi_nhan_ids, nhom = [nguoi_nhan], None
    chung = dict(nguoi_tao=me, tieu_de=tieu_de, thoi_diem=thoi_diem, han_hoan_thanh=han,
                 nhom=nhom,
                 ma_lien_quan=(data.ma_lien_quan or "").strip()[:120] or None,
                 chuan_bi=(data.chuan_bi or "").strip() or None,
                 nguoi_ho_tro_id=data.nguoi_ho_tro_id,
                 ho_tro_gi=(data.ho_tro_gi or "").strip() or None,
                 muc_do=(data.muc_do if data.muc_do in _NV_MUC_DO else "BINH_THUONG"),
                 ghi_chu=(data.ghi_chu or "").strip() or None)
    tao = [NhacViec(nhan_vien_id=i, **chung) for i in nguoi_nhan_ids]
    for r in tao:
        db.add(r)
    db.flush()
    ghi_audit(db, nd.id, "TAO", "nhac_viec", tao[0].id,
              moi={"tieu_de": tieu_de, "so_nguoi": len(tao), "nhom": nhom,
                   "thoi_diem": data.thoi_diem, "han_hoan_thanh": data.han_hoan_thanh})
    db.commit()
    # Gửi Google Chat — best-effort, KHÔNG được làm hỏng việc tạo lời nhắc
    chat = {"da_gui": 0}
    try:
        from ..nhac_viec_service import gui_khi_tao
        chat = gui_khi_tao(db, tao)
        db.commit()
    except Exception as e:
        chat = {"da_gui": 0, "loi": [f"{type(e).__name__}: {e}"]}
    return {"ok": True, "id": tao[0].id, "so_nguoi": len(tao), "nhom": nhom, "chat": chat}


@router.get("/nhac-viec/cua-toi")
def nhac_viec_cua_toi(db: Session = Depends(get_db),
                      nd: NguoiDung = Depends(lay_nguoi_dung_hien_tai)):
    """Việc người khác nhắc tôi + việc tôi nhắc người khác."""
    me = _wt_nv_cua_toi(db, nd)
    rows = db.query(NhacViec).filter(
        (NhacViec.nhan_vien_id == me) | (NhacViec.nguoi_tao == me)
    ).order_by(NhacViec.thoi_diem).all()
    ten = _nv_ten_map(db, rows)
    ra = [_nv_ra(r, ten) for r in rows]
    return {"nhan_vien_id": me,
            "nhac_toi": [x for x in ra if x["nhan_vien_id"] == me],
            "toi_nhac": [x for x in ra if x["nguoi_tao"] == me and x["nhan_vien_id"] != me]}


@router.get("/nhac-viec")
def ds_nhac_viec(trang_thai: str | None = None, nhan_vien_id: int | None = None,
                 db: Session = Depends(get_db), _=Depends(yeu_cau(MODULE, "XEM"))):
    """Quản lý xem toàn bộ lời nhắc của công ty."""
    q = db.query(NhacViec)
    if trang_thai:
        q = q.filter(NhacViec.trang_thai == trang_thai)
    if nhan_vien_id:
        q = q.filter(NhacViec.nhan_vien_id == nhan_vien_id)
    rows = q.order_by(NhacViec.thoi_diem).limit(400).all()
    ten = _nv_ten_map(db, rows)
    return {"danh_sach": [_nv_ra(r, ten) for r in rows]}


@router.put("/nhac-viec/{nv_id}")
def sua_nhac_viec(nv_id: int, data: NhacViecVao, db: Session = Depends(get_db),
                  nd: NguoiDung = Depends(lay_nguoi_dung_hien_tai)):
    r = db.get(NhacViec, nv_id)
    if r is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Không tìm thấy lời nhắc")
    if not _nv_sua_duoc(db, r, nd):
        raise HTTPException(status.HTTP_403_FORBIDDEN,
                            "Chỉ người tạo, người được nhắc hoặc CEO/Admin mới sửa được lời nhắc này")
    if (data.tieu_de or "").strip():
        r.tieu_de = data.tieu_de.strip()
    if (data.thoi_diem or "").strip():
        r.thoi_diem = _nv_doc_thoi_diem(data.thoi_diem)
    r.han_hoan_thanh = (_nv_doc_thoi_diem(data.han_hoan_thanh)
                        if (data.han_hoan_thanh or "").strip() else None)
    if r.han_hoan_thanh is not None and r.han_hoan_thanh < r.thoi_diem:
        raise HTTPException(status.HTTP_400_BAD_REQUEST,
                            "Hạn hoàn thành phải sau thời điểm nhắc")
    if data.nhan_vien_id and db.get(NhanVien, data.nhan_vien_id) is not None:
        r.nhan_vien_id = data.nhan_vien_id
    r.ma_lien_quan = (data.ma_lien_quan or "").strip()[:120] or None
    r.chuan_bi = (data.chuan_bi or "").strip() or None
    r.nguoi_ho_tro_id = data.nguoi_ho_tro_id
    r.ho_tro_gi = (data.ho_tro_gi or "").strip() or None
    r.ghi_chu = (data.ghi_chu or "").strip() or None
    if data.muc_do in _NV_MUC_DO:
        r.muc_do = data.muc_do
    if data.trang_thai in _NV_TRANG_THAI:
        r.trang_thai = data.trang_thai
        r.xong_luc = _gio_vn() if data.trang_thai == "XONG" else None
    ghi_audit(db, nd.id, "SUA", "nhac_viec", nv_id, moi={"tieu_de": r.tieu_de})
    db.commit()
    return {"ok": True}


@router.post("/nhac-viec/{nv_id}/trang-thai")
def doi_trang_thai_nhac_viec(nv_id: int, trang_thai: str, db: Session = Depends(get_db),
                             nd: NguoiDung = Depends(lay_nguoi_dung_hien_tai)):
    """Đánh dấu nhanh: đang làm / xong / hủy."""
    if trang_thai not in _NV_TRANG_THAI:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Trạng thái không hợp lệ")
    r = db.get(NhacViec, nv_id)
    if r is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Không tìm thấy lời nhắc")
    if not _nv_sua_duoc(db, r, nd):
        raise HTTPException(status.HTTP_403_FORBIDDEN,
                            "Chỉ người tạo, người được nhắc hoặc CEO/Admin mới đổi được trạng thái")
    r.trang_thai = trang_thai
    r.xong_luc = _gio_vn() if trang_thai == "XONG" else None
    ghi_audit(db, nd.id, "SUA", "nhac_viec", nv_id, moi={"trang_thai": trang_thai})
    db.commit()
    return {"ok": True, "trang_thai": trang_thai}


@router.delete("/nhac-viec/{nv_id}")
def xoa_nhac_viec(nv_id: int, db: Session = Depends(get_db),
                  nd: NguoiDung = Depends(lay_nguoi_dung_hien_tai)):
    """Xóa lời nhắc — người TẠO hoặc CEO/ADMIN."""
    r = db.get(NhacViec, nv_id)
    if r is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Không tìm thấy lời nhắc")
    me = nhan_vien_id_cua(db, nd.id)
    la_qtv = (nd.vai_tro.ma if nd.vai_tro else "").upper() in ("CEO", "ADMIN")
    if r.nguoi_tao != me and not la_qtv:
        raise HTTPException(status.HTTP_403_FORBIDDEN,
                            "Chỉ người đặt lời nhắc hoặc CEO/Admin mới xóa được")
    ghi_audit(db, nd.id, "XOA", "nhac_viec", nv_id, cu={"tieu_de": r.tieu_de})
    db.delete(r); db.commit()
    return {"da_xoa": True}


@router.get("/chat-trang-thai")
def chat_trang_thai(nd: NguoiDung = Depends(chi_vai_tro("CEO", "ADMIN"))):
    """Cấu hình Google Chat đang ở chế độ nào (không lộ khóa bí mật)."""
    from ..config import settings as _st
    sa = (_st.gchat_service_account or "").strip()
    email_sa = ""
    if sa:
        try:
            email_sa = _json.loads(sa).get("client_email", "")
        except Exception:
            email_sa = "(JSON không hợp lệ)"
    import os as _os
    raw = _st.chat_provider or ""
    # Chẩn đoán: máy chủ thực sự đọc được gì (không lộ khóa) — giúp tìm lỗi cấu hình
    chan_doan = {
        "chat_provider_raw": repr(raw),                      # thấy dấu cách/hoa-thường thừa
        "env_CHAT_PROVIDER": repr(_os.environ.get("CHAT_PROVIDER")),
        "co_GCHAT_SERVICE_ACCOUNT": bool(sa),
        "do_dai_service_account": len(sa),
        "env_co_GCHAT_SERVICE_ACCOUNT": bool(_os.environ.get("GCHAT_SERVICE_ACCOUNT")),
    }
    return {"che_do": (_st.chat_provider or "DEMO").strip().upper(),
            "co_webhook": bool((_st.gchat_webhook_url or "").strip()),
            "service_account": email_sa,
            "gui_khi_tao": _st.nhac_viec_gui_khi_tao,
            "ban_tin": _st.nhac_viec_ban_tin,
            "gio_ban_tin": _st.nhac_viec_gio_ban_tin,
            "chan_doan": chan_doan}


@router.post("/chat-thu")
def chat_gui_thu(email: str | None = None, db: Session = Depends(get_db),
                 nd: NguoiDung = Depends(chi_vai_tro("CEO", "ADMIN"))):
    """Gửi một tin THỬ để kiểm tra kết nối Google Chat.
    Không truyền email thì gửi cho chính người đang đăng nhập."""
    from ..chat_gateway import lay_chat_provider
    from ..nhac_viec_service import gui_toi_nguoi
    if not email:
        me = nhan_vien_id_cua(db, nd.id)
        nv = db.get(NhanVien, me) if me else None
        email = (nv.email if nv else "") or ""
    if not email:
        raise HTTPException(status.HTTP_400_BAD_REQUEST,
                            "Tài khoản của bạn chưa có email — truyền ?email=... để thử.")
    kq = gui_toi_nguoi(
        db, lay_chat_provider(), email,
        "✅ *SVWS thử kết nối Google Chat*\nNếu bạn đọc được tin này thì cấu hình đã đúng.")
    return {"toi": email, **kq}


@router.get("/chat-dm")
def chat_dm_ds(db: Session = Depends(get_db),
               nd: NguoiDung = Depends(chi_vai_tro("CEO", "ADMIN"))):
    """Danh sách người đã nhắn bot (đã ghi nhớ phòng nhắn riêng) — để kiểm tra."""
    from ..models import ChatDM
    rows = db.query(ChatDM).order_by(ChatDM.cap_nhat.desc()).limit(200).all()
    return {"danh_sach": [{"email": r.google_email, "ten": r.ten_hthi,
                           "co_space": bool(r.space_name), "user_id": r.user_id}
                          for r in rows]}


@router.post("/nhac-viec/ban-tin-thu")
def ban_tin_thu(db: Session = Depends(get_db),
                nd: NguoiDung = Depends(chi_vai_tro("CEO", "ADMIN"))):
    """Gửi NGAY bản tin tổng hợp (bỏ qua chốt ngày) để xem thử nội dung."""
    from ..nhac_viec_service import gui_ban_tin_ngay
    return gui_ban_tin_ngay(db, ep=True)


@router.delete("/nhac-viec/nhom/{nhom}")
def xoa_nhom_nhac_viec(nhom: str, db: Session = Depends(get_db),
                       nd: NguoiDung = Depends(lay_nguoi_dung_hien_tai)):
    """Xóa CẢ LÔ lời nhắc “Nhắc tất cả” — người TẠO hoặc CEO/ADMIN."""
    rows = db.query(NhacViec).filter(NhacViec.nhom == nhom).all()
    if not rows:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Không tìm thấy lô lời nhắc")
    me = nhan_vien_id_cua(db, nd.id)
    la_qtv = (nd.vai_tro.ma if nd.vai_tro else "").upper() in ("CEO", "ADMIN")
    if rows[0].nguoi_tao != me and not la_qtv:
        raise HTTPException(status.HTTP_403_FORBIDDEN,
                            "Chỉ người đặt lời nhắc hoặc CEO/Admin mới xóa được")
    ghi_audit(db, nd.id, "XOA", "nhac_viec", rows[0].id,
              cu={"nhom": nhom, "tieu_de": rows[0].tieu_de, "so_dong": len(rows)})
    for r in rows:
        db.delete(r)
    db.commit()
    return {"da_xoa": True, "so_dong": len(rows)}


# ----- Hồ sơ lương -----
@router.get("/ho-so")
def ds_ho_so(db: Session = Depends(get_db), _=Depends(yeu_cau(MODULE, "XEM"))):
    out = []
    for nv in db.query(NhanVien).order_by(NhanVien.id).all():
        out.append({"id": nv.id, "ma": nv.ma, "ho_ten": nv.ho_ten, "chuc_danh": nv.chuc_danh,
                    "trang_thai": nv.trang_thai, "luong_co_ban": _f(nv.luong_co_ban),
                    "luong_dong_bh": _f(nv.luong_dong_bh), "so_phu_thuoc": nv.so_phu_thuoc,
                    "phu_cap_an": _f(nv.phu_cap_an), "phu_cap_di_lai": _f(nv.phu_cap_di_lai),
                    "phu_cap_dien_thoai": _f(nv.phu_cap_dien_thoai),
                    "phu_cap_trach_nhiem": _f(nv.phu_cap_trach_nhiem),
                    "ma_so_thue": nv.ma_so_thue, "so_tai_khoan": nv.so_tai_khoan,
                    "ngan_hang": nv.ngan_hang, "email": nv.email, "tk_chi_phi": nv.tk_chi_phi})
    return out


@router.get("/ho-so-xuat-excel")
def xuat_excel_ho_so(db: Session = Depends(get_db), _=Depends(yeu_cau(MODULE, "XEM"))):
    """Xuất bảng hồ sơ lương nhân viên ra Excel."""
    import io
    from datetime import date as _date
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook(); ws = wb.active; ws.title = "Hồ sơ lương"
    hdr_fill = PatternFill("solid", fgColor="0E7490")
    hdr_font = Font(bold=True, color="FFFFFF")
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    thin = Border(*[Side(style="thin", color="D7DEE3")] * 4)
    money = "#,##0"

    ws.merge_cells("A1:P1")
    ws["A1"] = "HỒ SƠ LƯƠNG NHÂN VIÊN"
    ws["A1"].font = Font(bold=True, size=15, color="0E7490"); ws["A1"].alignment = center
    ws.append(["Ngày lập:", _date.today().strftime("%d/%m/%Y")])
    ws.cell(row=2, column=1).font = bold
    ws.append([])

    head = ["STT", "Mã NV", "Họ tên", "Chức danh", "Trạng thái", "Lương cơ bản",
            "Lương đóng BH", "PC ăn", "PC đi lại", "PC điện thoại", "PC trách nhiệm",
            "Tổng phụ cấp", "Phụ thuộc", "Email", "Số tài khoản", "Ngân hàng"]
    ws.append(head)
    hr = ws.max_row
    for c in ws[hr]:
        c.fill = hdr_fill; c.font = hdr_font; c.alignment = center; c.border = thin

    for i, nv in enumerate(db.query(NhanVien).order_by(NhanVien.id).all(), start=1):
        pc = [_f(nv.phu_cap_an) or 0, _f(nv.phu_cap_di_lai) or 0,
              _f(nv.phu_cap_dien_thoai) or 0, _f(nv.phu_cap_trach_nhiem) or 0]
        ws.append([i, nv.ma or nv.id, nv.ho_ten, nv.chuc_danh or "", nv.trang_thai or "",
                   _f(nv.luong_co_ban) or 0, _f(nv.luong_dong_bh) or _f(nv.luong_co_ban) or 0,
                   pc[0], pc[1], pc[2], pc[3], sum(pc), nv.so_phu_thuoc or 0,
                   nv.email or "", nv.so_tai_khoan or "", nv.ngan_hang or ""])
        rr = ws.max_row
        for c in ws[rr]:
            c.border = thin
        for col in range(6, 13):
            ws.cell(row=rr, column=col).number_format = money

    for w, i in zip([6, 10, 26, 22, 12, 14, 14, 11, 11, 12, 13, 13, 10, 26, 16, 14], range(1, 17)):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = f"A{hr + 1}"

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    fn = f"ho-so-luong_{_date.today():%Y%m%d}.xlsx"
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fn}"'})


def _hs_chuan(s) -> str:
    """Chuẩn hóa tiêu đề cột Excel: bỏ dấu, thường, chỉ giữ chữ + số."""
    import re as _re
    import unicodedata
    s = str(s or "").replace("Đ", "D").replace("đ", "d")
    s = "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")
    return _re.sub(r"[^a-z0-9]+", " ", s.lower()).strip()


_HS_COT = {   # tên cột chuẩn hóa -> trường hồ sơ lương
    "ma nv": "ma", "manv": "ma", "ma nhan vien": "ma",
    "ho ten": "ho_ten", "ho va ten": "ho_ten", "ten nhan vien": "ho_ten",
    "chuc danh": "chuc_danh", "vi tri": "chuc_danh",
    "luong co ban": "luong_co_ban", "luong cb": "luong_co_ban",
    "luong dong bh": "luong_dong_bh", "luong bh": "luong_dong_bh",
    "luong dong bao hiem": "luong_dong_bh",
    "pc an": "phu_cap_an", "phu cap an": "phu_cap_an",
    "pc di lai": "phu_cap_di_lai", "phu cap di lai": "phu_cap_di_lai",
    "pc dien thoai": "phu_cap_dien_thoai", "phu cap dien thoai": "phu_cap_dien_thoai",
    "pc trach nhiem": "phu_cap_trach_nhiem", "phu cap trach nhiem": "phu_cap_trach_nhiem",
    "phu thuoc": "so_phu_thuoc", "so phu thuoc": "so_phu_thuoc",
    "so nguoi phu thuoc": "so_phu_thuoc",
    "email": "email", "mst": "ma_so_thue", "ma so thue": "ma_so_thue",
    "so tai khoan": "so_tai_khoan", "stk": "so_tai_khoan",
    "ngan hang": "ngan_hang",
}
_HS_SO = ("luong_co_ban", "luong_dong_bh", "phu_cap_an", "phu_cap_di_lai",
          "phu_cap_dien_thoai", "phu_cap_trach_nhiem", "so_phu_thuoc")


@router.post("/ho-so-nhap-excel")
async def nhap_excel_ho_so(file: UploadFile = File(...), db: Session = Depends(get_db),
                           nd: NguoiDung = Depends(yeu_cau(MODULE, "THAO_TAC"))):
    """Cập nhật hồ sơ lương từ file Excel (.xlsx): khớp theo Mã NV rồi đến Họ tên —
    có rồi thì cập nhật các cột có trong file; chưa có thì tạo nhân viên mới.
    Dùng đúng bố cục file '⬇ Xuất Excel' hoặc file tự tạo có dòng tiêu đề cột."""
    import io
    import re as _re
    from openpyxl import load_workbook
    data = await file.read()
    if len(data) > 10 * 1024 * 1024:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "File quá lớn (tối đa 10MB)")
    try:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(status.HTTP_400_BAD_REQUEST,
                            "Không mở được file — cần file Excel .xlsx (file .xls đời cũ hãy lưu lại thành .xlsx)")
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    # tìm dòng tiêu đề: có >= 2 cột khớp và phải có Họ tên hoặc Mã NV
    cot, dong_bd = None, None
    for i, r in enumerate(rows[:15]):
        m = {}
        for j, v in enumerate(r or ()):
            f = _HS_COT.get(_hs_chuan(v))
            if f and f not in m.values():
                m[j] = f
        if len(m) >= 2 and ("ho_ten" in m.values() or "ma" in m.values()):
            cot, dong_bd = m, i + 1
            break
    if not cot:
        raise HTTPException(status.HTTP_400_BAD_REQUEST,
                            "Không tìm thấy dòng tiêu đề cột (cần các cột như Mã NV, Họ tên, Lương cơ bản...). "
                            "Dễ nhất: bấm ⬇ Xuất Excel, sửa số liệu trong file đó rồi nhập lại.")

    def so(v):
        if v is None or v == "":
            return None
        if isinstance(v, (int, float)):
            return round(float(v))
        d = _re.sub(r"[^\d]", "", str(v))
        return int(d) if d else None

    cap_nhat, tao_moi, bo_qua, chi_tiet = 0, 0, 0, []
    for r in rows[dong_bd:]:
        if not r or all(v is None or str(v).strip() == "" for v in r):
            continue
        it = {}
        for j, f in cot.items():
            v = r[j] if j < len(r) else None
            if v is None or str(v).strip() == "":
                continue
            it[f] = so(v) if f in _HS_SO else str(v).strip()
        ten = (it.get("ho_ten") or "").strip()
        ma = (it.get("ma") or "").strip()
        if not ten and not ma:
            continue
        nv = None
        if ma:
            nv = db.query(NhanVien).filter(NhanVien.ma == ma).first()
        if nv is None and ten:
            nv = db.query(NhanVien).filter(NhanVien.ho_ten.ilike(ten)).first()
        if nv is not None:
            doi = []
            for f, v in it.items():
                if f in ("ma", "ho_ten"):
                    continue
                if v is not None and str(getattr(nv, f) or "") != str(v):
                    setattr(nv, f, v)
                    doi.append(f)
            if doi:
                cap_nhat += 1
                chi_tiet.append({"ten": nv.ho_ten, "kq": "Cập nhật (" + str(len(doi)) + " cột)"})
            else:
                bo_qua += 1
                chi_tiet.append({"ten": nv.ho_ten, "kq": "Không đổi"})
        elif ten:
            nv = NhanVien(ma=ma or None, ho_ten=ten, chuc_danh=it.get("chuc_danh"),
                          luong_co_ban=it.get("luong_co_ban") or 0,
                          luong_dong_bh=it.get("luong_dong_bh") or 0,
                          so_phu_thuoc=it.get("so_phu_thuoc") or 0,
                          phu_cap_an=it.get("phu_cap_an") or 0,
                          phu_cap_di_lai=it.get("phu_cap_di_lai") or 0,
                          phu_cap_dien_thoai=it.get("phu_cap_dien_thoai") or 0,
                          phu_cap_trach_nhiem=it.get("phu_cap_trach_nhiem") or 0,
                          ma_so_thue=it.get("ma_so_thue"), email=it.get("email"),
                          so_tai_khoan=it.get("so_tai_khoan"), ngan_hang=it.get("ngan_hang"),
                          tk_chi_phi="642", trang_thai="DANG_LAM")
            db.add(nv); db.flush()
            tao_moi += 1
            chi_tiet.append({"ten": ten, "kq": "Tạo mới"})
        else:
            bo_qua += 1
            chi_tiet.append({"ten": ma, "kq": "Bỏ qua — không có họ tên"})
        if cap_nhat + tao_moi + bo_qua >= 300:
            break
    ghi_audit(db, nd.id, "NHAP_EXCEL", "nhan_vien", None,
              moi={"file": file.filename, "cap_nhat": cap_nhat, "tao_moi": tao_moi})
    db.commit()
    return {"cap_nhat": cap_nhat, "tao_moi": tao_moi, "bo_qua": bo_qua,
            "chi_tiet": chi_tiet[:60]}


@router.put("/nhan-vien/{nv_id}/ho-so")
def cap_nhat_ho_so(nv_id: int, data: HoSoLuongVao, db: Session = Depends(get_db),
                   nd: NguoiDung = Depends(yeu_cau(MODULE, "THAO_TAC"))):
    nv = db.get(NhanVien, nv_id)
    if nv is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Không tìm thấy nhân viên")
    for k in ("luong_co_ban", "luong_dong_bh", "so_phu_thuoc", "phu_cap_an", "phu_cap_di_lai",
              "phu_cap_dien_thoai", "phu_cap_trach_nhiem", "ma_so_thue", "so_tai_khoan",
              "ngan_hang", "email", "tk_chi_phi"):
        setattr(nv, k, getattr(data, k))
    if data.chuc_danh is not None:
        nv.chuc_danh = data.chuc_danh
    if data.ho_ten is not None and data.ho_ten.strip():
        nv.ho_ten = data.ho_ten.strip()
    if data.ma is not None and data.ma.strip():
        trung = db.query(NhanVien).filter(NhanVien.ma == data.ma.strip(),
                                          NhanVien.id != nv.id).first()
        if trung:
            raise HTTPException(status.HTTP_400_BAD_REQUEST, f"Mã NV '{data.ma}' đã tồn tại")
        nv.ma = data.ma.strip()
    ghi_audit(db, nd.id, "CAP_NHAT", "nhan_vien", nv.id, moi={"luong_co_ban": _f(nv.luong_co_ban)})
    db.commit()
    return {"id": nv.id, "trang_thai": "DA_LUU"}


# ----- Kỳ lương: tạo & sinh bảng lương cho toàn bộ NV đang làm -----
@router.post("/ky-luong")
def tao_ky_luong(data: KyLuongVao, db: Session = Depends(get_db),
                 nd: NguoiDung = Depends(yeu_cau(MODULE, "THAO_TAC"))):
    y, mth = int(data.thang[:4]), int(data.thang[5:7])
    ngay_chot = data.ngay_chot or date(y, mth, 7) if mth != 12 else (data.ngay_chot or date(y, 12, 7))
    ky = db.get(KyLuong, data.thang)
    if ky is None:
        ky = KyLuong(thang=data.thang); db.add(ky)
    if ky.trang_thai == "DA_CHOT":
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Kỳ lương đã chốt, không thể sinh lại.")
    ky.cong_chuan = data.cong_chuan
    ky.ngay_chot = ngay_chot
    n, dong_bo_wt = 0, 0
    for nv in db.query(NhanVien).filter_by(trang_thai="DANG_LAM").all():
        bl = db.query(BangLuong).filter_by(nhan_vien_id=nv.id, thang=data.thang).first()
        moi = bl is None
        if bl is None:
            bl = BangLuong(nhan_vien_id=nv.id, thang=data.thang); db.add(bl)
        bl.cong_chuan = data.cong_chuan
        if moi:
            bl.cong_thuc_te = data.cong_chuan
        bl.trang_thai = "CHO_DUYET"
        db.flush()
        # kéo số liệu Working time & Overtime ĐÃ DUYỆT vào cột tăng ca + khấu trừ
        wt = _wt_so_lieu_luong(db, nv.id, data.thang)
        if wt is not None:
            bl.gio_ot_thuong = wt["gio_ot_thuong"]
            bl.gio_ot_cuoi_tuan = wt["gio_ot_cuoi_tuan"]
            bl.gio_ot_le = wt["gio_ot_le"]
            bl.ngay_nghi_kpep = wt["ngay_nghi_kpep"]
            dong_bo_wt += 1
        _ap_dung_tinh(db, nv, bl)
        n += 1
    _cap_nhat_tong_ky(db, ky)
    ghi_audit(db, nd.id, "TAO", "ky_luong", None,
              moi={"thang": data.thang, "so_nv": n, "dong_bo_wt": dong_bo_wt})
    db.commit()
    return {"thang": data.thang, "so_nhan_vien": n, "ngay_chot": str(ngay_chot),
            "cong_chuan": _f(ky.cong_chuan), "trang_thai": ky.trang_thai,
            "dong_bo_wt": dong_bo_wt}


@router.get("/ky-luong")
def ds_ky_luong(db: Session = Depends(get_db), _=Depends(yeu_cau(MODULE, "XEM"))):
    out = []
    for ky in db.query(KyLuong).order_by(KyLuong.thang.desc()).all():
        out.append({"thang": ky.thang, "cong_chuan": _f(ky.cong_chuan),
                    "ngay_chot": str(ky.ngay_chot) if ky.ngay_chot else None,
                    "trang_thai": ky.trang_thai, "da_gui_email": bool(ky.da_gui_email),
                    "tong_thu_nhap": _f(ky.tong_thu_nhap), "tong_thuc_linh": _f(ky.tong_thuc_linh),
                    "tong_chi_phi_dn": _f(ky.tong_chi_phi_dn)})
    return out


@router.get("/ky-luong/{thang}")
def chi_tiet_ky(thang: str, db: Session = Depends(get_db), _=Depends(yeu_cau(MODULE, "XEM"))):
    ky = db.get(KyLuong, thang)
    rows = db.query(BangLuong).filter_by(thang=thang).order_by(BangLuong.id).all()
    pl = [_pl_dict(db, r) for r in rows]
    tre_han = bool(ky and ky.ngay_chot and date.today() > ky.ngay_chot and ky.trang_thai != "DA_CHOT")
    head = ({"thang": ky.thang, "cong_chuan": _f(ky.cong_chuan),
             "ngay_chot": str(ky.ngay_chot) if ky.ngay_chot else None, "trang_thai": ky.trang_thai,
             "da_gui_email": bool(ky.da_gui_email)} if ky else
            {"thang": thang, "trang_thai": "CHUA_TAO"})
    return {"ky": head, "tre_han": tre_han, "phieu": pl,
            "tong": {"thu_nhap": sum(p["tong_thu_nhap"] for p in pl),
                     "thuc_linh": sum(p["thuc_linh"] for p in pl),
                     "bh_nv": sum(p["bhxh"] + p["bhyt"] + p["bhtn"] for p in pl),
                     "bh_dn": sum(p["bhxh_dn"] + p["bhyt_dn"] + p["bhtn_dn"] for p in pl),
                     "thue_tncn": sum(p["thue_tncn"] for p in pl),
                     "chi_phi_dn": sum(p["chi_phi_dn"] for p in pl)}}


@router.get("/phieu-luong/{bl_id}")
def chi_tiet_phieu(bl_id: int, db: Session = Depends(get_db), _=Depends(yeu_cau(MODULE, "XEM"))):
    bl = db.get(BangLuong, bl_id)
    if bl is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Không tìm thấy phiếu lương")
    return _pl_dict(db, bl)


# ----- Chấm công / OT / tạm ứng cho 1 nhân viên -> tính lại -----
@router.put("/phieu-luong/{bl_id}")
def cham_cong_phieu(bl_id: int, data: ChamCongLuongVao, db: Session = Depends(get_db),
                    nd: NguoiDung = Depends(yeu_cau(MODULE, "THAO_TAC"))):
    bl = db.get(BangLuong, bl_id)
    if bl is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Không tìm thấy phiếu lương")
    if bl.trang_thai == "DA_DUYET":
        raise HTTPException(status.HTTP_400_BAD_REQUEST,
                            "Kỳ lương đã chốt — hãy bấm “Mở lại kỳ” (CEO/Admin) để điều chỉnh, rồi chốt lại.")
    for k in ("cong_thuc_te", "gio_ot_thuong", "gio_ot_cuoi_tuan", "gio_ot_le", "tam_ung",
              "phu_cap_khac", "ngay_nghi_kpep", "so_phut_di_tre", "khau_tru_khac"):
        v = getattr(data, k)
        if v is not None:
            setattr(bl, k, v)
    nv = db.get(NhanVien, bl.nhan_vien_id)
    _ap_dung_tinh(db, nv, bl)
    ky = db.get(KyLuong, bl.thang)
    if ky:
        _cap_nhat_tong_ky(db, ky)
    db.commit()
    return _pl_dict(db, bl, nv)


# ----- CHỐT lương: tính lại + hạch toán chi phí + khóa kỳ -----
@router.post("/ky-luong/{thang}/chot")
def chot_ky_luong(thang: str, db: Session = Depends(get_db),
                  nd: NguoiDung = Depends(yeu_cau(MODULE, "XEM")),
                  __: NguoiDung = Depends(chi_vai_tro("KTT", "CEO"))):
    ky = db.get(KyLuong, thang)
    if ky is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Chưa tạo kỳ lương này")
    if ky.trang_thai == "DA_CHOT":
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Kỳ lương đã chốt rồi.")
    rows = db.query(BangLuong).filter_by(thang=thang).all()
    if not rows:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Kỳ chưa có bảng lương.")
    me = nhan_vien_id_cua(db, nd.id)
    so_bt = 0
    for bl in rows:
        nv = db.get(NhanVien, bl.nhan_vien_id)
        _ap_dung_tinh(db, nv, bl)
        so_bt += len(hach_toan_luong(db, bl))
        bl.trang_thai = "DA_DUYET"
        bl.nguoi_duyet_ktt = me
        bl.nguoi_ky_ceo = me
    ky.trang_thai = "DA_CHOT"
    ky.nguoi_chot = me
    _cap_nhat_tong_ky(db, ky)
    ghi_audit(db, nd.id, "DUYET", "ky_luong", None,
              moi={"thang": thang, "so_bang": len(rows), "so_but_toan": so_bt})
    db.commit()
    return {"thang": thang, "so_bang": len(rows), "so_but_toan_luong": so_bt,
            "tong_chi_phi_dn": _f(ky.tong_chi_phi_dn), "trang_thai": "DA_CHOT"}


def _bt_luong_cua(db, bl_ids: list[int]):
    """Bút toán lương đã hạch toán của các phiếu (nguon='LUONG', nguon_id=bang_luong.id)."""
    if not bl_ids:
        return []
    return db.query(ButToan).filter(ButToan.nguon == "LUONG",
                                    ButToan.nguon_id.in_(bl_ids)).all()


# ----- MỞ LẠI kỳ đã chốt để điều chỉnh (gỡ bút toán lương) — CEO/ADMIN -----
@router.post("/ky-luong/{thang}/mo-lai")
def mo_lai_ky_luong(thang: str, db: Session = Depends(get_db),
                    nd: NguoiDung = Depends(chi_vai_tro("CEO", "ADMIN"))):
    """Mở khóa kỳ lương đã chốt: GỠ toàn bộ bút toán lương của kỳ rồi đưa kỳ về NHAP
    để sửa/xóa. Chốt lại sẽ hạch toán mới hoàn toàn."""
    ky = db.get(KyLuong, thang)
    if ky is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Chưa tạo kỳ lương này")
    if ky.trang_thai != "DA_CHOT":
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Kỳ lương chưa chốt — không cần mở lại.")
    rows = db.query(BangLuong).filter_by(thang=thang).all()
    bts = _bt_luong_cua(db, [b.id for b in rows])
    for bt in bts:
        db.delete(bt)
    for bl in rows:
        bl.trang_thai = "CHO_DUYET"
        bl.nguoi_duyet_ktt = None
        bl.nguoi_ky_ceo = None
    ky.trang_thai = "NHAP"
    ky.nguoi_chot = None
    db.flush()
    _cap_nhat_tong_ky(db, ky)
    ghi_audit(db, nd.id, "SUA", "ky_luong", None,
              moi={"thang": thang, "mo_lai_ky": True, "but_toan_da_go": len(bts),
                   "so_bang": len(rows)})
    db.commit()
    return {"thang": thang, "trang_thai": "NHAP", "but_toan_da_go": len(bts),
            "so_bang": len(rows)}


# ----- XÓA 1 phiếu lương trong kỳ — CEO/ADMIN, giữ vết kế toán -----
@router.delete("/phieu-luong/{bl_id}")
def xoa_phieu_luong(bl_id: int, db: Session = Depends(get_db),
                    nd: NguoiDung = Depends(chi_vai_tro("CEO", "ADMIN"))):
    bl = db.get(BangLuong, bl_id)
    if bl is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Không tìm thấy phiếu lương")
    ky = db.get(KyLuong, bl.thang)
    if (ky is not None and ky.trang_thai == "DA_CHOT") or bl.trang_thai == "DA_DUYET":
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            f"Kỳ lương {bl.thang} đã chốt và đã hạch toán — hãy bấm “Mở lại kỳ” trước khi xóa phiếu.")
    con_bt = len(_bt_luong_cua(db, [bl.id]))
    if con_bt:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            f"Phiếu lương này còn {con_bt} bút toán lương đã ghi sổ — hãy “Mở lại kỳ” để gỡ bút toán trước.")
    nv = db.get(NhanVien, bl.nhan_vien_id)
    ghi_audit(db, nd.id, "XOA", "bang_luong", bl_id,
              cu={"thang": bl.thang, "nhan_vien": (nv.ho_ten if nv else bl.nhan_vien_id),
                  "thuc_linh": _f(bl.thuc_linh)})
    db.delete(bl)
    db.flush()
    if ky is not None:
        _cap_nhat_tong_ky(db, ky)
    db.commit()
    return {"da_xoa": True}


# ----- XÓA CẢ KỲ lương (mọi phiếu trong tháng) — CEO/ADMIN -----
@router.delete("/ky-luong/{thang}")
def xoa_ky_luong(thang: str, db: Session = Depends(get_db),
                 nd: NguoiDung = Depends(chi_vai_tro("CEO", "ADMIN"))):
    ky = db.get(KyLuong, thang)
    rows = db.query(BangLuong).filter_by(thang=thang).all()
    if ky is None and not rows:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Chưa tạo kỳ lương này")
    if ky is not None and ky.trang_thai == "DA_CHOT":
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            "Kỳ lương đã chốt và đã hạch toán — hãy bấm “Mở lại kỳ” trước khi xóa.")
    con_bt = len(_bt_luong_cua(db, [b.id for b in rows]))
    if con_bt:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            f"Kỳ còn {con_bt} bút toán lương đã ghi sổ — hãy “Mở lại kỳ” để gỡ bút toán trước.")
    for bl in rows:
        db.delete(bl)
    if ky is not None:
        db.delete(ky)
    ghi_audit(db, nd.id, "XOA", "ky_luong", None,
              cu={"thang": thang, "so_bang": len(rows)})
    db.commit()
    return {"da_xoa": True, "so_bang": len(rows)}


# ----- GỬI EMAIL phiếu lương từng người (chậm nhất ngày 7) -----
def _email_phieu_luong_html(nv, p):
    def vn(x):
        return f"{x:,.0f}đ"
    return f"""<div style="font-family:Arial,sans-serif;max-width:600px;margin:auto;color:#1f2937">
<div style="background:#0e7490;color:#fff;padding:14px 18px;border-radius:8px 8px 0 0">
  <div style="font-size:16px;font-weight:bold">{settings.cong_ty_ten}</div>
  <div style="font-size:12px;opacity:.9">{settings.cong_ty_dia_chi}</div></div>
<div style="border:1px solid #e5e7eb;border-top:0;padding:18px;border-radius:0 0 8px 8px">
  <h2 style="margin:0 0 4px">PHIẾU LƯƠNG THÁNG {p['thang']}</h2>
  <div style="color:#6b7280;font-size:13px;margin-bottom:12px">Nhân viên: <b>{nv.ho_ten}</b>{' · ' + (nv.chuc_danh or '') if nv.chuc_danh else ''} · Mã NV: {nv.ma or nv.id}</div>
  <table style="width:100%;border-collapse:collapse;font-size:14px">
    <tr><td style="padding:5px 0">Công chuẩn / thực tế</td><td style="text-align:right">{p['cong_chuan']:.0f} / {p['cong_thuc_te']:.0f}</td></tr>
    <tr><td>Lương theo công</td><td style="text-align:right">{vn(p['luong_thuc_te'] + p.get('khau_tru_nghi',0) + p.get('khau_tru_tre',0))}</td></tr>
    <tr><td>Tăng ca (thường {p['gio_ot_thuong']:.0f}h · cuối tuần {p['gio_ot_cuoi_tuan']:.0f}h · lễ {p['gio_ot_le']:.0f}h)</td><td style="text-align:right">{vn(p['ot'])}</td></tr>
    <tr><td>Phụ cấp{(' (gồm phát sinh ' + vn(p['phu_cap_khac']) + ')') if p.get('phu_cap_khac') else ''}</td><td style="text-align:right">{vn(p['phu_cap'])}</td></tr>
    {f'<tr><td style="color:#b91c1c">Trừ nghỉ không phép ({p["ngay_nghi_kpep"]:.1f} ngày)</td><td style="text-align:right;color:#b91c1c">-{vn(p["khau_tru_nghi"])}</td></tr>' if p.get('khau_tru_nghi') else ''}
    {f'<tr><td style="color:#b91c1c">Trừ đi trễ ({p["so_phut_di_tre"]} phút)</td><td style="text-align:right;color:#b91c1c">-{vn(p["khau_tru_tre"])}</td></tr>' if p.get('khau_tru_tre') else ''}
    <tr style="border-top:1px solid #e5e7eb;font-weight:bold"><td style="padding-top:6px">TỔNG THU NHẬP</td><td style="text-align:right;padding-top:6px">{vn(p['tong_thu_nhap'])}</td></tr>
    <tr><td style="color:#b91c1c">BHXH (8%)</td><td style="text-align:right;color:#b91c1c">-{vn(p['bhxh'])}</td></tr>
    <tr><td style="color:#b91c1c">BHYT (1,5%)</td><td style="text-align:right;color:#b91c1c">-{vn(p['bhyt'])}</td></tr>
    <tr><td style="color:#b91c1c">BHTN (1%)</td><td style="text-align:right;color:#b91c1c">-{vn(p['bhtn'])}</td></tr>
    <tr><td style="color:#b91c1c">Thuế TNCN</td><td style="text-align:right;color:#b91c1c">-{vn(p['thue_tncn'])}</td></tr>
    {f'<tr><td style="color:#b91c1c">Khấu trừ khác</td><td style="text-align:right;color:#b91c1c">-{vn(p["khau_tru_khac"])}</td></tr>' if p.get('khau_tru_khac') else ''}
    {f'<tr><td style="color:#b91c1c">Tạm ứng</td><td style="text-align:right;color:#b91c1c">-{vn(p["tam_ung"])}</td></tr>' if p['tam_ung'] else ''}
    <tr style="border-top:2px solid #0e7490;font-weight:bold;font-size:16px"><td style="padding-top:8px;color:#0e7490">THỰC LĨNH</td><td style="text-align:right;padding-top:8px;color:#0e7490">{vn(p['thuc_linh'])}</td></tr>
  </table>
  {f'<div style="color:#6b7280;font-size:12px;margin-top:10px">Chuyển khoản: {nv.so_tai_khoan} - {nv.ngan_hang or ""}</div>' if nv.so_tai_khoan else ''}
  <div style="color:#9ca3af;font-size:11px;margin-top:14px;border-top:1px solid #f3f4f6;padding-top:8px">
    Phiếu lương tự động từ hệ thống ERP SVWS. Mọi thắc mắc vui lòng liên hệ Phòng Hành chính - Nhân sự.</div>
</div></div>"""


@router.post("/ky-luong/{thang}/gui-email")
def gui_email_luong(thang: str, db: Session = Depends(get_db),
                    nd: NguoiDung = Depends(yeu_cau(MODULE, "THAO_TAC"))):
    ky = db.get(KyLuong, thang)
    if ky is None or ky.trang_thai != "DA_CHOT":
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Cần CHỐT kỳ lương trước khi gửi email.")
    rows = db.query(BangLuong).filter_by(thang=thang).all()
    provider = lay_email_provider()
    da_gui, bo_qua = 0, []
    for bl in rows:
        nv = db.get(NhanVien, bl.nhan_vien_id)
        to = nv.email
        if not to and nv.nguoi_dung_id:
            u = db.get(NguoiDung, nv.nguoi_dung_id)
            to = u.email if u else None
        if not to:
            bo_qua.append(nv.ho_ten); continue
        p = _pl_dict(db, bl, nv)
        provider.gui(to, f"[SVWS] Phiếu lương tháng {thang} - {nv.ho_ten}",
                     _email_phieu_luong_html(nv, p), gui_tu=settings.cong_ty_email)
        bl.email_sent = True
        bl.ngay_gui_email = datetime.now()
        da_gui += 1
    ky.da_gui_email = True
    tre_han = bool(ky.ngay_chot and date.today() > ky.ngay_chot)
    ghi_audit(db, nd.id, "GUI_EMAIL", "ky_luong", None, moi={"thang": thang, "da_gui": da_gui})
    db.commit()
    return {"thang": thang, "da_gui": da_gui, "bo_qua_thieu_email": bo_qua,
            "tre_han": tre_han, "han_chot": str(ky.ngay_chot) if ky.ngay_chot else None}
